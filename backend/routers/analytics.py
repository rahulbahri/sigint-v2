"""
routers/analytics.py — Fingerprint, monthly KPIs, summary, bridge, NLP query,
                        smart actions, board deck, weekly briefing, and export endpoints.
"""
import io
import json
import re
import numpy as np
import pandas as pd
from datetime import datetime
from typing import Optional, List

from fastapi import APIRouter, File, HTTPException, Request, UploadFile
from fastapi.responses import StreamingResponse, HTMLResponse
from pydantic import BaseModel

from core.database import get_db
from core.deps import _get_workspace, _require_workspace
from core.kpi_defs import (
    KPI_DEFS, CAUSATION_RULES, ALL_CAUSATION_RULES, BENCHMARKS,
    EXTENDED_ONTOLOGY_METRICS, ONTOLOGY_DOMAIN, compute_gap_status,
)
from core.criticality import DOMAIN_URGENCY, DEFAULT_WEIGHTS as CRIT_DEFAULT_WEIGHTS

# ── Board deck export moved to routers/board_pack.py ─────────────────────────

router = APIRouter()

@router.get("/api/monthly", tags=["KPIs"])
def monthly_kpis(request: Request, year: Optional[int] = None):
    """Return computed monthly KPI values. Optionally filter by year."""
    workspace_id = _require_workspace(request)
    conn = get_db()
    query = "SELECT * FROM monthly_data WHERE workspace_id=?"
    params: list = [workspace_id]
    if year:
        query += " AND year = ?"
        params.append(year)
    rows = conn.execute(query, params).fetchall()
    conn.close()
    result = []
    for row in rows:
        entry = {"year": row["year"], "month": row["month"], "kpis": json.loads(row["data_json"])}
        result.append(entry)
    return sorted(result, key=lambda x: (x["year"], x["month"]))

@router.get("/api/fingerprint", tags=["Analytics"])
def fingerprint(request: Request, year: Optional[int] = None):
    """
    Returns the 12-month KPI fingerprint for the organisation.
    Each KPI shows its monthly values, target, trend direction, and status (green/yellow/red).
    """
    workspace_id = _require_workspace(request)
    conn = get_db()
    if year:
        query = "SELECT * FROM monthly_data WHERE workspace_id=? AND year=?"
        rows = conn.execute(query, [workspace_id, year]).fetchall()
    else:
        rows = conn.execute("SELECT * FROM monthly_data WHERE workspace_id=?", [workspace_id]).fetchall()
    targets = {r["kpi_key"]: {"target": r["target_value"], "direction": r["direction"], "unit": r["unit"]}
               for r in conn.execute("SELECT * FROM kpi_targets WHERE workspace_id=?", [workspace_id]).fetchall()}
    conn.close()

    # Organise by KPI
    kpi_monthly: dict = {}
    for row in rows:
        mo_key = f"{row['year']}-{row['month']:02d}"
        data   = json.loads(row["data_json"])
        for kpi_key, val in data.items():
            if kpi_key in ("year", "month") or kpi_key.startswith("_"):
                continue
            kpi_monthly.setdefault(kpi_key, {})[mo_key] = val

    fingerprint_out = []
    for kdef in KPI_DEFS:
        key  = kdef["key"]
        vals = kpi_monthly.get(key, {})
        t    = targets.get(key, {})
        tval = t.get("target")
        dirn = t.get("direction", "higher")
        unit = t.get("unit", kdef["unit"])

        from core.kpi_utils import compute_kpi_avg as _cka
        monthly_list = [{"period": k, "value": v} for k, v in sorted(vals.items())]
        values       = [m["value"] for m in monthly_list]
        avg          = _cka(values, window=6, period_filtered=False)

        def status(val, target, direction):
            if val is None or target is None: return "grey"
            pct = val / target if target else 0
            if direction == "higher":
                return "green" if pct >= 0.98 else ("yellow" if pct >= 0.90 else "red")
            else:
                return "green" if pct <= 1.02 else ("yellow" if pct <= 1.10 else "red")

        trend = None
        if len(values) >= 2:
            trend = "up" if values[-1] > values[0] else ("down" if values[-1] < values[0] else "flat")

        fingerprint_out.append({
            "key":           key,
            "name":          kdef["name"],
            "unit":          unit,
            "target":        tval,
            "direction":     dirn,
            "avg":           avg,
            "trend":         trend,
            "fy_status":     status(avg, tval, dirn),
            "monthly":       monthly_list,
            "causation":     CAUSATION_RULES.get(key, {
                                 "root_causes": [], "downstream_impact": [], "corrective_actions": []
                             }),
            "gaap_status":   kdef.get("gaap_status", "operating"),
        })

    return fingerprint_out

@router.get("/api/summary", tags=["Analytics"])
def summary(request: Request, year: Optional[int] = None):
    """High-level dashboard summary: upload count, KPI coverage, status breakdown."""
    workspace_id = _require_workspace(request)
    conn = get_db()
    uploads = conn.execute("SELECT COUNT(*) as c FROM uploads WHERE workspace_id=?", [workspace_id]).fetchone()["c"]
    # Filter by year when provided so status counts match the fingerprint tab
    if year:
        monthly_rows = conn.execute("SELECT * FROM monthly_data WHERE workspace_id=? AND year=?", [workspace_id, year]).fetchall()
    else:
        monthly_rows = conn.execute("SELECT * FROM monthly_data WHERE workspace_id=?", [workspace_id]).fetchall()
    all_rows_count = conn.execute("SELECT COUNT(*) as c FROM monthly_data WHERE workspace_id=?", [workspace_id]).fetchone()["c"]
    targets = {r["kpi_key"]: {"target": r["target_value"], "direction": r["direction"]}
               for r in conn.execute("SELECT * FROM kpi_targets WHERE workspace_id=?", [workspace_id]).fetchall()}
    conn.close()

    all_kpis: dict = {}
    for row in monthly_rows:
        for k, v in json.loads(row["data_json"]).items():
            if k.startswith("_") or k in ("year", "month"):
                continue
            all_kpis.setdefault(k, []).append(v)

    status_counts = {"green": 0, "yellow": 0, "red": 0, "grey": 0}
    for key, vals in all_kpis.items():
        # Filter out None/NaN values before computing average
        clean_vals = [v for v in vals if v is not None and not (isinstance(v, float) and (np.isnan(v) or np.isinf(v)))]
        if not clean_vals:
            status_counts["grey"] += 1
            continue
        avg  = round(float(np.mean(clean_vals)), 2)
        t    = targets.get(key, {})
        tval = t.get("target")
        dirn = t.get("direction", "higher")
        if tval is None:
            status_counts["grey"] += 1
            continue
        pct = avg / tval if tval else 0
        if dirn == "higher":
            s = "green" if pct >= 0.98 else ("yellow" if pct >= 0.90 else "red")
        else:
            s = "green" if pct <= 1.02 else ("yellow" if pct <= 1.10 else "red")
        status_counts[s] += 1

    # kpis_tracked = KPI keys that have data AND a target definition
    kpis_with_definition = [k for k in all_kpis if k in targets]
    return {
        "uploads":         uploads,
        "kpis_tracked":    len(kpis_with_definition),
        "kpis_available":  len(KPI_DEFS),
        "months_of_data":  all_rows_count,   # always total across all years
        "status_breakdown": status_counts,
    }

@router.get("/api/available-years", tags=["Analytics"])
def available_years(request: Request):
    """Return distinct years that have monthly KPI data."""
    workspace_id = _require_workspace(request)
    conn = get_db()
    rows = conn.execute("SELECT DISTINCT year FROM monthly_data WHERE workspace_id=? ORDER BY year", [workspace_id]).fetchall()
    conn.close()
    return [r["year"] for r in rows]

# ─── NLP Query Endpoint ─────────────────────────────────────────────────────

@router.post("/api/query", tags=["Analytics"])
async def query_kpi(request: Request, payload: dict):
    """
    Natural-language KPI query powered by Claude.
    Accepts { "question": "...", "years": [2024] } and returns { "answer": "...", "kpis_referenced": [...] }.
    Builds full context from the live DB fingerprint on every call, filtered to requested years.
    """
    workspace_id = _require_workspace(request)
    question = payload.get("question", "").strip()
    if not question:
        raise HTTPException(status_code=400, detail="question is required")
    years_filter = payload.get("years", None)  # list of ints or None = all years

    # ── Build context from DB (replicate fingerprint + summary logic inline) ──
    conn = get_db()
    if years_filter:
        placeholders = ",".join("?" * len(years_filter))
        rows = conn.execute(
            f"SELECT * FROM monthly_data WHERE workspace_id=? AND year IN ({placeholders})",
            [workspace_id] + list(years_filter)
        ).fetchall()
    else:
        rows = conn.execute("SELECT * FROM monthly_data WHERE workspace_id=?", [workspace_id]).fetchall()
    targets  = {r["kpi_key"]: {"target": r["target_value"], "direction": r["direction"], "unit": r["unit"]}
                for r in conn.execute("SELECT * FROM kpi_targets WHERE workspace_id=?", [workspace_id]).fetchall()}
    uploads  = conn.execute("SELECT COUNT(*) as c FROM uploads WHERE workspace_id=?", [workspace_id]).fetchone()["c"]
    conn.close()

    # Organise monthly values by KPI key
    kpi_monthly: dict = {}
    for row in rows:
        mo_key = f"{row['year']}-{row['month']:02d}"
        for kpi_key, val in json.loads(row["data_json"]).items():
            if kpi_key.startswith("_") or kpi_key in ("year", "month"):
                continue
            kpi_monthly.setdefault(kpi_key, {})[mo_key] = val

    def _status(val, target, direction):
        if val is None or target is None:
            return "grey"
        pct = val / target if target else 0
        if direction == "higher":
            return "green" if pct >= 0.98 else ("yellow" if pct >= 0.90 else "red")
        return "green" if pct <= 1.02 else ("yellow" if pct <= 1.10 else "red")

    kpi_lines   = []
    status_counts = {"green": 0, "yellow": 0, "red": 0, "grey": 0}

    for kdef in KPI_DEFS:
        key  = kdef["key"]
        vals = kpi_monthly.get(key, {})
        t    = targets.get(key, {})
        tval = t.get("target")
        dirn = t.get("direction", "higher")

        from core.kpi_utils import compute_kpi_avg as _cka
        monthly_sorted = sorted(vals.items())
        values         = [v for _, v in monthly_sorted if v is not None]
        avg            = _cka(values, window=6, period_filtered=False)
        status         = _status(avg, tval, dirn)
        status_counts[status] += 1

        trend = None
        if len(values) >= 2:
            trend = "up" if values[-1] > values[0] else ("down" if values[-1] < values[0] else "flat")

        monthly_str = ", ".join(f"{p}: {v}" for p, v in monthly_sorted) or "no data"

        kpi_lines.append(
            f"- {kdef['name']} (key:{key}, unit:{kdef['unit']}): "
            f"avg={avg}, target={tval}, direction={dirn}, status={status}, trend={trend}\n"
            f"  monthly → {monthly_str}"
        )

    months_of_data = len(rows)
    kpis_tracked   = len([k for k in KPI_DEFS if kpi_monthly.get(k["key"])])

    # ── Projection context (if available) ─────────────────────────────────────
    proj_context_lines = []
    try:
        proj_conn  = get_db()
        proj_rows  = proj_conn.execute("SELECT * FROM projection_monthly_data WHERE workspace_id=?", [workspace_id]).fetchall()
        proj_conn.close()
        if proj_rows:
            proj_by_period: dict = {}
            for pr in proj_rows:
                proj_by_period[(pr["year"], pr["month"])] = json.loads(pr["data_json"])

            actual_by_period2: dict = {}
            for row in rows:
                k2 = (row["year"], row["month"])
                actual_by_period2.setdefault(k2, {})
                actual_by_period2[k2].update(json.loads(row["data_json"]))

            overlap2 = sorted(set(proj_by_period.keys()) & set(actual_by_period2.keys()))
            if overlap2:
                for kdef in KPI_DEFS:
                    key2      = kdef["key"]
                    direction2 = kdef["direction"]
                    gap_pcts2 = []
                    actuals2  = []
                    projs2    = []
                    for (yr2, mo2) in overlap2:
                        pv = proj_by_period[(yr2, mo2)].get(key2)
                        av = actual_by_period2[(yr2, mo2)].get(key2)
                        if pv and av and pv != 0:
                            actuals2.append(av)
                            projs2.append(pv)
                            if direction2 == "higher":
                                gap_pcts2.append((av - pv) / abs(pv) * 100)
                            else:
                                gap_pcts2.append((pv - av) / abs(pv) * 100)
                    if actuals2:
                        avg_a2 = round(float(np.mean(actuals2)), 2)
                        avg_p2 = round(float(np.mean(projs2)), 2)
                        avg_g2 = round(float(np.mean(gap_pcts2)), 2)
                        status2 = compute_gap_status(avg_g2)
                        proj_context_lines.append(
                            f"- {kdef['name']}: actual avg={avg_a2}, projected avg={avg_p2}, gap={avg_g2:+.1f}% [{status2}]"
                        )
    except Exception:
        pass

    proj_section = ""
    if proj_context_lines:
        proj_section = f"""
PROJECTION vs ACTUAL CONTEXT ({len(proj_context_lines)} KPIs compared):
{chr(10).join(proj_context_lines)}
"""

    # Build a human-readable period description for the prompt header
    if years_filter:
        yr_list = sorted(years_filter)
        period_desc = f"FY {yr_list[0]}" if len(yr_list) == 1 else f"FY {yr_list[0]}–{yr_list[-1]}"
    else:
        all_years = sorted({row["year"] for row in rows}) if rows else []
        period_desc = (f"FY {all_years[0]}–{all_years[-1]}" if len(all_years) > 1
                       else f"FY {all_years[0]}" if all_years else "All Available Data")

    system_prompt = f"""You are an expert financial analyst embedded in the Axiom KPI Intelligence Dashboard.
You have access to the following organisational performance data for {period_desc}:

SUMMARY: {months_of_data} months of data | {kpis_tracked}/{len(KPI_DEFS)} KPIs tracked | Period: {period_desc}
Status breakdown: {status_counts.get('green', 0)} on target, {status_counts.get('yellow', 0)} needs attention, {status_counts.get('red', 0)} critical

KPI DATA:
{chr(10).join(kpi_lines)}
{proj_section}
Rules:
- Answer concisely (2-4 sentences max, or a short bullet list)
- Always cite specific numbers and months when relevant
- Flag critical KPIs (status=red) clearly
- When projection data is available, reference the gap percentages and status in your analysis
- Do NOT make up data beyond what is provided above
- Respond in plain text — no markdown headers, no asterisks, no bold"""

    try:
        client = anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY"))
        msg = client.messages.create(
            model="claude-3-5-haiku-20241022",
            max_tokens=400,
            system=system_prompt,
            messages=[{"role": "user", "content": question}],
        )
        answer = msg.content[0].text
        kpis_referenced = [k["key"] for k in KPI_DEFS if k["name"].lower() in answer.lower()]
        return {"answer": answer, "kpis_referenced": kpis_referenced}
    except Exception as e:
        return {"answer": f"Query unavailable: {str(e)}", "kpis_referenced": []}

def _compute_fingerprint_data(targets_override=None, workspace_id: str = ""):
    """Reuse fingerprint logic without HTTP call."""
    conn = get_db()
    rows = conn.execute("SELECT * FROM monthly_data WHERE workspace_id=?", [workspace_id]).fetchall()
    if targets_override is not None:
        targets = targets_override
    else:
        targets = {r["kpi_key"]: {"target": r["target_value"], "direction": r["direction"], "unit": r["unit"]}
                   for r in conn.execute("SELECT * FROM kpi_targets WHERE workspace_id=?", [workspace_id]).fetchall()}
    conn.close()

    kpi_monthly: dict = {}
    for row in rows:
        mo_key = f"{row['year']}-{row['month']:02d}"
        data = json.loads(row["data_json"])
        for kpi_key, val in data.items():
            if kpi_key in ("year", "month") or kpi_key.startswith("_"):
                continue
            kpi_monthly.setdefault(kpi_key, {})[mo_key] = val

    def _status(val, target, direction):
        if val is None or target is None:
            return "grey"
        pct = val / target if target else 0
        if direction == "higher":
            return "green" if pct >= 0.98 else ("yellow" if pct >= 0.90 else "red")
        else:
            return "green" if pct <= 1.02 else ("yellow" if pct <= 1.10 else "red")

    # Build lookup for all known KPI metadata (KPI_DEFS + EXTENDED_ONTOLOGY_METRICS)
    all_kpi_meta = {kd["key"]: kd for kd in KPI_DEFS}
    for em in EXTENDED_ONTOLOGY_METRICS:
        if em["key"] not in all_kpi_meta:
            all_kpi_meta[em["key"]] = em

    # All KPIs that have both a target AND monthly data
    keys_to_include = set(KPI_DEFS[i]["key"] for i in range(len(KPI_DEFS)))
    for key in kpi_monthly:
        if key in targets:
            keys_to_include.add(key)

    fingerprint_out = []
    for key in keys_to_include:
        kdef = all_kpi_meta.get(key, {"key": key, "name": key.replace("_", " ").title(), "unit": "ratio", "direction": "higher", "domain": "other"})
        vals = kpi_monthly.get(key, {})
        t = targets.get(key, {})
        tval = t.get("target")
        dirn = t.get("direction", kdef.get("direction", "higher"))
        unit = t.get("unit", kdef.get("unit", "ratio"))

        from core.kpi_utils import compute_kpi_avg as _cka
        monthly_list = [{"period": k, "value": v} for k, v in sorted(vals.items())]
        values = [m["value"] for m in monthly_list]
        avg = _cka(values, window=6, period_filtered=False)

        trend = None
        if len(values) >= 2:
            trend = "up" if values[-1] > values[0] else ("down" if values[-1] < values[0] else "flat")

        fingerprint_out.append({
            "key": key,
            "name": kdef.get("name", key.replace("_", " ").title()),
            "unit": unit,
            "target": tval,
            "direction": dirn,
            "avg": avg,
            "trend": trend,
            "fy_status": _status(avg, tval, dirn),
            "monthly": monthly_list,
        })

    # Sort: KPI_DEFS order first, then extended KPIs alphabetically
    kpi_def_order = {kd["key"]: i for i, kd in enumerate(KPI_DEFS)}
    fingerprint_out.sort(key=lambda x: (kpi_def_order.get(x["key"], 9999), x["key"]))
    return fingerprint_out

@router.get("/api/export/board-deck.pptx", tags=["Board Deck"])
def export_board_deck(request: Request, stage: str = "series_b"):
    """DEPRECATED: Use POST /api/board-pack/generate instead.
    Redirects to the new board pack endpoint with default settings.
    """
    from fastapi.responses import JSONResponse
    return JSONResponse(
        status_code=410,
        content={
            "detail": "This endpoint has been replaced. Use POST /api/board-pack/generate with date range and mode selection.",
            "new_endpoint": "/api/board-pack/generate",
            "method": "POST",
        },
    )


def _export_board_deck_REMOVED():
    """Old board-deck export removed. See routers/board_pack.py."""
    raise NotImplementedError("Moved to routers/board_pack.py")


# NOTE: ~500 lines of old board-deck PPTX generation code was removed here.
# The functionality has been rewritten with intelligent chart selection,
# knowledge-graph-based causal narratives, and flexible date ranges in:
#   - core/chart_engine.py (chart rendering)
#   - core/board_narrative.py (narrative generation)
#   - routers/board_pack.py (slide assembly)


_valid_stages_UNUSED = {"seed", "series_a", "series_b", "series_c"}
# Old board-deck code (bench = {} through StreamingResponse) removed.
# See git history for the original ~500-line PPTX generator.


# ─── KPI Targets ─────────────────────────────────────────────────────────────

@router.put("/api/targets/{kpi_key}", tags=["Configuration"])
async def update_target(request: Request, kpi_key: str):
    """Upsert the target value (and optionally unit/direction) for a KPI."""
    import re as _re
    workspace_id = _require_workspace(request)
    if not workspace_id:
        raise HTTPException(status_code=401, detail="Not authenticated")
    if not _re.match(r'^[a-z_]+$', kpi_key):
        raise HTTPException(status_code=400, detail="Invalid KPI key format")

    body = await request.json()
    target_value = float(body.get("target", body.get("target_value", 0)))
    unit         = body.get("unit") or next(
        (k["unit"]      for k in KPI_DEFS if k["key"] == kpi_key), "pct"
    )
    direction    = body.get("direction") or next(
        (k["direction"] for k in KPI_DEFS if k["key"] == kpi_key), "higher"
    )

    conn = get_db()
    try:
        conn.execute(
            """INSERT INTO kpi_targets (kpi_key, target_value, unit, direction, workspace_id)
               VALUES (?,?,?,?,?)
               ON CONFLICT(kpi_key, workspace_id) DO UPDATE
               SET target_value=excluded.target_value,
                   unit=excluded.unit,
                   direction=excluded.direction""",
            (kpi_key, target_value, unit, direction, workspace_id)
        )
        conn.commit()
    finally:
        conn.close()

    from core.database import _audit
    _audit("target_changed", "kpi_target", kpi_key,
           f"Target for {kpi_key} set to {target_value}{unit} ({direction} is better)",
           workspace_id=workspace_id)

    return {"kpi_key": kpi_key, "target_value": target_value, "unit": unit, "direction": direction}


# ─── Weekly Briefing HTML ────────────────────────────────────────────────────

@router.get("/api/export/weekly-briefing.html", tags=["Export"], response_class=HTMLResponse)
def weekly_briefing(request: Request, stage: Optional[str] = "series_b"):
    """
    Generate and return a self-contained HTML weekly briefing document.
    Opens directly in the browser — no download needed.
    """
    workspace_id = _get_workspace(request)
    if not workspace_id:
        raise HTTPException(status_code=401, detail="Unauthorised")
    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT year, month, data_json FROM monthly_data WHERE workspace_id=? ORDER BY year DESC, month DESC LIMIT 3",
            [workspace_id]
        ).fetchall()
        targets_rows = conn.execute(
            "SELECT kpi_key, target_value, direction, unit FROM kpi_targets WHERE workspace_id=?",
            [workspace_id]
        ).fetchall()
        settings_row = conn.execute(
            "SELECT value FROM company_settings WHERE key='company_name' AND workspace_id=?",
            [workspace_id]
        ).fetchone() if True else None
        # Fetch active decisions for briefing section
        try:
            decision_rows = conn.execute(
                "SELECT title, decided_by, decided_at FROM decisions "
                "WHERE workspace_id=? AND status='active' ORDER BY decided_at DESC LIMIT 5",
                [workspace_id],
            ).fetchall()
        except Exception:
            decision_rows = []
    finally:
        conn.close()

    company_name = (settings_row["value"] if settings_row else None) or workspace_id or "Your Company"
    targets = {r["kpi_key"]: {"target": r["target_value"], "direction": r["direction"] or "higher", "unit": r["unit"] or ""}
               for r in targets_rows}

    # Build KPI aggregates from last 3 months
    kpi_vals: dict = {}
    for row in rows:
        for k, v in json.loads(row["data_json"]).items():
            if k.startswith("_") or v is None:
                continue
            try:
                kpi_vals.setdefault(k, []).append(float(v))
            except (ValueError, TypeError):
                pass
    kpi_avgs = {k: round(sum(v)/len(v), 2) for k, v in kpi_vals.items() if v}

    # Classify KPIs
    red, yellow, green = [], [], []
    for kdef in KPI_DEFS:
        key = kdef["key"]
        avg = kpi_avgs.get(key)
        t   = targets.get(key, {})
        tval, dirn, unit = t.get("target"), t.get("direction","higher"), t.get("unit","")
        if avg is None or tval is None:
            continue
        pct = (avg/tval) if dirn=="higher" else (tval/avg if avg else 0)
        gap_sign = "+" if (dirn=="higher" and avg>tval) or (dirn=="lower" and avg<tval) else "-"
        gap_abs = abs(avg-tval)
        entry = {"name": kdef["name"], "key": key, "avg": avg, "target": tval,
                 "unit": unit, "dirn": dirn, "pct": pct, "gap_sign": gap_sign, "gap_abs": gap_abs}
        if pct >= 0.98:   green.append(entry)
        elif pct >= 0.90: yellow.append(entry)
        else:             red.append(entry)

    red.sort(key=lambda x: x["pct"])
    yellow.sort(key=lambda x: x["pct"])
    green.sort(key=lambda x: -x["pct"])

    now_str = datetime.utcnow().strftime("%B %d, %Y")
    score_color = "#DC2626" if len(red) > len(green) else ("#D97706" if yellow else "#059669")

    def fmt(entry):
        v, t, u = entry["avg"], entry["target"], entry["unit"]
        if u in ("pct","%"): return f"{v:.1f}% vs {t:.1f}%"
        if u in ("days","day"): return f"{v:.1f}d vs {t:.1f}d"
        if u == "ratio": return f"{v:.2f}x vs {t:.2f}x"
        if u in ("months","mo"): return f"{v:.1f}mo vs {t:.1f}mo"
        return f"{v:.2f} vs {t:.2f}"

    def rows_html(items, color):
        if not items: return f'<tr><td colspan="3" style="color:#94a3b8;font-style:italic;padding:12px 0">None</td></tr>'
        out = ""
        for e in items[:8]:
            gap_pct = abs(e["pct"]-1)*100
            direction = "above" if e["pct"]>=1 else "below"
            out += f"""
            <tr style="border-bottom:1px solid #f1f5f9">
              <td style="padding:10px 0;font-weight:600;color:#1e293b">{e['name']}</td>
              <td style="padding:10px 8px;color:#64748b;font-size:13px">{fmt(e)}</td>
              <td style="padding:10px 0;text-align:right">
                <span style="background:{color}15;color:{color};padding:2px 8px;border-radius:99px;font-size:12px;font-weight:700">
                  {gap_pct:.1f}% {direction} target
                </span>
              </td>
            </tr>"""
        return out

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>{company_name} · Weekly KPI Briefing · {now_str}</title>
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;background:#f8fafc;color:#1e293b;line-height:1.5}}
  .wrap{{max-width:860px;margin:0 auto;padding:40px 24px}}
  .header{{background:linear-gradient(135deg,#0055A4 0%,#00AEEF 100%);border-radius:16px;padding:36px 40px;color:#fff;margin-bottom:32px}}
  .header h1{{font-size:28px;font-weight:800;letter-spacing:-0.5px;margin-bottom:4px}}
  .header .meta{{font-size:13px;opacity:0.75;margin-top:8px}}
  .stat-row{{display:grid;grid-template-columns:repeat(3,1fr);gap:16px;margin-bottom:32px}}
  .stat{{background:#fff;border-radius:12px;padding:20px;border:1px solid #e2e8f0;text-align:center;box-shadow:0 1px 3px rgba(0,0,0,.06)}}
  .stat .num{{font-size:36px;font-weight:800;margin-bottom:4px}}
  .stat .lbl{{font-size:12px;color:#94a3b8;text-transform:uppercase;letter-spacing:.05em;font-weight:600}}
  .section{{background:#fff;border-radius:12px;padding:24px;margin-bottom:20px;border:1px solid #e2e8f0;box-shadow:0 1px 3px rgba(0,0,0,.06)}}
  .section-title{{font-size:13px;font-weight:700;text-transform:uppercase;letter-spacing:.06em;margin-bottom:16px;display:flex;align-items:center;gap:8px}}
  .dot{{width:8px;height:8px;border-radius:50%;display:inline-block}}
  table{{width:100%;border-collapse:collapse}}
  .callout{{background:#fffbeb;border:1px solid #fde68a;border-radius:10px;padding:16px 20px;margin-bottom:20px;font-size:13px;color:#92400e;line-height:1.6}}
  .callout strong{{color:#78350f}}
  .footer{{text-align:center;font-size:12px;color:#94a3b8;margin-top:32px;padding-top:16px;border-top:1px solid #e2e8f0}}
  @media print{{body{{background:#fff}}.wrap{{padding:0 16px}}}}
</style>
</head>
<body>
<div class="wrap">
  <div class="header">
    <div style="font-size:12px;opacity:.7;margin-bottom:6px;text-transform:uppercase;letter-spacing:.1em">Weekly KPI Briefing</div>
    <h1>{company_name}</h1>
    <div class="meta">Generated {now_str} &nbsp;·&nbsp; Based on last 3 months of data &nbsp;·&nbsp; Powered by Axiom Intelligence</div>
  </div>

  <div class="stat-row">
    <div class="stat">
      <div class="num" style="color:#DC2626">{len(red)}</div>
      <div class="lbl">Require Action</div>
    </div>
    <div class="stat">
      <div class="num" style="color:#D97706">{len(yellow)}</div>
      <div class="lbl">Watch List</div>
    </div>
    <div class="stat">
      <div class="num" style="color:#059669">{len(green)}</div>
      <div class="lbl">On Target</div>
    </div>
  </div>

  {'<div class="callout"><strong>Executive Summary:</strong> ' + str(len(red)) + ' KPI' + ('s' if len(red)!=1 else '') + ' require immediate action this week. ' + ('The primary concerns are ' + ', '.join(e['name'] for e in red[:3]) + (' among others' if len(red) > 3 else '') + '.' if red else 'All tracked KPIs are performing at or above target. Maintain current trajectory.') + ' ' + str(len(green)) + ' KPI' + ('s are' if len(green)!=1 else ' is') + ' exceeding targets, providing a strong foundation.</div>' if True else ''}

  <div class="section">
    <div class="section-title"><span class="dot" style="background:#DC2626"></span> Requires Action ({len(red)} KPIs)</div>
    <table><tbody>{rows_html(red, "#DC2626")}</tbody></table>
  </div>

  <div class="section">
    <div class="section-title"><span class="dot" style="background:#D97706"></span> Watch List ({len(yellow)} KPIs)</div>
    <table><tbody>{rows_html(yellow, "#D97706")}</tbody></table>
  </div>

  <div class="section">
    <div class="section-title"><span class="dot" style="background:#059669"></span> On Target ({len(green)} KPIs)</div>
    <table><tbody>{rows_html(green, "#059669")}</tbody></table>
  </div>

  {"".join(f'''
  <div class="section">
    <div class="section-title"><span class="dot" style="background:#0055A4"></span> Active Decisions ({len(decision_rows)})</div>
    <table><tbody>
    {"".join(f'<tr style="border-bottom:1px solid #f1f5f9"><td style="padding:10px 0;font-weight:600;color:#1e293b">{dr["title"]}</td><td style="padding:10px 8px;color:#64748b;font-size:13px">{dr["decided_by"] or ""}</td><td style="padding:10px 0;text-align:right;font-size:12px;color:#94a3b8">{(dr["decided_at"] or "")[:10]}</td></tr>' for dr in decision_rows)}
    </tbody></table>
  </div>
  ''' if decision_rows else [])}

  {f'''<div class="section">
    <h2 style="color:#003087;margin:0 0 10px 0;font-size:16px">SEC-Grade Analytics</h2>
    <table style="width:100%;border-collapse:collapse;font-size:13px">
      <tr style="border-bottom:2px solid #e2e8f0">
        <th style="text-align:left;padding:8px 0;color:#64748b;font-weight:600">Metric</th>
        <th style="text-align:right;padding:8px 0;color:#64748b;font-weight:600">Value</th>
        <th style="text-align:right;padding:8px 0;color:#64748b;font-weight:600">Classification</th>
      </tr>
      {"".join(f'<tr style="border-bottom:1px solid #f1f5f9"><td style="padding:8px 0;color:#1e293b">{label}</td><td style="padding:8px 0;text-align:right;font-weight:600;color:#0f172a">{val}</td><td style="padding:8px 0;text-align:right;font-size:11px;color:#64748b">{cls}</td></tr>' for label, val, cls in [
        ("Rule of 40", f"{(float(latest.get('revenue_growth', 0) or 0) + float(latest.get('ebitda_margin', 0) or 0)):.1f}", "Non-GAAP"),
        ("Gross Margin", f"{float(latest.get('gross_margin', 0) or 0):.1f}%", "GAAP"),
        ("NRR", f"{float(latest.get('nrr', 0) or 0):.1f}%", "Non-GAAP"),
        ("Burn Multiple", f"{float(latest.get('burn_multiple', 0) or 0):.2f}x", "Non-GAAP"),
        ("Customer Concentration", f"{float(latest.get('customer_concentration', 0) or 0):.1f}%", "Operating"),
        ("Churn Rate", f"{float(latest.get('churn_rate', 0) or 0):.1f}%", "Operating"),
      ])}
    </table>
    <p style="font-size:10px;color:#94a3b8;margin-top:8px">GAAP/Non-GAAP classification per SEC Regulation G. Non-GAAP metrics are operational measures not derived from GAAP financial statements.</p>
  </div>''' if latest else ''}

  <div class="footer">
    This briefing is generated automatically by Axiom Intelligence from your live KPI data.
    Targets and thresholds are configured in your KPI Targets settings. &nbsp;·&nbsp;
    <a href="https://app.axiomsync.ai" style="color:#0055A4;text-decoration:none">Open platform →</a>
  </div>
</div>
</body>
</html>"""

    return HTMLResponse(content=html, media_type="text/html")


# ─── Plan vs Actual Bridge ───────────────────────────────────────────────────

@router.get("/api/bridge", tags=["Projection"])
def bridge_analysis(request: Request, year: Optional[int] = None):
    """
    Compare projected vs actual KPIs month-by-month.
    Returns gap analysis, status (green/yellow/red), and causation rules for each KPI.
    """
    workspace_id = _require_workspace(request)
    conn = get_db()
    try:
        if year:
            proj_rows   = conn.execute("SELECT * FROM projection_monthly_data WHERE workspace_id=? AND year=?", [workspace_id, year]).fetchall()
            actual_rows = conn.execute("SELECT * FROM monthly_data WHERE workspace_id=? AND year=?", [workspace_id, year]).fetchall()
        else:
            proj_rows   = conn.execute("SELECT * FROM projection_monthly_data WHERE workspace_id=?", [workspace_id]).fetchall()
            actual_rows = conn.execute("SELECT * FROM monthly_data WHERE workspace_id=?", [workspace_id]).fetchall()
        # Load targets for three-way comparison
        target_rows = conn.execute(
            "SELECT kpi_key, target_value, direction FROM kpi_targets WHERE workspace_id=?",
            [workspace_id],
        ).fetchall()
        targets_by_kpi = {r["kpi_key"]: r["target_value"] for r in target_rows if r["target_value"] is not None}
    finally:
        conn.close()

    if not proj_rows:
        return {"has_projection": False}

    # Build projection lookup: (year, month) -> kpi_dict
    proj_by_period: dict = {}
    for row in proj_rows:
        proj_by_period[(row["year"], row["month"])] = json.loads(row["data_json"])

    # Build actuals lookup: (year, month) -> kpi_dict (merge if multiple uploads)
    actual_by_period: dict = {}
    for row in actual_rows:
        key = (row["year"], row["month"])
        actual_by_period.setdefault(key, {})
        actual_by_period[key].update(json.loads(row["data_json"]))

    # Find overlapping periods
    overlap = sorted(set(proj_by_period.keys()) & set(actual_by_period.keys()))
    if not overlap:
        return {"has_projection": True, "has_overlap": False, "summary": {}, "kpis": {}}

    kpis_out: dict = {}
    for kdef in KPI_DEFS:
        key       = kdef["key"]
        direction = kdef["direction"]
        months_data: dict = {}

        for (yr, mo) in overlap:
            proj_val   = proj_by_period[(yr, mo)].get(key)
            actual_val = actual_by_period[(yr, mo)].get(key)
            if proj_val is None or actual_val is None:
                continue
            if proj_val == 0:
                continue

            if direction == "higher":
                gap_pct = (actual_val - proj_val) / abs(proj_val) * 100
            else:
                gap_pct = (proj_val - actual_val) / abs(proj_val) * 100

            period_key = f"{yr}-{mo:02d}"
            months_data[period_key] = {
                "actual":    round(float(actual_val), 2),
                "projected": round(float(proj_val), 2),
                "gap":       round(float(actual_val - proj_val), 2),
                "gap_pct":   round(float(gap_pct), 2),
            }

        if not months_data:
            continue

        actuals       = [v["actual"]    for v in months_data.values()]
        projecteds    = [v["projected"] for v in months_data.values()]
        avg_actual    = round(float(np.mean(actuals)), 2)
        avg_projected = round(float(np.mean(projecteds)), 2)
        avg_gap       = round(float(avg_actual - avg_projected), 2)
        if avg_projected != 0:
            if direction == "higher":
                avg_gap_pct = round((avg_actual - avg_projected) / abs(avg_projected) * 100, 2)
            else:
                avg_gap_pct = round((avg_projected - avg_actual) / abs(avg_projected) * 100, 2)
        else:
            avg_gap_pct = 0.0
        overall_status = compute_gap_status(avg_gap_pct)

        # Three-way: include target comparison (additive, backward-compatible)
        target_val = targets_by_kpi.get(key)
        actual_vs_target_pct = None
        projected_vs_target_pct = None
        if target_val is not None and target_val != 0:
            if direction == "higher":
                actual_vs_target_pct = round((avg_actual - target_val) / abs(target_val) * 100, 2)
                projected_vs_target_pct = round((avg_projected - target_val) / abs(target_val) * 100, 2)
            else:
                actual_vs_target_pct = round((target_val - avg_actual) / abs(target_val) * 100, 2)
                projected_vs_target_pct = round((target_val - avg_projected) / abs(target_val) * 100, 2)

        kpis_out[key] = {
            "name":           kdef["name"],
            "unit":           kdef["unit"],
            "direction":      direction,
            "avg_actual":     avg_actual,
            "avg_projected":  avg_projected,
            "avg_gap":        avg_gap,
            "avg_gap_pct":    avg_gap_pct,
            "target_value":   target_val,
            "actual_vs_target_pct":    actual_vs_target_pct,
            "projected_vs_target_pct": projected_vs_target_pct,
            "overall_status": overall_status,
            "months":         months_data,
            "causation":      CAUSATION_RULES.get(key, {
                "root_causes": [], "downstream_impact": [], "corrective_actions": []
            }),
        }

    on_track = sum(1 for k in kpis_out.values() if -3 <= k["avg_gap_pct"])
    behind   = sum(1 for k in kpis_out.values() if k["avg_gap_pct"] < -3)
    ahead    = sum(1 for k in kpis_out.values() if k["avg_gap_pct"] >= 3)
    on_track = on_track - ahead

    return {
        "has_projection":  True,
        "has_overlap":     True,
        "summary": {
            "on_track":              on_track,
            "behind":                behind,
            "ahead":                 ahead,
            "total_months_compared": len(overlap),
        },
        "kpis": kpis_out,
    }


# ─── KPI Audit Excel Export ─────────────────────────────────────────────────

@router.get("/api/export/kpi-audit.xlsx", tags=["Export"])
def export_kpi_audit(request: Request):
    """
    Generate a comprehensive Excel workbook that shows every KPI tracked by the
    platform: raw monthly data, formula, computation rationale, health-score
    component membership, criticality ranking breakdown, benchmarks, causation
    graph (depends on / feeds into), and which platform tabs use each KPI.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from core.health_score import compute_health_score

    workspace_id = _require_workspace(request)
    conn = get_db()

    # ── Gather all data ────────────────────────────────────────────────────

    # 1. Build master KPI lookup
    ALL_KPIS = {}
    for d in KPI_DEFS:
        ALL_KPIS[d["key"]] = {**d, "source": "core"}
    for d in EXTENDED_ONTOLOGY_METRICS:
        if d["key"] not in ALL_KPIS:
            ALL_KPIS[d["key"]] = {**d, "source": "extended", "formula": "—"}

    # 2. Fetch monthly data (stored as JSON blobs per month)
    rows = conn.execute(
        "SELECT year, month, data_json FROM monthly_data WHERE workspace_id=? ORDER BY year, month",
        [workspace_id],
    ).fetchall()
    monthly = {}  # key -> [(year, month, value), ...]
    for r in rows:
        year, month = r["year"], r["month"]
        try:
            data = json.loads(r["data_json"]) if isinstance(r["data_json"], str) else {}
        except (json.JSONDecodeError, TypeError):
            continue
        for kpi_key, value in data.items():
            if kpi_key.startswith("_"):
                continue
            if value is not None:
                monthly.setdefault(kpi_key, []).append((year, month, value))

    # 3. Fetch targets
    tgt_rows = conn.execute(
        "SELECT kpi_key, target_value FROM kpi_targets WHERE workspace_id=?",
        [workspace_id],
    ).fetchall()
    targets = {r["kpi_key"]: r["target_value"] for r in tgt_rows}

    # 4. Compute health score & criticality
    hs = compute_health_score(conn, workspace_id)
    comp_detail = hs.get("component_detail", {})
    composite_ranked = hs.get("composite_ranked", [])

    conn.close()

    # Build lookup dicts
    crit_lookup = {c["key"]: c for c in composite_ranked}

    # Momentum / target / risk KPI sets
    momentum_set = {}
    for k in comp_detail.get("momentum", {}).get("kpis", []):
        momentum_set[k["key"]] = k
    target_set = {}
    for k in comp_detail.get("target_achievement", {}).get("kpis", []):
        target_set[k["key"]] = k
    risk_set = set()
    for k in comp_detail.get("risk", {}).get("kpis", []):
        risk_set.add(k["key"])

    # Domain lookup
    def get_domain(key):
        d = ALL_KPIS.get(key, {}).get("domain")
        if d:
            return d
        return ONTOLOGY_DOMAIN.get(key, "other")

    # Causation helpers
    def depends_on(key):
        rules = ALL_CAUSATION_RULES.get(key, {})
        return "; ".join(rules.get("root_causes", []))

    def feeds_into(key):
        return ", ".join(ALL_CAUSATION_RULES.get(key, {}).get("downstream_impact", []))

    # Reverse feeds: which KPIs list this one in their downstream_impact
    reverse_feeds = {}
    for k, v in ALL_CAUSATION_RULES.items():
        for downstream in v.get("downstream_impact", []):
            reverse_feeds.setdefault(downstream, []).append(k)

    # Tab usage mapping
    TAB_USAGE = {
        # Home: all scored KPIs
        "Home (Health Score)": set(k for k in ALL_KPIS if monthly.get(k)),
        "Variance Command": set(k for k in ALL_KPIS if k in targets and monthly.get(k)),
        "Forward Signals": set(k for k in ALL_KPIS if monthly.get(k) and len(monthly[k]) >= 3),
        "Trend Explorer": set(k for k in ALL_KPIS if monthly.get(k)),
        "Performance Fingerprint": set(k for k in ALL_KPIS if monthly.get(k)),
        "Plan vs Actual": set(k for k in ALL_KPIS if k in targets and monthly.get(k)),
        "Executive Brief": set(k for k in ALL_KPIS if monthly.get(k)),
        "Board Pack": set(k for k in ALL_KPIS if monthly.get(k)),
    }

    def tabs_for_kpi(key):
        tabs = [tab for tab, kset in TAB_USAGE.items() if key in kset]
        return ", ".join(tabs) if tabs else "—"

    # ── Build workbook ─────────────────────────────────────────────────────

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="0055A4", end_color="0055A4", fill_type="solid")
    alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    thin_border = Border(
        bottom=Side(style="thin", color="E2E8F0"),
    )
    wrap = Alignment(wrap_text=True, vertical="top")

    def style_header(ws, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.freeze_panes = "A2"

    def auto_width(ws, max_w=45):
        for col in ws.columns:
            mx = 0
            letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    mx = max(mx, min(len(str(cell.value)), max_w))
            ws.column_dimensions[letter].width = max(mx + 3, 12)

    # ── Computation map: how each KPI is derived from basis numbers ────────
    # Each entry: list of (input_label, input_kpi_key_or_None, fallback_desc)
    # plus an excel_formula template using {A}, {B}, {C} placeholders for input
    # cell references, and a human-readable explanation.

    COMPUTATION_MAP = {
        # ── Margin & Profitability ──
        "gross_margin":        {"inputs": [("Revenue", None, "Total Revenue from data source"),
                                           ("COGS", None, "Cost of Goods Sold from data source")],
                                "excel": "=({A}-{B})/{A}*100",
                                "explain": "(Revenue - COGS) / Revenue x 100"},
        "operating_margin":    {"inputs": [("Revenue", None, "Total Revenue"),
                                           ("COGS", None, "Cost of Goods Sold"),
                                           ("OpEx", None, "Operating Expenses")],
                                "excel": "=({A}-{B}-{C})/{A}*100",
                                "explain": "(Revenue - COGS - OpEx) / Revenue x 100"},
        "ebitda_margin":       {"inputs": [("Revenue", None, "Total Revenue"),
                                           ("COGS", None, "Cost of Goods Sold"),
                                           ("OpEx", None, "Operating Expenses")],
                                "excel": "=(({A}-{B}-{C})*1.15)/{A}*100",
                                "explain": "((Revenue - COGS - OpEx) x 1.15) / Revenue x 100  [1.15 = D&A add-back proxy]"},
        "opex_ratio":          {"inputs": [("OpEx", None, "Operating Expenses"),
                                           ("Revenue", None, "Total Revenue")],
                                "excel": "={A}/{B}*100",
                                "explain": "OpEx / Revenue x 100"},
        "contribution_margin": {"inputs": [("Revenue", None, "Total Revenue"),
                                           ("COGS", None, "Cost of Goods Sold"),
                                           ("OpEx", None, "Operating Expenses")],
                                "excel": "=({A}-{B}-{C}*0.3)/{A}*100",
                                "explain": "(Revenue - COGS - 30% of OpEx) / Revenue x 100"},
        "operating_leverage":  {"inputs": [("Operating Margin This Month", "operating_margin", None),
                                           ("Operating Margin Prev Month", None, "Prior month operating_margin")],
                                "excel": "=({A}-{B})/{B}",
                                "explain": "% change in Operating Income / % change in Revenue (approx via margin delta)"},
        "margin_volatility":   {"inputs": [("Gross Margin (6 months)", "gross_margin", "Rolling 6-month series")],
                                "excel": "=STDEV of last 6 gross_margin values",
                                "explain": "Standard deviation of gross_margin over trailing 6 months"},
        # ── Revenue ──
        "revenue_growth":      {"inputs": [("Revenue This Month", None, "Current month total revenue"),
                                           ("Revenue Prev Month", None, "Prior month total revenue")],
                                "excel": "=({A}-{B})/{B}*100",
                                "explain": "(Revenue_Month - Revenue_PrevMonth) / Revenue_PrevMonth x 100"},
        "arr_growth":          {"inputs": [("ARR This Month", None, "Current month ARR (MRR x 12)"),
                                           ("ARR Prev Month", None, "Prior month ARR")],
                                "excel": "=({A}-{B})/{B}*100",
                                "explain": "(ARR_Month - ARR_PrevMonth) / ARR_PrevMonth x 100"},
        "revenue_quality":     {"inputs": [("Recurring Revenue", None, "Subscription/recurring revenue"),
                                           ("Total Revenue", None, "Total revenue")],
                                "excel": "={A}/{B}*100",
                                "explain": "Recurring_Revenue / Total_Revenue x 100"},
        "recurring_revenue":   {"inputs": [("Recurring Revenue", None, "Subscription/recurring revenue"),
                                           ("Total Revenue", None, "Total revenue")],
                                "excel": "={A}/{B}*100",
                                "explain": "Recurring_Revenue / Total_Revenue x 100"},
        "customer_concentration": {"inputs": [("Top Customer Rev", None, "Revenue from largest customer"),
                                              ("Total Revenue", None, "Total revenue")],
                                   "excel": "={A}/{B}*100",
                                   "explain": "Top_Customer_Revenue / Total_Revenue x 100 (or Pareto approx: min(100/N_customers x 2.5, 100))"},
        "revenue_momentum":    {"inputs": [("Current Rev Growth", "revenue_growth", None),
                                           ("12M Avg Rev Growth", None, "Average revenue_growth over 12 months")],
                                "excel": "={A}/{B}",
                                "explain": "Current_Revenue_Growth / 12M_Avg_Revenue_Growth (>1 = accelerating)"},
        "revenue_fragility":   {"inputs": [("Customer Concentration", "customer_concentration", None),
                                           ("Churn Rate", "churn_rate", None),
                                           ("NRR", "nrr", None)],
                                "excel": "=({A}*{B})/{C}",
                                "explain": "(Customer_Concentration x Churn_Rate) / NRR"},
        "pricing_power_index": {"inputs": [("ARPU Change %", None, "Month-over-month ARPU % change"),
                                           ("Customer Volume Change %", None, "Month-over-month customer count % change")],
                                "excel": "={A}-{B}",
                                "explain": "Delta_ARPU% - Delta_Customer_Volume% (positive = pricing power)"},
        "avg_deal_size":       {"inputs": [("Won Deal Value", None, "Total value of won deals"),
                                           ("Deals Won", None, "Count of won deals")],
                                "excel": "={A}/{B}",
                                "explain": "Total_Won_Value / Number_of_Deals_Won"},
        "expansion_rate":      {"inputs": [("Expansion Revenue", None, "Upsell + cross-sell revenue"),
                                           ("Beginning MRR", None, "MRR at start of period")],
                                "excel": "={A}/{B}*100",
                                "explain": "Expansion_Revenue / Beginning_MRR x 100"},
        "gross_dollar_ret":    {"inputs": [("End MRR excl Expansion", None, "Ending MRR minus expansion"),
                                           ("Beginning MRR", None, "MRR at start of period")],
                                "excel": "={A}/{B}*100",
                                "explain": "(MRR_End - Expansion) / MRR_Start x 100"},
        "customer_ltv":        {"inputs": [("ARPU", None, "Average Revenue Per User"),
                                           ("Gross Margin %", "gross_margin", None),
                                           ("Monthly Churn Rate", "churn_rate", None)],
                                "excel": "=({A}*{B}/100)/({C}/100)",
                                "explain": "(ARPU x Gross_Margin%) / Monthly_Churn_Rate"},
        "ltv_cac":             {"inputs": [("Customer LTV", "customer_ltv", None),
                                           ("CAC", None, "Customer Acquisition Cost = S&M / New_Customers")],
                                "excel": "={A}/{B}",
                                "explain": "LTV / CAC"},
        # ── Growth & Acquisition ──
        "cpl":                 {"inputs": [("Marketing Spend", None, "Total marketing spend"),
                                           ("Leads Generated", None, "Total new leads")],
                                "excel": "={A}/{B}",
                                "explain": "Marketing_Spend / Leads_Generated"},
        "mql_sql_rate":        {"inputs": [("SQLs", None, "Sales Qualified Leads"),
                                           ("MQLs", None, "Marketing Qualified Leads")],
                                "excel": "={A}/{B}*100",
                                "explain": "SQLs / MQLs x 100"},
        "pipeline_velocity":   {"inputs": [("Opportunities", None, "Number of open opportunities"),
                                           ("Win Rate", "win_rate", None),
                                           ("Avg Deal Size", "avg_deal_size", None),
                                           ("Sales Cycle Days", None, "Average days to close")],
                                "excel": "={A}*{B}/100*{C}/{D}",
                                "explain": "Opportunities x Win_Rate x Avg_Deal_Size / Sales_Cycle_Days"},
        "win_rate":            {"inputs": [("Deals Won", None, "Count of deals won"),
                                           ("Total Deals", None, "Total deals in period")],
                                "excel": "={A}/{B}*100",
                                "explain": "Deals_Won / Total_Deals x 100"},
        "pipeline_conversion": {"inputs": [("Won Value", None, "Total value of won deals"),
                                           ("Total Pipeline Value", None, "Total pipeline value")],
                                "excel": "={A}/{B}*100",
                                "explain": "Won_Value / Total_Pipeline_Value x 100"},
        "quota_attainment":    {"inputs": [("Actual Bookings", None, "Total bookings achieved"),
                                           ("Quota Target", None, "Sales quota for period")],
                                "excel": "={A}/{B}*100",
                                "explain": "Actual_Bookings / Quota_Target x 100"},
        "marketing_roi":       {"inputs": [("Revenue from Marketing", None, "Revenue attributed to marketing"),
                                           ("Marketing Spend", None, "Total marketing investment")],
                                "excel": "={A}/{B}",
                                "explain": "Marketing_Attributed_Revenue / Marketing_Spend"},
        "growth_efficiency":   {"inputs": [("ARR Growth Rate", "arr_growth", None),
                                           ("Burn Multiple", "burn_multiple", None)],
                                "excel": "={A}/{B}",
                                "explain": "ARR_Growth_Rate / Burn_Multiple"},
        # ── Retention ──
        "churn_rate":          {"inputs": [("Lost Customers", None, "Customers lost this month"),
                                           ("Total Customers (prev)", None, "Total customers at start of month")],
                                "excel": "={A}/{B}*100",
                                "explain": "Lost_Customers / Starting_Customers x 100"},
        "nrr":                 {"inputs": [("Starting MRR", None, "MRR at beginning of period"),
                                           ("Expansion", None, "Expansion revenue"),
                                           ("Churn", None, "Churned revenue"),
                                           ("Contraction", None, "Contraction revenue")],
                                "excel": "=({A}+{B}-{C}-{D})/{A}*100",
                                "explain": "(MRR_Start + Expansion - Churn - Contraction) / MRR_Start x 100"},
        "logo_retention":      {"inputs": [("Retained Customers", None, "Customers still active"),
                                           ("Total Customers (prev)", None, "Total customers at start of month")],
                                "excel": "={A}/{B}*100",
                                "explain": "Retained_Customers / Starting_Customers x 100"},
        "customer_decay_slope":{"inputs": [("Churn Rate This Month", "churn_rate", None),
                                           ("Churn Rate Prev Month", None, "Prior month churn_rate")],
                                "excel": "={A}-{B}",
                                "explain": "Delta_Churn_Rate month-over-month (positive = worsening)"},
        # ── Efficiency ──
        "sales_efficiency":    {"inputs": [("New ARR", None, "New Annual Recurring Revenue added"),
                                           ("S&M Spend", None, "Sales & Marketing spend")],
                                "excel": "={A}/{B}",
                                "explain": "New_ARR / Sales_Marketing_Spend"},
        "cac_payback":         {"inputs": [("CAC", None, "Customer Acquisition Cost"),
                                           ("ARPU", None, "Average Revenue Per User per month"),
                                           ("Gross Margin %", "gross_margin", None)],
                                "excel": "={A}/({B}*{C}/100)",
                                "explain": "CAC / (ARPU x Gross_Margin%) — result in months"},
        "burn_multiple":       {"inputs": [("Net Burn", None, "Total Expenses - Total Revenue"),
                                           ("Net New ARR", None, "New ARR added in period")],
                                "excel": "={A}/{B}",
                                "explain": "Net_Burn / Net_New_ARR (lower is better; <1 = efficient)"},
        "burn_convexity":      {"inputs": [("Burn Multiple This Month", "burn_multiple", None),
                                           ("Burn Multiple Prev Month", None, "Prior month burn_multiple")],
                                "excel": "={A}-{B}",
                                "explain": "Delta_Burn_Multiple month-over-month (negative = improving)"},
        "headcount_eff":       {"inputs": [("Revenue", None, "Monthly revenue"),
                                           ("Headcount", None, "Total employees")],
                                "excel": "={A}/{B}",
                                "explain": "Monthly_Revenue / Headcount"},
        "rev_per_employee":    {"inputs": [("Annual Revenue", None, "Monthly revenue x 12"),
                                           ("Headcount", None, "Total employees")],
                                "excel": "={A}/{B}",
                                "explain": "Annualized_Revenue / Headcount"},
        "billable_utilization":{"inputs": [("Billable Hours", None, "Hours billed to clients"),
                                           ("Total Hours", None, "Total available working hours")],
                                "excel": "={A}/{B}*100",
                                "explain": "Billable_Hours / Total_Available_Hours x 100"},
        # ── Cash Flow ──
        "dso":                 {"inputs": [("Accounts Receivable", None, "Total AR balance"),
                                           ("Revenue", None, "Monthly revenue")],
                                "excel": "={A}/{B}*30",
                                "explain": "(Accounts_Receivable / Revenue) x 30 days"},
        "ar_turnover":         {"inputs": [("Annual Revenue", None, "Net credit sales (annualized)"),
                                           ("Average AR", None, "Average Accounts Receivable balance")],
                                "excel": "={A}/{B}",
                                "explain": "Net_Credit_Sales / Average_AR"},
        "avg_collection_period":{"inputs": [("DSO", "dso", None)],
                                 "excel": "={A}",
                                 "explain": "365 / AR_Turnover (equivalent to DSO)"},
        "cash_conv_cycle":     {"inputs": [("DSO", "dso", None),
                                           ("DIO (est)", None, "Days Inventory Outstanding"),
                                           ("DPO (est)", None, "Days Payable Outstanding")],
                                "excel": "={A}+{B}-{C}",
                                "explain": "DSO + DIO - DPO  [simplified: DSO + 10 day buffer when DIO/DPO unavailable]"},
        "cei":                 {"inputs": [("Beginning AR", None, "AR at period start"),
                                           ("Credit Sales", None, "Credit sales in period"),
                                           ("Ending AR", None, "AR at period end"),
                                           ("Current AR", None, "AR still within terms")],
                                "excel": "=({A}+{B}-{C})/({A}+{B}-{D})*100",
                                "explain": "(Beg_AR + Sales - End_AR) / (Beg_AR + Sales - Current_AR) x 100"},
        "ar_aging_current":    {"inputs": [("Current Invoices (0-30d)", None, "Invoices within terms"),
                                           ("Total Invoices", None, "All outstanding invoices")],
                                "excel": "={A}/{B}*100",
                                "explain": "Current_AR / Total_AR x 100"},
        "ar_aging_overdue":    {"inputs": [("Overdue Invoices (30+d)", None, "Invoices past terms"),
                                           ("Total Invoices", None, "All outstanding invoices")],
                                "excel": "={A}/{B}*100",
                                "explain": "Overdue_AR / Total_AR x 100"},
        "cash_runway":         {"inputs": [("Cash Balance", None, "Current cash on hand"),
                                           ("Monthly Burn", None, "Net monthly cash outflow")],
                                "excel": "={A}/{B}",
                                "explain": "Cash_Balance / Monthly_Burn (result in months)"},
        "current_ratio":       {"inputs": [("Current Assets", None, "Total current assets"),
                                           ("Current Liabilities", None, "Total current liabilities")],
                                "excel": "={A}/{B}",
                                "explain": "Current_Assets / Current_Liabilities"},
        "working_capital":     {"inputs": [("Current Assets", None, "Total current assets"),
                                           ("Current Liabilities", None, "Total current liabilities"),
                                           ("Total Assets", None, "Total assets")],
                                "excel": "=({A}-{B})/{C}",
                                "explain": "(Current_Assets - Current_Liabilities) / Total_Assets"},
        # ── Risk ──
        "contraction_rate":    {"inputs": [("Contraction MRR", None, "MRR lost to downgrades"),
                                           ("Beginning MRR", None, "MRR at start of period")],
                                "excel": "={A}/{B}*100",
                                "explain": "Contraction_MRR / Beginning_MRR x 100"},
        "payback_period":      {"inputs": [("Total Investment", None, "Total capital invested"),
                                           ("Annual Cash Flow", None, "Annual net cash flow")],
                                "excel": "={A}/{B}",
                                "explain": "Total_Investment / Annual_Cash_Flow (result in months)"},
    }

    # Build a lookup: kpi_key -> latest value (for cross-referencing input KPIs)
    latest_values = {}
    for key in ALL_KPIS:
        series = monthly.get(key, [])
        if series:
            latest_values[key] = series[-1][2]

    # Also get prior-month values for delta-based KPIs
    prior_values = {}
    for key in ALL_KPIS:
        series = monthly.get(key, [])
        if len(series) >= 2:
            prior_values[key] = series[-2][2]

    # ── Back-calculate implied raw values (normalised to Revenue = 100) ────
    # This lets the user verify margin/ratio KPIs even when raw accounting
    # data isn't stored separately.  We label these "implied" so the auditor
    # knows they are derived, not source-of-truth.
    lv = latest_values
    pv = prior_values
    implied = {}

    rev_norm = 100.0  # normalised revenue base
    implied["Revenue (norm=100)"] = rev_norm
    gm = lv.get("gross_margin")
    if gm is not None:
        implied["COGS (implied)"] = round(rev_norm * (1 - gm / 100), 4)
    opr = lv.get("opex_ratio")
    if opr is not None:
        implied["OpEx (implied)"] = round(rev_norm * opr / 100, 4)
    elif lv.get("operating_margin") is not None and gm is not None:
        implied["OpEx (implied)"] = round(rev_norm * (gm / 100 - lv["operating_margin"] / 100), 4)

    # Prev-month values for growth KPIs
    rg = lv.get("revenue_growth")
    if rg is not None:
        implied["Revenue PrevMonth (implied)"] = round(rev_norm / (1 + rg / 100), 4)
    ag = lv.get("arr_growth")
    if ag is not None:
        arr_now = rev_norm * 12
        implied["ARR PrevMonth (implied)"] = round(arr_now / (1 + ag / 100), 4)
        implied["ARR This Month (implied)"] = round(arr_now, 4)

    # DSO -> implied AR
    dso_val = lv.get("dso")
    if dso_val is not None:
        implied["AR (implied)"] = round(rev_norm * dso_val / 30, 4)

    # Map raw-input labels to implied values
    RAW_IMPLIED = {
        "Revenue":              rev_norm,
        "Total Revenue":        rev_norm,
        "Monthly revenue":      rev_norm,
        "COGS":                 implied.get("COGS (implied)"),
        "OpEx":                 implied.get("OpEx (implied)") or implied.get("OpEx from ratio (implied)"),
        "Operating Expenses":   implied.get("OpEx (implied)") or implied.get("OpEx from ratio (implied)"),
        "Revenue This Month":   rev_norm,
        "Revenue Prev Month":   implied.get("Revenue PrevMonth (implied)"),
        "ARR This Month":       implied.get("ARR This Month (implied)"),
        "ARR Prev Month":       implied.get("ARR PrevMonth (implied)"),
        "Accounts Receivable":  implied.get("AR (implied)"),
        "Annual Revenue":       round(rev_norm * 12, 4),
    }

    # ── Sheet 1: KPI Master ────────────────────────────────────────────────

    ws1 = wb.active
    ws1.title = "KPI Master"
    # Columns A-R (18 columns)
    headers1 = [
        "KPI Key", "KPI Name", "Unit", "Direction", "Domain",
        "Computation Formula",
        "Input A", "Value A", "Input B", "Value B", "Input C", "Value C", "Input D", "Value D",
        "Latest Value (Platform)", "Verification Formula", "Verified Value",
        "Target", "Gap %",
        "Rationale / Root Causes", "Feeds Into (Downstream)",
        "Fed By (Upstream KPIs)", "Corrective Actions",
        "Used On Tabs",
    ]
    ws1.append(headers1)
    style_header(ws1, len(headers1))

    formula_font = Font(name="Consolas", size=9, color="6B21A8")
    input_label_font = Font(size=9, color="0055A4", italic=True)
    verify_fill = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")

    sorted_keys = sorted(ALL_KPIS.keys(), key=lambda k: (get_domain(k), ALL_KPIS[k].get("name", k)))
    for i, key in enumerate(sorted_keys):
        kpi = ALL_KPIS[key]
        series = monthly.get(key, [])
        latest = series[-1][2] if series else None
        target = targets.get(key)
        gap = None
        if latest is not None and target and target != 0:
            direction = kpi.get("direction", "higher")
            if direction == "higher":
                gap = round((latest - target) / abs(target) * 100, 1)
            else:
                gap = round((target - latest) / abs(target) * 100, 1)

        fed_by = ", ".join(reverse_feeds.get(key, []))
        actions = "; ".join(ALL_CAUSATION_RULES.get(key, {}).get("corrective_actions", []))

        comp = COMPUTATION_MAP.get(key)
        row_num = i + 2  # 1-indexed, row 1 is header

        # Input columns: up to 4 inputs (A-D), each with label + value
        input_labels = ["", "", "", ""]
        input_vals = [None, None, None, None]
        explain = kpi.get("formula", "—")
        excel_formula = None

        if comp:
            explain = comp["explain"]
            inputs = comp["inputs"]
            cell_refs = {}
            for j, inp in enumerate(inputs[:4]):
                label, ref_kpi, fallback = inp
                input_labels[j] = label
                if ref_kpi and ref_kpi in latest_values:
                    input_vals[j] = round(latest_values[ref_kpi], 4)
                elif ref_kpi and ref_kpi in prior_values:
                    input_vals[j] = round(prior_values[ref_kpi], 4)
                elif label in RAW_IMPLIED and RAW_IMPLIED[label] is not None:
                    # Use back-calculated implied value (normalised to Rev=100)
                    input_vals[j] = RAW_IMPLIED[label]
                    input_labels[j] = f"{label} *implied, Rev=100"
                else:
                    input_vals[j] = fallback or "from data source"
                # Map {A},{B},{C},{D} to cell refs for the verification formula
                # Value columns are H, J, L, N (columns 8, 10, 12, 14)
                val_col = get_column_letter(8 + j * 2)  # H, J, L, N
                cell_refs[chr(65 + j)] = f"{val_col}{row_num}"

            # Build the Excel verification formula if all numeric inputs are available
            has_all_numeric = all(
                isinstance(input_vals[j], (int, float))
                for j in range(len(inputs[:4]))
            )
            if has_all_numeric and comp.get("excel"):
                excel_formula = comp["excel"]
                for placeholder, ref in cell_refs.items():
                    excel_formula = excel_formula.replace("{" + placeholder + "}", ref)

        row = [
            key,                                        # A: KPI Key
            kpi.get("name", key),                       # B: KPI Name
            kpi.get("unit", "—"),                       # C: Unit
            kpi.get("direction", "—"),                  # D: Direction
            get_domain(key),                            # E: Domain
            explain,                                    # F: Computation Formula
            input_labels[0],                            # G: Input A label
            input_vals[0],                              # H: Value A
            input_labels[1],                            # I: Input B label
            input_vals[1],                              # J: Value B
            input_labels[2],                            # K: Input C label
            input_vals[2],                              # L: Value C
            input_labels[3],                            # M: Input D label
            input_vals[3],                              # N: Value D
            round(latest, 4) if latest is not None else None,  # O: Latest Value
            None,                                       # P: Verification Formula (set below)
            None,                                       # Q: Verified Value (set below)
            target,                                     # R: Target
            gap,                                        # S: Gap %
            depends_on(key),                            # T: Rationale
            feeds_into(key),                            # U: Feeds Into
            fed_by,                                     # V: Fed By
            actions,                                    # W: Corrective Actions
            tabs_for_kpi(key),                          # X: Used On Tabs
        ]
        ws1.append(row)

        # Set the verification formula in column Q (17) if we have one
        if excel_formula:
            ws1.cell(row=row_num, column=17).value = excel_formula
            ws1.cell(row=row_num, column=17).font = formula_font
            ws1.cell(row=row_num, column=17).fill = verify_fill
        else:
            ws1.cell(row=row_num, column=17).value = "N/A (raw inputs from data source)"
            ws1.cell(row=row_num, column=17).font = Font(size=9, color="9CA3AF", italic=True)

        # Show the formula text in column P
        if comp and comp.get("excel"):
            ws1.cell(row=row_num, column=16).value = comp["excel"]
            ws1.cell(row=row_num, column=16).font = formula_font

        # Style input label cells
        for j in range(4):
            label_col = 7 + j * 2  # G, I, K, M
            ws1.cell(row=row_num, column=label_col).font = input_label_font
            # If value is a string (not numeric), style it differently
            val_col = 8 + j * 2
            cell = ws1.cell(row=row_num, column=val_col)
            if isinstance(cell.value, str):
                cell.font = Font(size=9, color="9CA3AF", italic=True)

        if i % 2 == 1:
            for c in range(1, len(headers1) + 1):
                ws1.cell(row=i + 2, column=c).fill = alt_fill

    # Add audit legend at bottom
    ws1.append([])
    ws1.append([])
    legend_row = ws1.max_row + 1
    ws1.cell(row=legend_row, column=1, value="AUDIT NOTES").font = Font(bold=True, size=11, color="0055A4")
    ws1.cell(row=legend_row + 1, column=1, value="* Values marked 'implied, Rev=100' are back-calculated by normalising Revenue to 100. Replace with your actual accounting values and the Verified Value column will recompute.").font = Font(size=9, color="6B7280", italic=True)
    ws1.cell(row=legend_row + 2, column=1, value="* Column P shows the formula template. Column Q contains a live Excel formula referencing the input value cells (H, J, L, N). Edit those cells to audit with your own numbers.").font = Font(size=9, color="6B7280", italic=True)
    ws1.cell(row=legend_row + 3, column=1, value="* KPIs without verification formulas depend on raw accounting data (invoices, headcount, pipeline) not stored as separate KPI values. Their formula is shown in column F for manual verification.").font = Font(size=9, color="6B7280", italic=True)
    ws1.cell(row=legend_row + 4, column=1, value=f"* Latest period shown. Full monthly history is on the 'Monthly Data' sheet. Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}.").font = Font(size=9, color="6B7280", italic=True)

    auto_width(ws1)

    # ── Sheet 2: Monthly Data ──────────────────────────────────────────────

    ws2 = wb.create_sheet("Monthly Data")

    # Collect all periods
    all_periods = set()
    for series in monthly.values():
        for y, m, _ in series:
            all_periods.add((y, m))
    all_periods = sorted(all_periods)

    if all_periods:
        headers2 = ["KPI Key", "KPI Name", "Unit", "Direction"] + [
            f"{y}-{m:02d}" for y, m in all_periods
        ]
        ws2.append(headers2)
        style_header(ws2, len(headers2))

        for i, key in enumerate(sorted_keys):
            kpi = ALL_KPIS[key]
            series = monthly.get(key, [])
            val_map = {(y, m): v for y, m, v in series}
            row = [key, kpi.get("name", key), kpi.get("unit", "—"), kpi.get("direction", "—")]
            for p in all_periods:
                v = val_map.get(p)
                row.append(round(v, 4) if v is not None else None)
            ws2.append(row)
            if i % 2 == 1:
                for c in range(1, len(headers2) + 1):
                    ws2.cell(row=i + 2, column=c).fill = alt_fill
        auto_width(ws2, max_w=12)
    else:
        ws2.append(["No monthly data available"])

    # ── Sheet 3: Health Score Breakdown ────────────────────────────────────

    ws3 = wb.create_sheet("Health Score")

    # Summary row
    ws3.append(["Overall Health Score", hs.get("score"), "", "Grade", hs.get("grade"), "", "Label", hs.get("label")])
    ws3.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws3.cell(row=1, column=2).font = Font(bold=True, size=14, color="0055A4")
    ws3.append([])

    weights = hs.get("weights", {})
    ws3.append(["Component", "Score (0-100)", "Weight", "Weighted Contribution"])
    style_header(ws3, 4)
    for comp_name, w_key in [("Momentum", "momentum"), ("Target Achievement", "target"), ("Risk", "risk")]:
        sc = comp_detail.get(w_key if w_key != "target" else "target_achievement", {}).get("score", 0)
        w = weights.get(w_key, 0)
        ws3.append([comp_name, round(sc, 1), f"{w*100:.0f}%", round(sc * w, 1)])

    ws3.append([])
    ws3.append([])

    # Momentum KPIs
    ws3.append(["MOMENTUM COMPONENT — Per-KPI Detail"])
    ws3.cell(row=ws3.max_row, column=1).font = Font(bold=True, size=11, color="0055A4")
    ws3.append(["KPI Key", "KPI Name", "Status", "Delta %", "Interpretation"])
    style_header(ws3, 5)
    for k in comp_detail.get("momentum", {}).get("kpis", []):
        interp = {
            "improving": "Trending upward — positive momentum",
            "declining": "Trending downward — losing momentum",
            "stable": "No significant change in trend",
        }.get(k["status"], "")
        ws3.append([k["key"], k.get("name", k["key"]), k["status"],
                     round(k.get("delta_pct", 0), 2), interp])

    ws3.append([])
    ws3.append([])

    # Target Achievement KPIs
    ws3.append(["TARGET ACHIEVEMENT COMPONENT — Per-KPI Detail"])
    ws3.cell(row=ws3.max_row, column=1).font = Font(bold=True, size=11, color="0055A4")
    ws3.append(["KPI Key", "KPI Name", "Avg Value", "Target", "On Target?", "Direction"])
    style_header(ws3, 6)
    for k in comp_detail.get("target_achievement", {}).get("kpis", []):
        ws3.append([k["key"], k.get("name", k["key"]),
                     round(k.get("avg", 0), 4), round(k.get("target", 0), 4),
                     "Yes" if k.get("on_target") else "No",
                     k.get("direction", "—")])

    ws3.append([])
    ws3.append([])

    # Risk KPIs
    ws3.append(["RISK COMPONENT — Per-KPI Detail"])
    ws3.cell(row=ws3.max_row, column=1).font = Font(bold=True, size=11, color="0055A4")
    ws3.append(["KPI Key", "KPI Name", "Avg Value", "Target", "Direction", "Risk Note"])
    style_header(ws3, 6)
    for k in comp_detail.get("risk", {}).get("kpis", []):
        ws3.append([k["key"], k.get("name", k["key"]),
                     round(k.get("avg", 0), 4),
                     round(k.get("target", 0), 4) if k.get("target") else "—",
                     k.get("direction", "—"),
                     "Red status — significantly off target"])

    auto_width(ws3)

    # ── Sheet 4: Criticality Ranking ───────────────────────────────────────

    ws4 = wb.create_sheet("Criticality Ranking")
    headers4 = [
        "Rank", "KPI Key", "KPI Name", "Composite Score",
        "Gap Score", "Gap Weight", "Gap Contribution",
        "Trend Score", "Trend Weight", "Trend Contribution",
        "Impact Score", "Impact Weight", "Impact Contribution",
        "Domain Score", "Domain Weight", "Domain Contribution",
        "Domain", "Direction", "Avg Value", "Target", "Gap %",
    ]
    ws4.append(headers4)
    style_header(ws4, len(headers4))

    for i, c in enumerate(composite_ranked):
        w = c.get("weights_used", CRIT_DEFAULT_WEIGHTS)
        row = [
            c.get("rank", i + 1),
            c["key"],
            ALL_KPIS.get(c["key"], {}).get("name", c["key"]),
            round(c.get("composite", 0), 2),
            round(c.get("gap_score", 0), 2),
            f"{w.get('gap', 0.25)*100:.0f}%",
            round(c.get("gap_score", 0) * w.get("gap", 0.25), 2),
            round(c.get("trend_score", 0), 2),
            f"{w.get('trend', 0.25)*100:.0f}%",
            round(c.get("trend_score", 0) * w.get("trend", 0.25), 2),
            round(c.get("impact_score", 0), 2),
            f"{w.get('impact', 0.30)*100:.0f}%",
            round(c.get("impact_score", 0) * w.get("impact", 0.30), 2),
            round(c.get("domain_score", 0), 2),
            f"{w.get('domain', 0.20)*100:.0f}%",
            round(c.get("domain_score", 0) * w.get("domain", 0.20), 2),
            c.get("domain_label", c.get("domain", "—")),
            c.get("direction", "—"),
            round(c.get("avg", 0), 4) if c.get("avg") else None,
            round(c.get("target", 0), 4) if c.get("target") else None,
            round(c.get("gap_pct", 0), 2) if c.get("gap_pct") is not None else None,
        ]
        ws4.append(row)
        if i % 2 == 1:
            for col in range(1, len(headers4) + 1):
                ws4.cell(row=i + 2, column=col).fill = alt_fill

    # Formula explanation row
    ws4.append([])
    ws4.append(["Formula: Composite = (Gap × Gap_Weight) + (Trend × Trend_Weight) + (Impact × Impact_Weight) + (Domain × Domain_Weight)"])
    ws4.cell(row=ws4.max_row, column=1).font = Font(italic=True, color="6B7280")
    auto_width(ws4, max_w=18)

    # ── Sheet 5: Benchmarks ────────────────────────────────────────────────

    ws5 = wb.create_sheet("Benchmarks")
    stages = ["seed", "series_a", "series_b", "series_c"]
    stage_labels = {"seed": "Seed", "series_a": "Series A", "series_b": "Series B", "series_c": "Series C+"}
    headers5 = ["KPI Key", "KPI Name", "Unit"]
    for s in stages:
        headers5 += [f"{stage_labels[s]} P25", f"{stage_labels[s]} P50", f"{stage_labels[s]} P75"]
    ws5.append(headers5)
    style_header(ws5, len(headers5))

    for i, (key, stage_data) in enumerate(sorted(BENCHMARKS.items())):
        kpi = ALL_KPIS.get(key, {})
        row = [key, kpi.get("name", key), kpi.get("unit", "—")]
        for s in stages:
            sd = stage_data.get(s, {})
            row += [sd.get("p25"), sd.get("p50"), sd.get("p75")]
        ws5.append(row)
        if i % 2 == 1:
            for col in range(1, len(headers5) + 1):
                ws5.cell(row=i + 2, column=col).fill = alt_fill
    auto_width(ws5, max_w=16)

    # ── Sheet 6: Causation Graph ───────────────────────────────────────────

    ws6 = wb.create_sheet("Causation Graph")
    headers6 = [
        "KPI Key", "KPI Name", "Domain",
        "Root Causes (why it misses)", "Downstream Impact (KPI keys affected)",
        "Upstream KPIs (that feed into this)", "Corrective Actions",
    ]
    ws6.append(headers6)
    style_header(ws6, len(headers6))

    for i, key in enumerate(sorted(ALL_CAUSATION_RULES.keys())):
        rules = ALL_CAUSATION_RULES[key]
        kpi = ALL_KPIS.get(key, {})
        row = [
            key,
            kpi.get("name", key),
            get_domain(key),
            "; ".join(rules.get("root_causes", [])),
            ", ".join(rules.get("downstream_impact", [])),
            ", ".join(reverse_feeds.get(key, [])),
            "; ".join(rules.get("corrective_actions", [])),
        ]
        ws6.append(row)
        if i % 2 == 1:
            for col in range(1, len(headers6) + 1):
                ws6.cell(row=i + 2, column=col).fill = alt_fill
    auto_width(ws6)

    # ── Sheet 7: Tab Usage Matrix ──────────────────────────────────────────

    ws7 = wb.create_sheet("Tab Usage")
    tab_names = sorted(TAB_USAGE.keys())
    headers7 = ["KPI Key", "KPI Name"] + tab_names + ["Total Tabs"]
    ws7.append(headers7)
    style_header(ws7, len(headers7))

    check_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    for i, key in enumerate(sorted_keys):
        kpi = ALL_KPIS[key]
        row = [key, kpi.get("name", key)]
        count = 0
        for tab in tab_names:
            present = key in TAB_USAGE[tab]
            row.append("✓" if present else "")
            count += 1 if present else 0
        row.append(count)
        ws7.append(row)
        # Highlight check marks
        for j, tab in enumerate(tab_names):
            if key in TAB_USAGE[tab]:
                ws7.cell(row=i + 2, column=j + 3).fill = check_fill
        if i % 2 == 1:
            for col in [1, 2, len(headers7)]:
                ws7.cell(row=i + 2, column=col).fill = alt_fill
    auto_width(ws7, max_w=20)

    # ── Sheet 8: Domain Urgency Reference ──────────────────────────────────

    ws8 = wb.create_sheet("Reference")
    ws8.append(["Domain Urgency Scores"])
    ws8.cell(row=1, column=1).font = Font(bold=True, size=12, color="0055A4")
    ws8.append(["Domain", "Urgency Score (0-100)", "Tier"])
    style_header(ws8, 3)
    tiers = {100: "Existential", 95: "Existential", 80: "Revenue Engine",
             75: "Revenue Engine", 70: "Retention", 55: "Profitability", 45: "Efficiency"}
    for domain, score in sorted(DOMAIN_URGENCY.items(), key=lambda x: -x[1]):
        ws8.append([domain.title(), score, tiers.get(score, "—")])

    ws8.append([])
    ws8.append(["Composite Criticality Weights (Default)"])
    ws8.cell(row=ws8.max_row, column=1).font = Font(bold=True, size=12, color="0055A4")
    ws8.append(["Component", "Default Weight", "Description"])
    style_header(ws8, 3)
    descs = {
        "gap": "How far from target (normalised distance)",
        "trend": "Rate of deterioration (OLS slope over 6 months)",
        "impact": "Downstream causal reach in KPI graph (BFS)",
        "domain": "Business area survival tier urgency",
    }
    for comp, w in CRIT_DEFAULT_WEIGHTS.items():
        ws8.append([comp.title(), f"{w*100:.0f}%", descs.get(comp, "")])

    ws8.append([])
    ws8.append(["Health Score Weights (Default)"])
    ws8.cell(row=ws8.max_row, column=1).font = Font(bold=True, size=12, color="0055A4")
    ws8.append(["Component", "Default Weight", "Description"])
    style_header(ws8, 3)
    ws8.append(["Momentum", "30%", "% of KPIs with improving 6-month OLS trend"])
    ws8.append(["Target Achievement", "40%", "% of KPIs meeting their target"])
    ws8.append(["Risk", "30%", "Inverse of % KPIs in red status (>8% miss)"])

    ws8.append([])
    ws8.append([f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"])
    ws8.append([f"Workspace: {workspace_id}"])
    ws8.append([f"Total KPIs Tracked: {len(ALL_KPIS)}"])
    ws8.append([f"KPIs with Data: {len(monthly)}"])

    auto_width(ws8, max_w=50)

    # ── Write to buffer and return ─────────────────────────────────────────

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"axiom-kpi-audit-{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


# ─── Integration Specification Export ──────────────────────────────────────────

@router.get("/api/export/integration-spec.xlsx", tags=["Export"])
def export_integration_spec():
    """Download the Integration Specification workbook — field-level data
    requirements for every supported source system, canonical schema mappings,
    KPI reference, and ELT pipeline documentation."""
    from core.integration_spec import generate_integration_spec_workbook

    wb = generate_integration_spec_workbook()
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"axiom-integration-spec-{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


# ─── Financial Model Export ──────────────────────────────────────────────────

@router.get("/api/export/financial-model.xlsx", tags=["Export"])
def export_financial_model(request: Request):
    """
    Generate a structured Excel financial model with live formulas,
    editable Assumptions sheet, P&L, scenarios, and confidence bands.
    """
    from core.excel_model import ModelWorkbookBuilder
    from core.database import _audit

    workspace_id = _require_workspace(request)
    conn = get_db()

    # ── Gather monthly data ──────────────────────────────────────────────
    rows = conn.execute(
        "SELECT year, month, data_json FROM monthly_data WHERE workspace_id=? "
        "ORDER BY year ASC, month ASC",
        [workspace_id],
    ).fetchall()

    monthly_data = []
    for r in rows:
        try:
            kpis = json.loads(r["data_json"])
            # Strip metadata keys
            clean = {k: v for k, v in kpis.items() if not k.startswith("_") and v is not None}
            monthly_data.append((int(r["year"]), int(r["month"]), clean))
        except (json.JSONDecodeError, TypeError):
            continue

    # ── Gather forecast data ─────────────────────────────────────────────
    forecast_data = None
    model_row = conn.execute(
        "SELECT * FROM markov_models WHERE workspace_id=? ORDER BY id DESC LIMIT 1",
        [workspace_id],
    ).fetchone()

    if model_row:
        # Check for a recent forecast run
        run_row = conn.execute(
            "SELECT trajectories, causal_paths FROM forecast_runs WHERE model_id=? ORDER BY id DESC LIMIT 1",
            [model_row["id"]],
        ).fetchone()
        if run_row and run_row["trajectories"]:
            try:
                forecast_data = {
                    "trajectories": json.loads(run_row["trajectories"]),
                    "kpis": json.loads(model_row["kpis"]),
                }
            except (json.JSONDecodeError, TypeError):
                pass

    # ── Gather scenarios ─────────────────────────────────────────────────
    scenario_rows = conn.execute(
        "SELECT name, levers_json FROM saved_scenarios WHERE workspace_id=? ORDER BY created_at DESC LIMIT 3",
        [workspace_id],
    ).fetchall()
    scenarios = [{"name": s["name"], "levers_json": s["levers_json"]} for s in scenario_rows]

    # ── Gather settings ──────────────────────────────────────────────────
    settings_rows = conn.execute(
        "SELECT key, value FROM company_settings WHERE workspace_id=?",
        [workspace_id],
    ).fetchall()
    settings = {r["key"]: r["value"] for r in settings_rows}

    # ── Gather targets ───────────────────────────────────────────────────
    target_rows = conn.execute(
        "SELECT kpi_key, target_value, unit, direction FROM kpi_targets WHERE workspace_id=?",
        [workspace_id],
    ).fetchall()
    targets = {
        r["kpi_key"]: {"target_value": r["target_value"], "unit": r["unit"], "direction": r["direction"]}
        for r in target_rows
    }

    # ── Gather projections ─────────────────────────────────────────────
    proj_rows = conn.execute(
        "SELECT year, month, data_json FROM projection_monthly_data WHERE workspace_id=? ORDER BY year, month",
        [workspace_id],
    ).fetchall()
    projection_data = []
    for pr in proj_rows:
        pd_json = json.loads(pr["data_json"]) if isinstance(pr["data_json"], str) else (pr["data_json"] or {})
        projection_data.append((int(pr["year"]), int(pr["month"]), pd_json))

    conn.close()

    # ── Build the workbook ───────────────────────────────────────────────
    builder = ModelWorkbookBuilder(monthly_data, forecast_data, scenarios, settings, targets, projection_data)
    wb = builder.build()

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    version = datetime.utcnow().strftime('%Y%m%d%H%M%S')
    fname = f"axiom-financial-model-{version}.xlsx"

    # ── Build assumption snapshot for Phase 3 diff engine ──────────────
    _snapshot_keys = (
        "revenue", "cogs", "opex", "headcount", "customers", "mrr", "arr",
        "ar", "sm_allocated", "revenue_growth", "gross_margin", "operating_margin",
        "ebitda_margin", "nrr", "churn_rate", "burn_multiple", "dso",
        "revenue_growth_rate", "gross_margin_target",
    )
    assumption_snapshot = {}
    if monthly_data:
        latest_kpis = monthly_data[-1][2]
        for k in _snapshot_keys:
            v = latest_kpis.get(k)
            if v is not None:
                try:
                    assumption_snapshot[k] = round(float(v), 6)
                except (ValueError, TypeError):
                    pass
    # Include any settings-based assumptions
    for sk in ("revenue_growth_rate", "gross_margin_target"):
        sv = settings.get(sk)
        if sv is not None and sk not in assumption_snapshot:
            try:
                assumption_snapshot[sk] = round(float(sv), 6)
            except (ValueError, TypeError):
                pass

    # ── Record export version ────────────────────────────────────────────
    conn = get_db()
    try:
        conn.execute(
            "INSERT INTO model_exports (version, filename, assumption_snapshot, months_of_data, "
            "forecast_kpis, scenarios_included, workspace_id) VALUES (?,?,?,?,?,?,?)",
            (version, fname, json.dumps(assumption_snapshot) if assumption_snapshot else None,
             len(monthly_data),
             len(forecast_data["kpis"]) if forecast_data else 0,
             len(scenarios), workspace_id),
        )
        conn.commit()
    except Exception:
        pass
    finally:
        conn.close()

    _audit("export_financial_model", "export", fname,
           f"Financial model exported: {len(monthly_data)} months, {len(scenarios)} scenarios",
           workspace_id=workspace_id)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


# ─── Financial Model Import + Diff Engine ────────────────────────────────────

@router.post("/api/import/financial-model", tags=["Import"])
async def import_financial_model(request: Request, file: UploadFile = File(...)):
    """
    Parse an uploaded financial model .xlsx, detect changed assumptions
    vs. the last export snapshot, and return a diff report.

    Does NOT apply changes — the user must confirm via /apply.
    """
    import openpyxl

    workspace_id = _require_workspace(request)

    fname = (file.filename or "").lower()
    if not fname.endswith(".xlsx"):
        raise HTTPException(400, "Please upload an .xlsx file exported from the platform.")

    contents = await file.read()
    if len(contents) > 10 * 1024 * 1024:
        raise HTTPException(413, "File too large. Maximum 10 MB.")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=False)
    except Exception as e:
        raise HTTPException(400, f"Could not open Excel file: {e}")

    # Read Assumptions sheet
    if "Assumptions" not in wb.sheetnames:
        raise HTTPException(400,
            "Sheet 'Assumptions' not found. Please upload a file exported from 'Export Financial Model'.")

    ws = wb["Assumptions"]
    imported_values = {}

    # Parse Assumptions layout: Row 4 = headers, Row 5+ = parameter rows
    # Column 1 = Label, Column 2 = Unit, Column 3 = Description, Column 4+ = monthly values
    from core.excel_model import ASSUMPTION_PARAMS

    param_labels = {label: key for key, label, _, _, _ in ASSUMPTION_PARAMS}
    # Map ASSUMPTION_PARAMS keys back to source KPI keys for snapshot comparison
    _param_to_kpi = {}
    for key, label, _, _, source_kpi in ASSUMPTION_PARAMS:
        _param_to_kpi[key] = source_kpi or key  # e.g., "revenue_growth_rate" -> "revenue_growth"

    for r in range(5, ws.max_row + 1):
        label = ws.cell(row=r, column=1).value
        if not label:
            continue
        param_key = param_labels.get(str(label).strip())
        if not param_key:
            continue

        # Read the last monthly value (rightmost non-empty data column)
        last_val = None
        for c in range(ws.max_column, 3, -1):
            cell_val = ws.cell(row=r, column=c).value
            if cell_val is not None and not str(cell_val).startswith("="):
                try:
                    last_val = round(float(cell_val), 6)
                except (ValueError, TypeError):
                    pass
                break
        if last_val is not None:
            imported_values[param_key] = last_val

    # Find the latest export snapshot for this workspace
    conn = get_db()
    export_row = conn.execute(
        "SELECT version, assumption_snapshot, exported_at FROM model_exports "
        "WHERE workspace_id=? ORDER BY id DESC LIMIT 1",
        [workspace_id],
    ).fetchone()
    conn.close()

    if not export_row or not export_row["assumption_snapshot"]:
        return {
            "status": "no_baseline",
            "message": "No previous export found. Cannot compute diff. The imported values are shown below.",
            "imported_values": imported_values,
            "changes": [],
            "scenario_mapping": {},
        }

    baseline = json.loads(export_row["assumption_snapshot"])
    version = export_row["version"]
    exported_at = export_row["exported_at"]

    # Compute diff (map ASSUMPTION_PARAMS keys to KPI keys for snapshot lookup)
    changes = []
    for key, imported_val in imported_values.items():
        kpi_key = _param_to_kpi.get(key, key)
        baseline_val = baseline.get(kpi_key) or baseline.get(key)
        if baseline_val is None:
            continue
        try:
            delta = round(imported_val - float(baseline_val), 6)
        except (ValueError, TypeError):
            continue
        if abs(delta) > 0.001:
            changes.append({
                "param": key,
                "exported": float(baseline_val),
                "imported": imported_val,
                "delta": delta,
            })

    # Map changes to scenario levers
    _PARAM_TO_LEVER = {
        "revenue_growth": "revenue_growth",
        "revenue_growth_rate": "revenue_growth",
        "gross_margin": "gross_margin_adj",
        "gross_margin_target": "gross_margin_adj",
        "churn_rate": "churn_adj",
        "headcount": "headcount_delta",
        "opex": "cost_reduction",
        "opex_ratio": "cost_reduction",
        "nrr": "expansion_adj",
    }
    scenario_mapping = {}
    for ch in changes:
        lever = _PARAM_TO_LEVER.get(ch["param"])
        if lever:
            # Use delta as the lever value (in pp or absolute units)
            scenario_mapping[lever] = round(ch["delta"], 4)

    from core.database import _audit
    _audit("import_financial_model", "import", file.filename or "unknown",
           f"Financial model imported: {len(changes)} changes detected vs export {version}",
           workspace_id=workspace_id)

    # ── Parse Projections sheet if it exists (Feature 5) ───────────────
    projection_changes = []
    if "Projections" in wb.sheetnames:
        try:
            proj_ws = wb["Projections"]
            # Parse: Column A = KPI key, Row 1 = month headers (YYYY-MM), data in grid
            proj_months = []
            for c in range(2, proj_ws.max_column + 1):
                header = str(proj_ws.cell(row=1, column=c).value or "")
                if header and "-" in header:
                    try:
                        parts = header.split("-")
                        yr, mo = int(parts[0]), int(parts[1])
                        proj_months.append((yr, mo, c))
                    except (ValueError, IndexError):
                        pass

            imported_projections = {}
            for r in range(2, proj_ws.max_row + 1):
                kpi = str(proj_ws.cell(row=r, column=1).value or "").strip()
                if not kpi:
                    continue
                for yr, mo, col in proj_months:
                    val = proj_ws.cell(row=r, column=col).value
                    if val is not None and isinstance(val, (int, float)):
                        imported_projections.setdefault((yr, mo), {})[kpi] = round(float(val), 4)

            # Diff against stored projections
            conn3 = get_db()
            try:
                stored_rows = conn3.execute(
                    "SELECT year, month, data_json FROM projection_monthly_data WHERE workspace_id=? ORDER BY year, month",
                    [workspace_id],
                ).fetchall()
                stored_proj = {}
                for sr in stored_rows:
                    d = json.loads(sr["data_json"]) if isinstance(sr["data_json"], str) else (sr["data_json"] or {})
                    stored_proj[(sr["year"], sr["month"])] = d
            finally:
                conn3.close()

            for (yr, mo), kpis in imported_projections.items():
                stored = stored_proj.get((yr, mo), {})
                for kpi, new_val in kpis.items():
                    old_val = stored.get(kpi)
                    if old_val is None or abs(new_val - (old_val or 0)) > 0.01:
                        projection_changes.append({
                            "period": f"{yr}-{mo:02d}", "kpi": kpi,
                            "old_value": round(float(old_val), 4) if old_val else None,
                            "new_value": new_val,
                            "delta": round(new_val - (old_val or 0), 4),
                        })
        except Exception:
            pass  # Projection parsing is optional — don't block assumption diff

    return {
        "status": "diff_computed",
        "export_version": version,
        "exported_at": exported_at,
        "imported_values": imported_values,
        "changes": changes,
        "scenario_mapping": scenario_mapping,
        "projection_changes": projection_changes,
        "message": f"{len(changes)} assumption(s) and {len(projection_changes)} projection value(s) changed." if (changes or projection_changes)
                   else "No changes detected since last export.",
    }


@router.post("/api/import/financial-model/apply", tags=["Import"])
async def apply_financial_model_changes(request: Request):
    """
    Apply imported assumption changes by creating a new saved scenario
    and optionally triggering a forecast re-run.
    """
    workspace_id = _require_workspace(request)
    body = await request.json()

    scenario_mapping = body.get("scenario_mapping", {})
    projection_changes = body.get("projection_changes", [])
    scenario_name = body.get("scenario_name", "Imported from Excel")
    run_forecast = body.get("run_forecast", False)

    if not scenario_mapping and not projection_changes:
        raise HTTPException(400, "No changes to apply.")

    # Apply projection changes to projection_monthly_data
    if projection_changes:
        conn_proj = get_db()
        try:
            # Group changes by (year, month)
            by_period = {}
            for ch in projection_changes:
                parts = ch["period"].split("-")
                yr, mo = int(parts[0]), int(parts[1])
                by_period.setdefault((yr, mo), {})[ch["kpi"]] = ch["new_value"]

            for (yr, mo), kpis in by_period.items():
                # Load existing, merge, save
                existing_row = conn_proj.execute(
                    "SELECT id, data_json FROM projection_monthly_data WHERE workspace_id=? AND year=? AND month=?",
                    [workspace_id, yr, mo],
                ).fetchone()
                if existing_row:
                    d = json.loads(existing_row["data_json"]) if isinstance(existing_row["data_json"], str) else (existing_row["data_json"] or {})
                    d.update(kpis)
                    conn_proj.execute(
                        "UPDATE projection_monthly_data SET data_json=? WHERE id=?",
                        [json.dumps(d), existing_row["id"]],
                    )
                else:
                    conn_proj.execute(
                        "INSERT INTO projection_monthly_data (projection_upload_id, year, month, data_json, version_label, workspace_id) "
                        "VALUES (?,?,?,?,?,?)",
                        [-1, yr, mo, json.dumps(kpis), "reimport", workspace_id],
                    )
            conn_proj.commit()
            from core.database import _audit
            _audit("projection_reimport", "projection", "reimport",
                   f"Projection re-import: {len(projection_changes)} values updated across {len(by_period)} months",
                   workspace_id=workspace_id)
        finally:
            conn_proj.close()

    # Create a saved scenario
    conn = get_db()
    cur = conn.execute(
        "INSERT INTO saved_scenarios (workspace_id, name, levers_json, notes) VALUES (?,?,?,?)",
        (workspace_id, scenario_name, json.dumps(scenario_mapping),
         f"Auto-created from re-imported financial model. Changes: {json.dumps(scenario_mapping)}"),
    )
    scenario_id = cur.lastrowid
    conn.commit()
    conn.close()

    result = {
        "status": "applied",
        "scenario_id": scenario_id,
        "scenario_name": scenario_name,
        "levers_applied": scenario_mapping,
    }

    # Optionally run forecast with the imported overrides
    if run_forecast:
        try:
            from routers.scenarios import _LEVER_TO_KPI
            from routers.forecast import _project_scenario

            conn = get_db()
            model_row = conn.execute(
                "SELECT current_states, thresholds FROM markov_models "
                "WHERE workspace_id=? ORDER BY id DESC LIMIT 1",
                [workspace_id],
            ).fetchone()
            conn.close()

            if model_row:
                current_values = json.loads(model_row["current_states"])
                value_ranges = json.loads(model_row["thresholds"])

                overrides = {}
                for lever_id, lever_value in scenario_mapping.items():
                    target_kpi = _LEVER_TO_KPI.get(lever_id)
                    if not target_kpi or target_kpi not in value_ranges:
                        continue
                    current = current_values.get(target_kpi, 0)
                    target = current + float(lever_value)
                    vr = value_ranges[target_kpi]
                    percentiles = [vr.get("p10", current), vr.get("p25", current),
                                   vr.get("p50", current), vr.get("p75", current),
                                   vr.get("p90", current)]
                    best_idx = min(range(5), key=lambda i: abs(target - percentiles[i]))
                    overrides[target_kpi] = best_idx

                fc_result = _project_scenario(90, overrides, 400, workspace_id)
                if isinstance(fc_result, dict) and "trajectories" in fc_result:
                    result["forecast"] = {
                        "trajectories": fc_result["trajectories"],
                        "overrides_applied": overrides,
                    }
        except Exception:
            pass  # Forecast is optional; don't fail the apply

    from core.database import _audit
    _audit("apply_imported_model", "scenario", scenario_name,
           f"Applied imported model changes: {len(scenario_mapping)} levers",
           workspace_id=workspace_id)

    return result


# ─── SEC-Grade Analytics Endpoints ────────────────────────────────────────────


@router.get("/api/analytics/arr-bridge", tags=["Analytics"])
async def arr_bridge(request: Request, from_year: int = 0, from_month: int = 1,
                     to_year: int = 0, to_month: int = 12):
    """ARR Bridge / Net New ARR Waterfall — the most important SaaS board chart.

    Decomposes ARR movement into: New, Expansion, Contraction, Churned.
    """
    workspace_id = _get_workspace(request)
    conn = get_db()
    try:
        # Build period filter
        sql = ("SELECT customer_id, period, amount, subscription_type "
               "FROM canonical_revenue WHERE workspace_id=? AND amount IS NOT NULL")
        args: list = [workspace_id]
        if from_year > 0:
            sql += " AND period >= ?"
            args.append(f"{from_year}-{from_month:02d}")
        if to_year > 0:
            sql += " AND period <= ?"
            args.append(f"{to_year}-{to_month:02d}")
        sql += " ORDER BY period"
        rows = conn.execute(sql, args).fetchall()
    except Exception:
        rows = []
    conn.close()

    if not rows:
        return {"periods": [], "summary": {}}

    # Build per-customer per-period revenue
    customer_period: dict = {}  # {period: {customer_id: total_amount}}
    for r in rows:
        cid = r[0] or "unknown"
        period = str(r[1] or "")[:7]  # YYYY-MM
        amt = float(r[2] or 0)
        if period not in customer_period:
            customer_period[period] = {}
        customer_period[period][cid] = customer_period[period].get(cid, 0) + amt

    sorted_periods = sorted(customer_period.keys())
    bridge_periods = []

    for i, period in enumerate(sorted_periods):
        current = customer_period[period]
        prev = customer_period[sorted_periods[i - 1]] if i > 0 else {}
        prev_total = sum(prev.values())
        curr_total = sum(current.values())

        new_arr = sum(v for k, v in current.items() if k not in prev) * 12
        churned_arr = sum(v for k, v in prev.items() if k not in current) * 12
        expansion_arr = sum(
            max(0, current.get(k, 0) - prev.get(k, 0))
            for k in current if k in prev and current[k] > prev[k]
        ) * 12
        contraction_arr = sum(
            max(0, prev.get(k, 0) - current.get(k, 0))
            for k in current if k in prev and current[k] < prev[k]
        ) * 12

        beginning_arr = prev_total * 12
        ending_arr = curr_total * 12

        bridge_periods.append({
            "period": period,
            "beginning_arr": round(beginning_arr, 2),
            "new_arr": round(new_arr, 2),
            "expansion_arr": round(expansion_arr, 2),
            "contraction_arr": round(contraction_arr, 2),
            "churned_arr": round(churned_arr, 2),
            "ending_arr": round(ending_arr, 2),
            "net_new_arr": round(new_arr + expansion_arr - contraction_arr - churned_arr, 2),
        })

    total_new = sum(p["new_arr"] for p in bridge_periods)
    total_exp = sum(p["expansion_arr"] for p in bridge_periods)
    total_con = sum(p["contraction_arr"] for p in bridge_periods)
    total_churn = sum(p["churned_arr"] for p in bridge_periods)

    return {
        "periods": bridge_periods,
        "summary": {
            "total_new_arr": round(total_new, 2),
            "total_expansion_arr": round(total_exp, 2),
            "total_contraction_arr": round(total_con, 2),
            "total_churned_arr": round(total_churn, 2),
            "net_change": round(total_new + total_exp - total_con - total_churn, 2),
        },
    }


@router.get("/api/analytics/cohort-retention", tags=["Analytics"])
async def cohort_retention(request: Request, metric: str = "revenue",
                           from_year: int = 0, from_month: int = 1,
                           to_year: int = 0, to_month: int = 12):
    """Cohort Revenue/Count Retention Matrix — vintage analysis."""
    workspace_id = _get_workspace(request)
    conn = get_db()
    try:
        sql = ("SELECT customer_id, period, amount FROM canonical_revenue "
               "WHERE workspace_id=? AND customer_id IS NOT NULL AND amount IS NOT NULL")
        args: list = [workspace_id]
        if from_year > 0:
            sql += " AND period >= ?"
            args.append(f"{from_year}-{from_month:02d}")
        if to_year > 0:
            sql += " AND period <= ?"
            args.append(f"{to_year}-{to_month:02d}")
        rev_rows = conn.execute(sql, args).fetchall()
    except Exception:
        rev_rows = []
    conn.close()

    if not rev_rows:
        return {"cohorts": []}

    # Determine each customer's first period (acquisition month)
    first_period: dict = {}  # customer_id -> first period
    customer_period_data: dict = {}  # (customer_id, period) -> amount
    for r in rev_rows:
        cid, period, amt = r[0], str(r[1] or "")[:7], float(r[2] or 0)
        if cid not in first_period or period < first_period[cid]:
            first_period[cid] = period
        key = (cid, period)
        customer_period_data[key] = customer_period_data.get(key, 0) + amt

    # Group customers by acquisition cohort
    cohort_customers: dict = {}  # acquisition_period -> set of customer_ids
    for cid, fp in first_period.items():
        cohort_customers.setdefault(fp, set()).add(cid)

    all_periods = sorted(set(p for _, p in customer_period_data.keys()))

    cohorts = []
    for acq_period in sorted(cohort_customers.keys()):
        customers = cohort_customers[acq_period]
        # Month 0 data
        m0_revenue = sum(customer_period_data.get((c, acq_period), 0) for c in customers)
        m0_count = sum(1 for c in customers if customer_period_data.get((c, acq_period), 0) > 0)

        months = []
        acq_idx = all_periods.index(acq_period) if acq_period in all_periods else -1
        if acq_idx < 0:
            continue

        for offset in range(len(all_periods) - acq_idx):
            p = all_periods[acq_idx + offset]
            if metric == "count":
                active = sum(1 for c in customers if customer_period_data.get((c, p), 0) > 0)
                retention = (active / m0_count * 100) if m0_count else 0
                months.append({"month_offset": offset, "period": p, "value": active, "retention_pct": round(retention, 1)})
            else:
                revenue = sum(customer_period_data.get((c, p), 0) for c in customers)
                retention = (revenue / m0_revenue * 100) if m0_revenue else 0
                months.append({"month_offset": offset, "period": p, "value": round(revenue, 2), "retention_pct": round(retention, 1)})

        cohorts.append({
            "acquisition_period": acq_period,
            "size": len(customers),
            "month_0_value": round(m0_revenue if metric == "revenue" else m0_count, 2),
            "months": months[:24],  # Cap at 24 months for display
        })

    return {"cohorts": cohorts, "metric": metric}


@router.get("/api/analytics/customer-concentration", tags=["Analytics"])
async def customer_concentration(request: Request, top_n: int = 10, period: str = "latest",
                                  from_year: int = 0, from_month: int = 1,
                                  to_year: int = 0, to_month: int = 12):
    """Top-N Customer Concentration — SEC disclosure trigger at 10%."""
    workspace_id = _get_workspace(request)
    conn = get_db()

    # Build period filter
    _pf = ""
    _pp = []
    if from_year > 0 and to_year > 0:
        _pf = " AND period >= ? AND period <= ?"
        _pp = [f"{from_year}-{from_month:02d}", f"{to_year}-{to_month:02d}"]
        period = f"{from_year}-{from_month:02d} to {to_year}-{to_month:02d}"
    elif period == "latest":
        p_row = conn.execute(
            "SELECT MAX(period) as p FROM canonical_revenue WHERE workspace_id=?",
            [workspace_id],
        ).fetchone()
        period = p_row[0] if p_row else None
        if period:
            _pf = " AND period = ?"
            _pp = [period]

    if not period:
        conn.close()
        return {"period": None, "customers": [], "total_revenue": 0}

    try:
        rows = conn.execute(
            "SELECT customer_id, SUM(amount) as total "
            "FROM canonical_revenue WHERE workspace_id=? "
            "AND customer_id IS NOT NULL AND amount IS NOT NULL"
            f"{_pf} GROUP BY customer_id ORDER BY total DESC",
            [workspace_id] + _pp,
        ).fetchall()

        name_rows = conn.execute(
            "SELECT source_id, name FROM canonical_customers WHERE workspace_id=?",
            [workspace_id],
        ).fetchall()
    except Exception:
        rows, name_rows = [], []
    conn.close()

    names = {r[0]: r[1] for r in name_rows}
    total_revenue = sum(float(r[1] or 0) for r in rows)
    customers = []
    running_pct = 0

    for i, r in enumerate(rows[:top_n]):
        cid = r[0]
        rev = float(r[1] or 0)
        pct = (rev / total_revenue * 100) if total_revenue else 0
        running_pct += pct
        customers.append({
            "rank": i + 1,
            "customer_id": cid,
            "customer_name": names.get(cid, cid),
            "revenue": round(rev, 2),
            "pct_of_total": round(pct, 1),
        })

    # HHI Index (sum of squared market shares)
    all_shares = [(float(r[1] or 0) / total_revenue * 100) if total_revenue else 0 for r in rows]
    hhi = sum(s ** 2 for s in all_shares)

    top_counts = [1, 5, 10, 20]
    top_pcts = {}
    for n in top_counts:
        top_pcts[f"top_{n}_pct"] = round(sum(
            (float(r[1] or 0) / total_revenue * 100) if total_revenue else 0
            for r in rows[:n]
        ), 1)

    return {
        "period": period,
        "total_revenue": round(total_revenue, 2),
        "customers": customers,
        "hhi_index": round(hhi, 1),
        **top_pcts,
        "sec_threshold_breached": any(c["pct_of_total"] > 10 for c in customers),
    }


@router.get("/api/analytics/margin-decomposition", tags=["Analytics"])
async def margin_decomposition(request: Request, from_year: int = 0, from_month: int = 1,
                                to_year: int = 0, to_month: int = 12):
    """Gross Margin Decomposition — breaks down COGS components over time."""
    workspace_id = _get_workspace(request)
    conn = get_db()

    _COGS_KEYWORDS = {"cogs", "cost of goods", "cost of revenue", "hosting",
                      "infrastructure", "direct", "support", "implementation"}

    # Build period filter clause
    _pf = ""
    _pa: list = []
    if from_year > 0:
        _pf += " AND period >= ?"
        _pa.append(f"{from_year}-{from_month:02d}")
    if to_year > 0:
        _pf += " AND period <= ?"
        _pa.append(f"{to_year}-{to_month:02d}")

    try:
        exp_rows = conn.execute(
            "SELECT period, category, SUM(amount) as total "
            "FROM canonical_expenses WHERE workspace_id=? AND amount IS NOT NULL"
            + _pf + " GROUP BY period, category ORDER BY period",
            [workspace_id] + _pa,
        ).fetchall()
        rev_rows = conn.execute(
            "SELECT period, SUM(amount) as total "
            "FROM canonical_revenue WHERE workspace_id=? AND amount IS NOT NULL"
            + _pf + " GROUP BY period ORDER BY period",
            [workspace_id] + _pa,
        ).fetchall()
    except Exception:
        exp_rows, rev_rows = [], []
    conn.close()

    revenue_by_period = {str(r[0])[:7]: float(r[1] or 0) for r in rev_rows}

    # Group expenses by period and category
    period_costs: dict = {}
    for r in exp_rows:
        period = str(r[0])[:7]
        cat = str(r[1] or "other").lower()
        amt = float(r[2] or 0)
        is_cogs = any(kw in cat for kw in _COGS_KEYWORDS)
        if not is_cogs:
            continue
        if period not in period_costs:
            period_costs[period] = {}
        # Normalize category names
        display_cat = cat
        if any(kw in cat for kw in ("hosting", "infrastructure")):
            display_cat = "Hosting & Infrastructure"
        elif "support" in cat:
            display_cat = "Support & CS"
        elif any(kw in cat for kw in ("cogs", "cost of goods", "cost of revenue", "direct")):
            display_cat = "Direct COGS"
        else:
            display_cat = cat.title()
        period_costs[period][display_cat] = period_costs[period].get(display_cat, 0) + amt

    periods = []
    for period in sorted(set(list(revenue_by_period.keys()) + list(period_costs.keys()))):
        rev = revenue_by_period.get(period, 0)
        costs = period_costs.get(period, {})
        total_cogs = sum(costs.values())
        gm_pct = ((rev - total_cogs) / rev * 100) if rev else 0
        periods.append({
            "period": period,
            "revenue": round(rev, 2),
            "total_cogs": round(total_cogs, 2),
            "cogs_components": [{"category": k, "amount": round(v, 2), "pct_of_revenue": round(v / rev * 100, 1) if rev else 0} for k, v in sorted(costs.items())],
            "gross_margin_pct": round(gm_pct, 1),
        })

    return {"periods": periods}


@router.get("/api/analytics/cash-waterfall", tags=["Analytics"])
async def cash_waterfall(request: Request, from_year: int = 0, from_month: int = 1,
                         to_year: int = 0, to_month: int = 12):
    """Cash Burn Waterfall — Opening Cash → Inflows → Outflows → Closing Cash."""
    workspace_id = _get_workspace(request)
    conn = get_db()

    _SM_KEYWORDS = {"sales", "marketing", "s&m", "advertising", "demand", "paid", "seo"}
    _RD_KEYWORDS = {"r&d", "research", "engineering", "development", "product"}
    _GA_KEYWORDS = {"g&a", "general", "admin", "office", "legal", "hr", "finance"}

    # Period filter
    _pf = ""
    _pp = []
    if from_year > 0 and to_year > 0:
        _pf = " AND period >= ? AND period <= ?"
        _pp = [f"{from_year}-{from_month:02d}", f"{to_year}-{to_month:02d}"]

    try:
        bs_rows = conn.execute(
            "SELECT period, cash_balance FROM canonical_balance_sheet "
            f"WHERE workspace_id=?{_pf} ORDER BY period",
            [workspace_id] + _pp,
        ).fetchall()
        rev_rows = conn.execute(
            "SELECT period, SUM(amount) as total FROM canonical_revenue "
            f"WHERE workspace_id=? AND amount IS NOT NULL{_pf} GROUP BY period",
            [workspace_id] + _pp,
        ).fetchall()
        exp_rows = conn.execute(
            "SELECT period, category, SUM(amount) as total FROM canonical_expenses "
            f"WHERE workspace_id=? AND amount IS NOT NULL{_pf} GROUP BY period, category",
            [workspace_id] + _pp,
        ).fetchall()
    except Exception:
        bs_rows, rev_rows, exp_rows = [], [], []
    conn.close()

    cash_by_period = {str(r[0])[:7]: float(r[1] or 0) for r in bs_rows}
    rev_by_period = {str(r[0])[:7]: float(r[1] or 0) for r in rev_rows}

    # Categorize expenses
    exp_by_period: dict = {}
    for r in exp_rows:
        period = str(r[0])[:7]
        cat = str(r[1] or "other").lower()
        amt = float(r[2] or 0)
        if period not in exp_by_period:
            exp_by_period[period] = {}
        if any(kw in cat for kw in _SM_KEYWORDS):
            bucket = "S&M"
        elif any(kw in cat for kw in _RD_KEYWORDS):
            bucket = "R&D"
        elif any(kw in cat for kw in _GA_KEYWORDS):
            bucket = "G&A"
        elif any(kw in cat for kw in ("cogs", "hosting", "infrastructure", "direct")):
            bucket = "COGS"
        else:
            bucket = "Other"
        exp_by_period[period][bucket] = exp_by_period[period].get(bucket, 0) + amt

    all_periods = sorted(set(list(cash_by_period.keys()) + list(rev_by_period.keys())))
    periods = []
    for i, period in enumerate(all_periods):
        opening = cash_by_period.get(all_periods[i - 1], 0) if i > 0 else cash_by_period.get(period, 0)
        revenue = rev_by_period.get(period, 0)
        expenses = exp_by_period.get(period, {})
        total_exp = sum(expenses.values())
        closing = cash_by_period.get(period, opening + revenue - total_exp)
        net_burn = revenue - total_exp
        runway = abs(closing / net_burn) if net_burn < 0 else 999

        periods.append({
            "period": period,
            "opening_cash": round(opening, 2),
            "revenue_inflow": round(revenue, 2),
            "expense_categories": [{"category": k, "amount": round(v, 2)} for k, v in sorted(expenses.items())],
            "total_expenses": round(total_exp, 2),
            "net_burn": round(net_burn, 2),
            "closing_cash": round(closing, 2),
            "runway_months": round(min(runway, 999), 1),
        })

    return {"periods": periods}


@router.get("/api/analytics/unit-economics", tags=["Analytics"])
async def unit_economics(request: Request, from_year: int = 0, from_month: int = 1,
                         to_year: int = 0, to_month: int = 12):
    """Unit Economics Waterfall — per-customer profitability decomposition."""
    workspace_id = _get_workspace(request)
    conn = get_db()

    # Period filter
    _pf = ""
    _pp = []
    _kpi_filter = ""
    _kpi_pp = []
    if from_year > 0 and to_year > 0:
        _pf = " AND period >= ? AND period <= ?"
        _pp = [f"{from_year}-{from_month:02d}", f"{to_year}-{to_month:02d}"]
        _kpi_filter = " AND (year > ? OR (year = ? AND month >= ?)) AND (year < ? OR (year = ? AND month <= ?))"
        _kpi_pp = [from_year, from_year, from_month, to_year, to_year, to_month]

    try:
        # Get KPI data for the selected period (latest month within range)
        row = conn.execute(
            "SELECT data_json FROM monthly_data WHERE workspace_id=?"
            f"{_kpi_filter} ORDER BY year DESC, month DESC LIMIT 1",
            [workspace_id] + _kpi_pp,
        ).fetchone()
        data = json.loads(row[0]) if row else {}

        # Count active customers within period (those with revenue)
        cust_row = conn.execute(
            "SELECT COUNT(DISTINCT customer_id) as cnt FROM canonical_revenue "
            f"WHERE workspace_id=? AND customer_id IS NOT NULL AND amount > 0{_pf}",
            [workspace_id] + _pp,
        ).fetchone()
        customer_count = cust_row[0] if cust_row else 0

        # Revenue and expenses within period
        rev_row = conn.execute(
            f"SELECT SUM(amount) FROM canonical_revenue WHERE workspace_id=? AND amount IS NOT NULL{_pf}",
            [workspace_id] + _pp,
        ).fetchone()
        exp_row = conn.execute(
            f"SELECT SUM(amount) FROM canonical_expenses WHERE workspace_id=? AND amount IS NOT NULL{_pf}",
            [workspace_id] + _pp,
        ).fetchone()
        period_count_row = conn.execute(
            f"SELECT COUNT(DISTINCT period) FROM canonical_revenue WHERE workspace_id=?{_pf}",
            [workspace_id] + _pp,
        ).fetchone()
    except Exception:
        data, customer_count, rev_row, exp_row, period_count_row = {}, 0, None, None, None
    conn.close()

    total_rev = float(rev_row[0] or 0) if rev_row else 0
    total_exp = float(exp_row[0] or 0) if exp_row else 0
    months = int(period_count_row[0] or 1) if period_count_row else 1

    arpu = (total_rev / months / max(customer_count, 1))
    gross_margin_pct = float(data.get("gross_margin", 70)) / 100
    churn_rate_pct = max(float(data.get("churn_rate", 5)), 0.1) / 100
    cogs_per_customer = arpu * (1 - gross_margin_pct)
    gross_profit_per_customer = arpu * gross_margin_pct
    cac = float(data.get("cac", 0)) or (total_exp / months / max(customer_count, 1) * 0.4)  # Estimate S&M portion
    contribution = gross_profit_per_customer - (cac / max(12 / churn_rate_pct, 1))
    ltv = (arpu * gross_margin_pct) / churn_rate_pct if churn_rate_pct > 0 else 0
    ltv_cac = ltv / cac if cac > 0 else 0

    return {
        "customer_count": customer_count,
        "months_of_data": months,
        "metrics": [
            {"label": "ARPU (Monthly)", "value": round(arpu, 2), "unit": "usd"},
            {"label": "COGS per Customer", "value": round(-cogs_per_customer, 2), "unit": "usd"},
            {"label": "Gross Profit per Customer", "value": round(gross_profit_per_customer, 2), "unit": "usd"},
            {"label": "CAC (Allocated)", "value": round(-cac, 2), "unit": "usd"},
            {"label": "Net Contribution", "value": round(contribution, 2), "unit": "usd"},
        ],
        "ltv": round(ltv, 2),
        "cac": round(cac, 2),
        "ltv_cac_ratio": round(ltv_cac, 2),
    }


@router.get("/api/analytics/accountability-rollup", tags=["Analytics"])
async def accountability_rollup(request: Request):
    """KPI Owner Accountability Dashboard — grouped by owner with status rollup."""
    workspace_id = _get_workspace(request)
    conn = get_db()
    try:
        acct_rows = conn.execute(
            "SELECT kpi_key, owner, due_date, status, last_updated "
            "FROM kpi_accountability WHERE workspace_id=?",
            [workspace_id],
        ).fetchall()
        # Get latest KPI values and targets
        kpi_row = conn.execute(
            "SELECT data_json FROM monthly_data WHERE workspace_id=? "
            "ORDER BY year DESC, month DESC LIMIT 1",
            [workspace_id],
        ).fetchone()
        kpi_data = json.loads(kpi_row[0]) if kpi_row else {}

        target_rows = conn.execute(
            "SELECT kpi_key, target_value, direction FROM kpi_targets WHERE workspace_id=?",
            [workspace_id],
        ).fetchall()
    except Exception:
        acct_rows, kpi_data, target_rows = [], {}, []
    conn.close()

    targets = {r[0]: {"target": float(r[1] or 0), "direction": r[2]} for r in target_rows}
    kpi_names = {d["key"]: d["name"] for d in KPI_DEFS + EXTENDED_ONTOLOGY_METRICS}

    owners: dict = {}
    for r in acct_rows:
        kpi_key, owner, due_date, status, updated = r[0], r[1] or "Unassigned", r[2], r[3], r[4]
        if owner not in owners:
            owners[owner] = {"name": owner, "kpis": [], "green": 0, "yellow": 0, "red": 0, "grey": 0}
        val = kpi_data.get(kpi_key)
        t = targets.get(kpi_key)
        if val is not None and t:
            ratio = val / t["target"] if t["target"] else 0
            if t["direction"] == "higher":
                kpi_status = "green" if ratio >= 0.98 else "yellow" if ratio >= 0.90 else "red"
            else:
                kpi_status = "green" if ratio <= 1.02 else "yellow" if ratio <= 1.10 else "red"
        else:
            kpi_status = "grey"
        owners[owner][kpi_status] += 1
        owners[owner]["kpis"].append({
            "key": kpi_key, "name": kpi_names.get(kpi_key, kpi_key),
            "status": kpi_status, "value": val, "target": t["target"] if t else None,
            "due_date": due_date, "resolution_status": status,
        })

    result = sorted(owners.values(), key=lambda o: o["red"], reverse=True)
    for o in result:
        total = len(o["kpis"])
        resolved = sum(1 for k in o["kpis"] if k["resolution_status"] == "resolved")
        o["total_kpis"] = total
        o["resolution_rate_pct"] = round(resolved / total * 100, 1) if total else 0

    return {"owners": result}


@router.get("/api/analytics/restatement-history", tags=["Analytics"])
async def restatement_history(request: Request):
    """Restatement / Correction History — tracks significant data changes."""
    workspace_id = _get_workspace(request)
    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT table_name, source, source_id, field_name, old_value, new_value, changed_at "
            "FROM canonical_change_log WHERE workspace_id=? "
            "AND field_name IN ('amount', 'period', 'status') "
            "ORDER BY changed_at DESC LIMIT 200",
            [workspace_id],
        ).fetchall()
    except Exception:
        rows = []
    conn.close()

    restatements = []
    for r in rows:
        old_val = r[4]
        new_val = r[5]
        try:
            old_f = float(old_val or 0)
            new_f = float(new_val or 0)
            pct_change = ((new_f - old_f) / abs(old_f) * 100) if old_f != 0 else 0
        except (ValueError, TypeError):
            pct_change = 0

        # Only include significant changes (>1% for amounts)
        if r[3] == "amount" and abs(pct_change) < 1.0:
            continue

        restatements.append({
            "table": r[0],
            "source": r[1],
            "source_id": r[2],
            "field": r[3],
            "old_value": old_val,
            "new_value": new_val,
            "pct_change": round(pct_change, 1),
            "changed_at": r[6],
        })

    return {"restatements": restatements[:50]}


@router.get("/api/analytics/seasonality", tags=["Analytics"])
async def seasonality_analysis(request: Request, kpi_key: str = "revenue_growth"):
    """Seasonality Analysis — compute seasonal indices for a KPI."""
    workspace_id = _get_workspace(request)
    conn = get_db()
    try:
        rows = conn.execute(
            "SELECT year, month, data_json FROM monthly_data WHERE workspace_id=? "
            "ORDER BY year, month",
            [workspace_id],
        ).fetchall()
    except Exception:
        rows = []
    conn.close()

    month_values: dict = {m: [] for m in range(1, 13)}
    all_values = []
    raw_points = []

    for r in rows:
        data = json.loads(r[2]) if isinstance(r[2], str) else (r[2] or {})
        val = data.get(kpi_key)
        if val is not None:
            try:
                v = float(val)
                month_values[r[1]].append(v)
                all_values.append(v)
                raw_points.append({"year": r[0], "month": r[1], "period": f"{r[0]}-{r[1]:02d}", "raw": round(v, 2)})
            except (ValueError, TypeError):
                pass

    if not all_values or len(all_values) < 6:
        return {"kpi_key": kpi_key, "seasonal_indices": {}, "adjusted_values": [], "message": "Insufficient data (need 6+ months)"}

    annual_avg = sum(all_values) / len(all_values)
    if annual_avg == 0:
        return {"kpi_key": kpi_key, "seasonal_indices": {}, "adjusted_values": []}

    # Compute seasonal indices
    indices = {}
    for m in range(1, 13):
        vals = month_values[m]
        if vals:
            month_avg = sum(vals) / len(vals)
            indices[m] = round(month_avg / annual_avg, 3)
        else:
            indices[m] = 1.0

    # Compute adjusted values
    adjusted = []
    for pt in raw_points:
        idx = indices.get(pt["month"], 1.0)
        adj_val = pt["raw"] / idx if idx != 0 else pt["raw"]
        adjusted.append({**pt, "adjusted": round(adj_val, 2), "seasonal_index": idx})

    return {"kpi_key": kpi_key, "seasonal_indices": indices, "adjusted_values": adjusted}


# ─── KPI Annotations CRUD ───────────────────────────────────────────────────

