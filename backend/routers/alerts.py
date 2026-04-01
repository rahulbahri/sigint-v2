"""
routers/alerts.py — Slack alerts, email alerts, Excel export/import, and weekly briefing.
"""
import io
import json
import urllib.request as _urllib_req
from datetime import datetime
from typing import Optional

import httpx
from fastapi import APIRouter, HTTPException, Request, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel as _BaseModel

from core.config import ALLOWED_EMAILS, APP_URL, RESEND_API_KEY, RESEND_FROM
from core.database import get_db
from core.deps import _get_workspace
from core.kpi_defs import BENCHMARKS, CAUSATION_RULES

router = APIRouter()


# ─── Slack Alerts ────────────────────────────────────────────────────────────

class SlackTestRequest(_BaseModel):
    webhook_url: str


class SlackAlertRequest(_BaseModel):
    webhook_url: str
    red_kpis: list[dict]          # [{key, name, value, target, pct_off}]
    company_name: str = "Your Company"


@router.post("/api/slack/test", tags=["Alerts"])
async def slack_test(request: Request, body: SlackTestRequest):
    """Send a test Slack message to verify the webhook URL."""
    workspace_id = _get_workspace(request)
    if not workspace_id:
        raise HTTPException(status_code=401, detail="Not authenticated")
    if not body.webhook_url.startswith("https://hooks.slack.com/"):
        raise HTTPException(status_code=400, detail="Invalid webhook URL. Must be a Slack incoming webhook.")
    payload = {
        "text": "✅ *Axiom Intelligence* — Slack alerts connected successfully. You'll receive KPI threshold alerts here.",
        "username": "Axiom Intelligence",
        "icon_emoji": ":bar_chart:",
    }
    try:
        data = json.dumps(payload).encode("utf-8")
        req  = _urllib_req.Request(
            body.webhook_url,
            data=data,
            headers={"Content-Type": "application/json"},
        )
        with _urllib_req.urlopen(req, timeout=8) as resp:
            resp_text = resp.read().decode("utf-8")
        if resp_text.strip() != "ok":
            raise HTTPException(status_code=502, detail=f"Slack returned: {resp_text}")
    except Exception as exc:
        raise HTTPException(status_code=502, detail=str(exc))
    return {"status": "sent"}


@router.post("/api/slack/notify", tags=["Alerts"])
async def slack_notify(request: Request, body: SlackAlertRequest):
    """Fire a KPI alert message to Slack for a batch of red KPIs."""
    workspace_id = _get_workspace(request)
    if not workspace_id:
        raise HTTPException(status_code=401, detail="Not authenticated")
    if not body.webhook_url.startswith("https://hooks.slack.com/"):
        raise HTTPException(status_code=400, detail="Invalid webhook URL. Must be a Slack incoming webhook.")
    if not body.red_kpis:
        return {"status": "no_alerts"}

    blocks: list[dict] = [
        {
            "type": "header",
            "text": {"type": "plain_text", "text": f"🚨 KPI Alert — {body.company_name}", "emoji": True},
        },
        {
            "type": "section",
            "text": {
                "type": "mrkdwn",
                "text": f"*{len(body.red_kpis)} KPI{'s' if len(body.red_kpis) > 1 else ''} below critical threshold* — immediate attention recommended.",
            },
        },
        {"type": "divider"},
    ]

    for kpi in body.red_kpis[:8]:   # cap at 8 to avoid giant messages
        pct  = abs(kpi.get("pct_off", 0))
        val  = kpi.get("value", "–")
        tgt  = kpi.get("target", "–")
        name = kpi.get("name", kpi.get("key", "?"))
        blocks.append({
            "type": "section",
            "text": {
                "type": "mrkdwn",
                "text": (
                    f"*{name}*\n"
                    f"Current: `{val}`   Target: `{tgt}`   Off by `{pct:.0f}%`"
                ),
            },
        })

    blocks.append({
        "type": "context",
        "elements": [
            {
                "type": "mrkdwn",
                "text": "Sent by *Axiom Intelligence V2* · Open the platform for full narrative analysis",
            }
        ],
    })

    payload = {"blocks": blocks, "username": "Axiom Intelligence", "icon_emoji": ":bar_chart:"}
    try:
        data = json.dumps(payload).encode("utf-8")
        req  = _urllib_req.Request(
            body.webhook_url,
            data=data,
            headers={"Content-Type": "application/json"},
        )
        with _urllib_req.urlopen(req, timeout=8) as resp:
            resp_text = resp.read().decode("utf-8")
        if resp_text.strip() != "ok":
            raise HTTPException(status_code=502, detail=f"Slack returned: {resp_text}")
    except Exception as exc:
        raise HTTPException(status_code=502, detail=str(exc))

    return {"status": "sent", "alerts_fired": len(body.red_kpis)}


# ─── KPI Alert Email ──────────────────────────────────────────────────────────

async def _send_kpi_alert_email(to_email: str, alerts: list) -> bool:
    if not RESEND_API_KEY or not alerts:
        return False
    rows_html = "".join([
        f'<tr>'
        f'<td style="padding:10px 12px;border-bottom:1px solid #f1f5f9;color:#1e293b">{a.get("kpi","")}</td>'
        f'<td style="padding:10px 12px;border-bottom:1px solid #f1f5f9;color:#ef4444;font-weight:600">{a.get("status","")}</td>'
        f'<td style="padding:10px 12px;border-bottom:1px solid #f1f5f9;color:#64748b">{a.get("value","")}</td>'
        f'</tr>'
        for a in alerts
    ])
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            await client.post(
                "https://api.resend.com/emails",
                headers={"Authorization": f"Bearer {RESEND_API_KEY}", "Content-Type": "application/json"},
                json={
                    "from": f"Axiom Alerts <{RESEND_FROM}>",
                    "to": [to_email],
                    "subject": f"⚠️ {len(alerts)} KPI alert{'s' if len(alerts)>1 else ''} require attention",
                    "html": f"""
                    <div style="font-family:-apple-system,sans-serif;max-width:560px;margin:0 auto;padding:40px 24px">
                      <h2 style="color:#0f172a;margin:0 0 6px">KPI Alert</h2>
                      <p style="color:#64748b;margin:0 0 20px">
                        {len(alerts)} metric{'s' if len(alerts)>1 else ''} require your attention:
                      </p>
                      <table style="width:100%;border-collapse:collapse;border:1px solid #e2e8f0;border-radius:8px;overflow:hidden">
                        <thead>
                          <tr style="background:#f8fafc">
                            <th style="padding:10px 12px;text-align:left;color:#475569;font-size:12px;font-weight:600;text-transform:uppercase">KPI</th>
                            <th style="padding:10px 12px;text-align:left;color:#475569;font-size:12px;font-weight:600;text-transform:uppercase">Status</th>
                            <th style="padding:10px 12px;text-align:left;color:#475569;font-size:12px;font-weight:600;text-transform:uppercase">Value</th>
                          </tr>
                        </thead>
                        <tbody>{rows_html}</tbody>
                      </table>
                      <a href="{APP_URL}"
                         style="background:#6366f1;color:#fff;padding:11px 22px;border-radius:6px;
                                text-decoration:none;display:inline-block;margin-top:24px;font-weight:600">
                        Open Dashboard →
                      </a>
                    </div>
                    """
                }
            )
        return True
    except Exception:
        return False


@router.post("/api/alerts/send-kpi-alert")
async def trigger_kpi_alert(request: Request):
    """Manually trigger KPI alert email for all critical KPIs."""
    from routers.analytics import _compute_fingerprint_data
    workspace_id = _get_workspace(request)
    conn = get_db()
    targets = {r["kpi_key"]: r["target_value"] for r in conn.execute("SELECT * FROM kpi_targets WHERE workspace_id=?", [workspace_id]).fetchall()}
    conn.close()
    alerts = []
    try:
        fingerprint = _compute_fingerprint_data(targets_override=targets, workspace_id=workspace_id)
        for kpi in fingerprint.get("kpis", []):
            if kpi.get("fy_status") == "critical":
                alerts.append({
                    "kpi": kpi.get("name", kpi.get("key", "")),
                    "status": "Critical",
                    "value": str(kpi.get("avg", ""))
                })
    except Exception:
        pass
    if not alerts:
        return {"message": "No critical KPIs to alert on", "sent": False}
    recipient = ALLOWED_EMAILS[0] if ALLOWED_EMAILS else RESEND_FROM
    sent = await _send_kpi_alert_email(recipient, alerts)
    return {"message": f"Alert sent for {len(alerts)} KPIs" if sent else "Email not configured", "alerts": len(alerts), "sent": sent}


@router.post("/api/alerts/test-email")
async def test_email(request: Request):
    """Send a test email to verify Resend configuration."""
    if not RESEND_API_KEY:
        return {"configured": False, "message": "RESEND_API_KEY not set"}
    recipient = ALLOWED_EMAILS[0] if ALLOWED_EMAILS else RESEND_FROM
    try:
        async with httpx.AsyncClient(timeout=10) as client:
            resp = await client.post(
                "https://api.resend.com/emails",
                headers={"Authorization": f"Bearer {RESEND_API_KEY}", "Content-Type": "application/json"},
                json={
                    "from": f"Axiom <{RESEND_FROM}>",
                    "to": [recipient],
                    "subject": "✅ Axiom email is configured",
                    "html": "<p>Your Axiom email integration is working correctly.</p>"
                }
            )
        return {"configured": True, "sent": resp.status_code in (200, 201), "status_code": resp.status_code}
    except Exception as e:
        return {"configured": True, "sent": False, "error": str(e)}


# ─── Weekly Briefing HTML Export ──────────────────────────────────────────────

@router.get("/api/export/weekly-briefing.html", tags=["Board Deck"])
def export_weekly_briefing(request: Request, stage: str = "series_b"):
    """Generate an HTML weekly briefing document."""
    from routers.analytics import _compute_fingerprint_data
    workspace_id = _get_workspace(request)
    fp_data = _compute_fingerprint_data(workspace_id=workspace_id)
    if not fp_data:
        raise HTTPException(status_code=404, detail="No data available")

    # Get benchmarks
    bench = {}
    for kpi_key, stages_data in BENCHMARKS.items():
        if stage in stages_data:
            bench[kpi_key] = stages_data[stage]

    # Get accountability data
    conn = get_db()
    acct_rows = conn.execute("SELECT * FROM kpi_accountability WHERE workspace_id=?", [workspace_id]).fetchall()
    conn.close()
    acct = {}
    for r in acct_rows:
        acct[r["kpi_key"]] = {"owner": r["owner"], "due_date": r["due_date"], "status": r["status"]}

    # Categorise
    green = [k for k in fp_data if k["fy_status"] == "green"]
    yellow = [k for k in fp_data if k["fy_status"] == "yellow"]
    red = [k for k in fp_data if k["fy_status"] == "red"]
    total = len(fp_data)

    # Sort red by gap magnitude
    def gap_pct(k):
        if k["avg"] is None or not k["target"]: return 0
        raw = (k["avg"] / k["target"] - 1) * 100
        return -raw if k["direction"] != "higher" else raw

    red_sorted = sorted(red, key=lambda k: abs(gap_pct(k)), reverse=True)
    yellow_sorted = sorted(yellow, key=lambda k: abs(gap_pct(k)), reverse=True)

    stage_label = {"seed": "Seed", "series_a": "Series A", "series_b": "Series B", "series_c": "Series C+"}.get(stage, stage)
    date_str = datetime.now().strftime("%B %d, %Y")

    def fmt_val(val, unit):
        if val is None: return "—"
        if unit == "pct": return f"{val:.1f}%"
        if unit == "days": return f"{val:.1f}d"
        if unit == "months": return f"{val:.1f}mo"
        if unit == "ratio": return f"{val:.2f}x"
        return f"{val:.2f}"

    # Get causation rules
    causation = {}
    try:
        causation = CAUSATION_RULES
    except Exception:
        pass

    # Build HTML
    html_parts = []
    html_parts.append(f"""<!DOCTYPE html>
<html><head><meta charset="utf-8"><title>Weekly Briefing — {date_str}</title>
<style>
  body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; max-width: 800px; margin: 40px auto; padding: 0 20px; color: #1e293b; line-height: 1.6; }}
  h1 {{ color: #0f172a; font-size: 24px; border-bottom: 2px solid #0055A4; padding-bottom: 8px; }}
  h2 {{ color: #334155; font-size: 18px; margin-top: 32px; }}
  .summary {{ display: flex; gap: 16px; margin: 16px 0; }}
  .stat {{ padding: 12px 20px; border-radius: 12px; text-align: center; flex: 1; }}
  .stat-red {{ background: #fef2f2; border: 1px solid #fecaca; }}
  .stat-yellow {{ background: #fffbeb; border: 1px solid #fde68a; }}
  .stat-green {{ background: #f0fdf4; border: 1px solid #bbf7d0; }}
  .stat .num {{ font-size: 28px; font-weight: 800; }}
  .stat-red .num {{ color: #dc2626; }}
  .stat-yellow .num {{ color: #d97706; }}
  .stat-green .num {{ color: #059669; }}
  .stat .label {{ font-size: 11px; color: #64748b; text-transform: uppercase; font-weight: 700; letter-spacing: 0.05em; }}
  table {{ width: 100%; border-collapse: collapse; margin: 12px 0; font-size: 13px; }}
  th {{ background: #0055A4; color: white; padding: 8px 12px; text-align: left; font-size: 11px; text-transform: uppercase; letter-spacing: 0.05em; }}
  td {{ padding: 8px 12px; border-bottom: 1px solid #e2e8f0; }}
  tr:nth-child(even) {{ background: #f8fafc; }}
  .red {{ color: #dc2626; font-weight: 700; }}
  .yellow {{ color: #d97706; font-weight: 700; }}
  .green {{ color: #059669; font-weight: 700; }}
  .owner {{ color: #0055A4; font-weight: 600; }}
  .footer {{ margin-top: 40px; padding-top: 16px; border-top: 1px solid #e2e8f0; font-size: 11px; color: #94a3b8; }}
  .action {{ background: #eff6ff; border-left: 3px solid #0055A4; padding: 8px 12px; margin: 4px 0; font-size: 12px; border-radius: 0 8px 8px 0; }}
</style></head><body>
<h1>Weekly Finance Briefing</h1>
<p style="color:#64748b; font-size:13px;">{date_str} · {stage_label} stage · {total} KPIs tracked</p>

<div class="summary">
  <div class="stat stat-red"><div class="num">{len(red)}</div><div class="label">Critical</div></div>
  <div class="stat stat-yellow"><div class="num">{len(yellow)}</div><div class="label">Watch</div></div>
  <div class="stat stat-green"><div class="num">{len(green)}</div><div class="label">On Target</div></div>
</div>
""")

    if red_sorted:
        html_parts.append("<h2>🔴 Critical KPIs — Immediate Attention Required</h2>")
        html_parts.append("<table><tr><th>KPI</th><th>Current</th><th>Target</th><th>Gap</th><th>Owner</th><th>Status</th></tr>")
        for k in red_sorted:
            gap = gap_pct(k)
            a = acct.get(k["key"], {})
            owner_str = a.get("owner", "—") or "—"
            status_str = a.get("status", "unassigned") or "unassigned"
            cause_data = causation.get(k["key"], {})
            root = cause_data.get("root_causes", [""])[0] if isinstance(cause_data.get("root_causes"), list) else ""
            fix = cause_data.get("corrective_actions", [""])[0] if isinstance(cause_data.get("corrective_actions"), list) else ""
            html_parts.append(f'<tr><td><strong>{k["name"]}</strong></td><td>{fmt_val(k["avg"], k["unit"])}</td><td>{fmt_val(k["target"], k["unit"])}</td><td class="red">{gap:+.1f}%</td><td class="owner">{owner_str}</td><td>{status_str}</td></tr>')
        html_parts.append("</table>")

    if yellow_sorted:
        html_parts.append("<h2>🟡 Watch Zone</h2>")
        html_parts.append("<table><tr><th>KPI</th><th>Current</th><th>Target</th><th>Gap</th><th>Owner</th></tr>")
        for k in yellow_sorted[:6]:
            gap = gap_pct(k)
            a = acct.get(k["key"], {})
            owner_str = a.get("owner", "—") or "—"
            html_parts.append(f'<tr><td>{k["name"]}</td><td>{fmt_val(k["avg"], k["unit"])}</td><td>{fmt_val(k["target"], k["unit"])}</td><td class="yellow">{gap:+.1f}%</td><td class="owner">{owner_str}</td></tr>')
        html_parts.append("</table>")

    if green[:5]:
        html_parts.append("<h2>🟢 Bright Spots</h2>")
        html_parts.append("<table><tr><th>KPI</th><th>Current</th><th>Target</th><th>Above Target</th></tr>")
        for k in sorted(green, key=lambda k: gap_pct(k), reverse=True)[:5]:
            gap = gap_pct(k)
            html_parts.append(f'<tr><td>{k["name"]}</td><td>{fmt_val(k["avg"], k["unit"])}</td><td>{fmt_val(k["target"], k["unit"])}</td><td class="green">+{abs(gap):.1f}%</td></tr>')
        html_parts.append("</table>")

    html_parts.append(f"""
<div class="footer">
  Generated by Axiom Intelligence · {date_str}<br>
  This briefing covers {total} KPIs for the {stage_label} stage.
</div>
</body></html>""")

    html_content = "\n".join(html_parts)
    buf = io.BytesIO(html_content.encode("utf-8"))
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="text/html",
        headers={"Content-Disposition": f'attachment; filename="weekly-briefing-{datetime.now().strftime("%Y%m%d")}.html"'}
    )


# ─── Data Export: Excel ───────────────────────────────────────────────────────

@router.get("/api/export/data.xlsx")
def export_data_xlsx(request: Request):
    """Export all monthly KPI data to Excel for offline editing."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    workspace_id = _get_workspace(request)

    # ── KPI metadata: key → (full_name, unit_label, direction, used_for) ──────
    KPI_META: dict[str, tuple[str, str, str, str]] = {
        "revenue_growth":        ("Revenue Growth Rate",          "%",              "higher is better",  "Status Distribution · Fingerprint · Signal Engine · Markov Forecast; feeds Revenue Momentum Index"),
        "gross_margin":          ("Gross Margin %",               "%",              "higher is better",  "Status Distribution · Fingerprint · Signal Engine · Markov Forecast; feeds Operating Margin, EBITDA Margin, Contribution Margin"),
        "operating_margin":      ("Operating Margin %",           "%",              "higher is better",  "Status Distribution · Fingerprint · Signal Engine; downstream of Gross Margin"),
        "ebitda_margin":         ("EBITDA Margin %",              "%",              "higher is better",  "Status Distribution · Fingerprint · Signal Engine; downstream of Operating Margin & Gross Margin"),
        "cash_conv_cycle":       ("Cash Conversion Cycle",        "days",           "lower is better",   "Status Distribution · Fingerprint · Signal Engine · Markov Forecast"),
        "dso":                   ("Days Sales Outstanding",       "days",           "lower is better",   "Status Distribution · Fingerprint · Signal Engine; component of Cash Conversion Cycle"),
        "arr_growth":            ("ARR Growth Rate",              "%",              "higher is better",  "Status Distribution · Fingerprint · Signal Engine · Markov Forecast; feeds Growth Efficiency Index, Revenue Momentum Index"),
        "nrr":                   ("Net Revenue Retention",        "%",              "higher is better",  "Status Distribution · Fingerprint · Signal Engine · Markov Forecast; feeds Revenue Fragility Index"),
        "burn_multiple":         ("Burn Multiple",                "ratio (×)",      "lower is better",   "Status Distribution · Fingerprint · Signal Engine · Markov Forecast; feeds Growth Efficiency Index, Burn Convexity"),
        "opex_ratio":            ("Operating Expense Ratio",      "%",              "lower is better",   "Status Distribution · Fingerprint · Signal Engine"),
        "contribution_margin":   ("Contribution Margin %",        "%",              "higher is better",  "Status Distribution · Fingerprint · Signal Engine; downstream of Gross Margin"),
        "revenue_quality":       ("Revenue Quality Ratio",        "%",              "higher is better",  "Status Distribution · Fingerprint · Signal Engine"),
        "cac_payback":           ("CAC Payback Period",           "months",         "lower is better",   "Status Distribution · Fingerprint · Signal Engine · Markov Forecast"),
        "sales_efficiency":      ("Sales Efficiency Ratio",       "ratio (×)",      "higher is better",  "Status Distribution · Fingerprint · Signal Engine; downstream of Pipeline Conversion"),
        "customer_concentration":("Customer Concentration",       "%",              "lower is better",   "Status Distribution · Fingerprint · Signal Engine; feeds Revenue Fragility Index"),
        "recurring_revenue":     ("Recurring Revenue Ratio",      "%",              "higher is better",  "Status Distribution · Fingerprint · Signal Engine"),
        "churn_rate":            ("Monthly Churn Rate",           "%",              "lower is better",   "Status Distribution · Fingerprint · Signal Engine · Markov Forecast; feeds Revenue Fragility Index, Customer Decay Curve Slope"),
        "operating_leverage":    ("Operating Leverage Index",     "ratio (×)",      "higher is better",  "Status Distribution · Fingerprint · Signal Engine"),
        "pipeline_conversion":   ("Pipeline Conversion Rate",     "%",              "higher is better",  "Status Distribution · Fingerprint · Signal Engine; feeds Sales Efficiency, ARR Growth"),
        "customer_ltv":          ("Customer Lifetime Value",      "$ (thousands)",  "higher is better",  "Status Distribution · Fingerprint · Signal Engine"),
        "pricing_power_index":   ("Pricing Power Index",          "%",              "higher is better",  "Status Distribution · Fingerprint · Signal Engine"),
        "growth_efficiency":     ("Growth Efficiency Index",      "ratio (×)",      "higher is better",  "Status Distribution · Fingerprint · Signal Engine; DERIVED = ARR Growth Rate ÷ Burn Multiple"),
        "revenue_momentum":      ("Revenue Momentum Index",       "ratio (×)",      "higher is better",  "Status Distribution · Fingerprint · Signal Engine; DERIVED = Current Rev Growth ÷ Annual Avg Rev Growth"),
        "revenue_fragility":     ("Strategic Revenue Fragility",  "ratio (×)",      "lower is better",   "Status Distribution · Fingerprint · Signal Engine; DERIVED = (Customer Concentration × Churn Rate) ÷ NRR"),
        "burn_convexity":        ("Burn Convexity",               "ratio (×)",      "lower is better",   "Status Distribution · Fingerprint · Signal Engine; DERIVED = Month-over-Month change in Burn Multiple"),
        "margin_volatility":     ("Margin Volatility Index",      "%",              "lower is better",   "Status Distribution · Fingerprint · Signal Engine; DERIVED = 6-Month rolling std dev of Gross Margin"),
        "customer_decay_slope":  ("Customer Decay Curve Slope",   "%",              "lower is better",   "Status Distribution · Fingerprint · Signal Engine; DERIVED = Month-over-Month change in Churn Rate"),
    }

    conn = get_db()
    rows = conn.execute(
        "SELECT year, month, data_json FROM monthly_data WHERE workspace_id=? ORDER BY year, month",
        [workspace_id]
    ).fetchall()
    conn.close()

    if not rows:
        raise HTTPException(status_code=404, detail="No data found")

    # Collect KPI keys in consistent order
    kpi_keys: list[str] = []
    parsed = []
    for r in rows:
        d = json.loads(r["data_json"])
        for k in d:
            if k not in kpi_keys:
                kpi_keys.append(k)
        parsed.append((r["year"], r["month"], d))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "KPI Data"

    # ── Colour palette ───────────────────────────────────────────────────────
    fill_key   = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")  # dark navy
    fill_name  = PatternFill(start_color="2C5282", end_color="2C5282", fill_type="solid")  # medium blue
    fill_unit  = PatternFill(start_color="2B6CB0", end_color="2B6CB0", fill_type="solid")  # blue
    fill_usage = PatternFill(start_color="EBF4FF", end_color="EBF4FF", fill_type="solid")  # pale blue
    fill_meta_left = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")  # dark slate for Year/Month cols

    font_white_bold = Font(color="FFFFFF", bold=True,  size=9)
    font_white      = Font(color="FFFFFF", bold=False, size=9)
    font_navy       = Font(color="1E3A5F", bold=False, size=8, italic=True)
    align_center    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_right     = Alignment(horizontal="right",  vertical="center")
    thin_border     = Border(
        bottom=Side(style="thin", color="CBD5E0"),
        right=Side(style="thin",  color="CBD5E0"),
    )

    month_names = ["Jan","Feb","Mar","Apr","May","Jun",
                   "Jul","Aug","Sep","Oct","Nov","Dec"]

    # ── Row 1: machine-readable KPI key (used by import) ────────────────────
    row1_values = ["Year", "Month"] + kpi_keys
    for ci, val in enumerate(row1_values, 1):
        c = ws.cell(row=1, column=ci, value=val)
        c.fill  = fill_key if ci > 2 else fill_meta_left
        c.font  = font_white_bold
        c.alignment = align_center

    # ── Row 2: human-readable full name ─────────────────────────────────────
    ws.cell(row=2, column=1, value="Full Name").fill   = fill_meta_left
    ws.cell(row=2, column=1).font                      = font_white_bold
    ws.cell(row=2, column=1).alignment                 = align_center
    ws.cell(row=2, column=2, value="").fill            = fill_meta_left
    for ci, k in enumerate(kpi_keys, 3):
        meta = KPI_META.get(k, (k, "", "", ""))
        c = ws.cell(row=2, column=ci, value=meta[0])
        c.fill      = fill_name
        c.font      = font_white
        c.alignment = align_center

    # ── Row 3: unit of measure ───────────────────────────────────────────────
    ws.cell(row=3, column=1, value="Unit").fill        = fill_meta_left
    ws.cell(row=3, column=1).font                      = font_white_bold
    ws.cell(row=3, column=1).alignment                 = align_center
    ws.cell(row=3, column=2, value="").fill            = fill_meta_left
    for ci, k in enumerate(kpi_keys, 3):
        meta = KPI_META.get(k, (k, "", "", ""))
        c = ws.cell(row=3, column=ci, value=meta[1])
        c.fill      = fill_unit
        c.font      = font_white_bold
        c.alignment = align_center

    # ── Row 4: used-for description ──────────────────────────────────────────
    ws.cell(row=4, column=1, value="Used For").fill    = fill_meta_left
    ws.cell(row=4, column=1).font                      = font_white_bold
    ws.cell(row=4, column=1).alignment                 = align_center
    ws.cell(row=4, column=2, value="").fill            = fill_meta_left
    for ci, k in enumerate(kpi_keys, 3):
        meta = KPI_META.get(k, (k, "", "", ""))
        c = ws.cell(row=4, column=ci, value=meta[3])
        c.fill      = fill_usage
        c.font      = font_navy
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # Row heights for header block
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 52

    # ── Data rows start at row 5 ─────────────────────────────────────────────
    for row_idx, (year, month, d) in enumerate(parsed, 5):
        ws.cell(row=row_idx, column=1, value=year).alignment   = align_right
        ws.cell(row=row_idx, column=2,
                value=month_names[month - 1] if 1 <= month <= 12 else month).alignment = align_center
        # Alternating row shading
        alt_fill = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid") if row_idx % 2 == 0 else None
        for ci, k in enumerate(kpi_keys, 3):
            val = d.get(k)
            c   = ws.cell(row=row_idx, column=ci,
                          value=round(val, 4) if isinstance(val, float) else val)
            c.alignment = align_right
            c.border    = thin_border
            if alt_fill:
                c.fill = alt_fill

    # ── Column widths ────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 7
    for ci in range(3, len(row1_values) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 20

    # Freeze below the 4-row header block
    ws.freeze_panes = "C5"

    # ── README / legend sheet ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("README")
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 80
    readme_rows = [
        ("INSTRUCTIONS", ""),
        ("Step 1", "Edit KPI values in the 'KPI Data' sheet — do NOT change row 1 (KPI keys) or columns A–B (Year / Month)."),
        ("Step 2", "Save the file as .xlsx."),
        ("Step 3", "Upload via Settings › Import Data in the platform."),
        ("", ""),
        ("HEADER GUIDE", ""),
        ("Row 1 — KPI Key",   "Machine-readable identifier used by the import engine. Do not edit."),
        ("Row 2 — Full Name", "Human-readable KPI name for reference only."),
        ("Row 3 — Unit",      "Unit of measure: % = percentage point value (e.g. 6.5 = 6.5%); ratio (×) = dimensionless multiplier; days / months = calendar count; $ (thousands) = USD thousands."),
        ("Row 4 — Used For",  "Platform features & model components that consume this KPI. 'DERIVED' means the value is computed from other KPIs stored in this sheet."),
        ("", ""),
        ("UNIT NOTES", ""),
        ("%",            "Store as a plain number, NOT as a decimal. e.g. 62.5 means 62.5%, not 0.625."),
        ("ratio (×)",    "Dimensionless multiplier. e.g. 1.2 means 1.2×."),
        ("days / months","Integer or decimal count of calendar days or months."),
        ("$ (thousands)","USD value in thousands. e.g. 80 means $80,000."),
        ("DERIVED KPIs", "Values are pre-computed for testing. You may override them; the model will use whatever value is in the cell."),
    ]
    for ri, (label, text) in enumerate(readme_rows, 1):
        ca = ws2.cell(row=ri, column=1, value=label)
        cb = ws2.cell(row=ri, column=2, value=text)
        if label.isupper() and label:
            ca.font = Font(bold=True, size=10, color="1E3A5F")
        ca.alignment = Alignment(vertical="top")
        cb.alignment = Alignment(wrap_text=True, vertical="top")
        ws2.row_dimensions[ri].height = 28 if text else 10

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=axiom_kpi_data.xlsx"},
    )


# ─── Data Import: Excel or CSV ────────────────────────────────────────────────

@router.post("/api/import/data")
async def import_data(request: Request, file: UploadFile = File(...)):
    """Import KPI data from a previously exported (and edited) Excel or CSV file.

    Accepted formats:
    - .xlsx  — exported from Download KPI Data; must contain a 'KPI Data' sheet.
    - .csv   — same structure as the xlsx KPI Data sheet (exported via Save As CSV).

    Both formats support the enriched 4-row header block (rows 2-4 are metadata
    and are skipped automatically). Row 1 always holds machine-readable KPI keys.
    Requires authentication; data is scoped to the current workspace.
    """
    import pandas as pd

    workspace_id = _get_workspace(request)
    if not workspace_id:
        raise HTTPException(status_code=401, detail="Authentication required")

    fname = (file.filename or "").lower()
    contents = await file.read()

    month_map = {"jan":1,"feb":2,"mar":3,"apr":4,"may":5,"jun":6,
                 "jul":7,"aug":8,"sep":9,"oct":10,"nov":11,"dec":12}

    records: list[tuple] = []

    if fname.endswith(".csv"):
        # ── CSV path ──────────────────────────────────────────────────────────
        try:
            df = pd.read_csv(io.StringIO(contents.decode("utf-8", errors="replace")))
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Could not parse CSV: {e}")

        if "Year" not in df.columns or "Month" not in df.columns:
            raise HTTPException(status_code=400,
                detail="Invalid file — expected 'Year' and 'Month' columns. "
                       "Please use a file exported from Download KPI Data.")

        kpi_keys = [c for c in df.columns if c not in ("Year", "Month")]

        # Detect enriched format: first data row is a metadata row (Full Name / Unit / Used For)
        if not df.empty and str(df.iloc[0, 0]).strip().lower() in ("full name", "unit", "used for", ""):
            df = df.iloc[3:].reset_index(drop=True)

        for _, row in df.iterrows():
            try:
                year = int(row["Year"])
                mo_raw = str(row["Month"]).strip()
                month = month_map.get(mo_raw.lower()) or int(mo_raw)
            except (ValueError, TypeError):
                continue
            data = {}
            for k in kpi_keys:
                v = row.get(k)
                if v is not None and str(v).strip() not in ("", "nan", "None"):
                    try:
                        data[k] = float(v)
                    except (ValueError, TypeError):
                        pass
            if data:
                records.append((year, month, json.dumps(data)))

    elif fname.endswith(".xlsx") or fname.endswith(".xls"):
        # ── Excel path ────────────────────────────────────────────────────────
        import openpyxl
        try:
            wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Could not open Excel file: {e}")

        if "KPI Data" not in wb.sheetnames:
            raise HTTPException(status_code=400,
                detail="Sheet 'KPI Data' not found. Please use a file exported from Download KPI Data.")
        ws = wb["KPI Data"]

        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        if headers[0] != "Year" or headers[1] != "Month":
            raise HTTPException(status_code=400,
                detail="Invalid file format — expected Year, Month in row 1 columns A–B")

        kpi_keys = headers[2:]
        row2_a = ws.cell(row=2, column=1).value
        data_start_row = 5 if str(row2_a).strip().lower() in ("full name", "unit", "used for") else 2

        for row in ws.iter_rows(min_row=data_start_row, values_only=True):
            if row[0] is None:
                continue
            try:
                year = int(row[0])
                mo_raw = str(row[1]).strip()
                month = month_map.get(mo_raw.lower()) or int(mo_raw)
            except (ValueError, TypeError):
                continue
            data = {}
            for i, k in enumerate(kpi_keys):
                if k is None:
                    continue
                v = row[2 + i]
                if v is not None:
                    try:
                        data[k] = float(v)
                    except (ValueError, TypeError):
                        pass
            if data:
                records.append((year, month, json.dumps(data)))
    else:
        raise HTTPException(status_code=400,
            detail=f"Unsupported file type. Please upload a .csv or .xlsx file exported from Download KPI Data.")

    if not records:
        raise HTTPException(status_code=400, detail="No data rows found in file.")

    conn = get_db()
    try:
        for year, month, data_json in records:
            existing = conn.execute(
                "SELECT id FROM monthly_data WHERE year=? AND month=? AND workspace_id=?",
                (year, month, workspace_id)
            ).fetchone()
            if existing:
                conn.execute(
                    "UPDATE monthly_data SET data_json=? WHERE year=? AND month=? AND workspace_id=?",
                    (data_json, year, month, workspace_id)
                )
            else:
                conn.execute(
                    "INSERT INTO monthly_data (year, month, data_json, workspace_id) VALUES (?,?,?,?)",
                    (year, month, data_json, workspace_id)
                )
        conn.commit()
    finally:
        conn.close()

    return {"status": "ok", "rows_imported": len(records)}
