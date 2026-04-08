"""
Generate a 12-sheet canonical workbook with realistic B2B SaaS data.

This workbook matches the platform's canonical table schemas. Users can
modify values and re-upload via POST /api/upload/canonical-xlsx to test
the full 61-KPI computation pipeline.

Usage:
    python3 scripts/generate_canonical_workbook.py [output_path]
"""
import math
import random
import sys
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ── Styling ──────────────────────────────────────────────────────────────────
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
HDR_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HDR_ALIGN = Alignment(horizontal="center", wrap_text=True)
NOTE_FONT = Font(italic=True, color="888888", size=10)


def _jitter(base, pct=0.05):
    """Add ±pct random jitter."""
    return round(base * (1 + random.uniform(-pct, pct)), 2)


def _date_str(yr, mo, day=15):
    return f"{yr}-{mo:02d}-{day:02d}"


def _write_sheet(ws, headers, rows, col_widths=None):
    """Write headers + data rows to a worksheet."""
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font, c.fill, c.alignment = HDR_FONT, HDR_FILL, HDR_ALIGN
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val)
    ws.freeze_panes = "A2"
    for ci, h in enumerate(headers, 1):
        w = (col_widths or {}).get(h, max(len(h) + 3, 14))
        ws.column_dimensions[get_column_letter(ci)].width = min(w, 25)


def generate_canonical_workbook(months=52, start_year=2022, start_month=1):
    """
    Generate a Workbook with 12 sheets of realistic B2B SaaS data.

    Returns an openpyxl.Workbook ready to save or stream.
    """
    random.seed(42)  # Reproducible
    wb = Workbook()

    # ── Date range ───────────────────────────────────────────────────────────
    periods = []
    yr, mo = start_year, start_month
    for _ in range(months):
        periods.append((yr, mo))
        mo += 1
        if mo > 12:
            mo = 1
            yr += 1

    # ── Business trajectory (realistic B2B SaaS growth) ──────────────────────
    # MRR: $40K → $500K over 52 months (Series A trajectory)
    base_mrr = 40000
    mrr_growth_monthly = 0.055  # ~5.5% MoM initially, decelerating

    # Customer base: 35 → 600
    base_customers = 35
    monthly_new_customers = 8  # grows over time

    # Headcount: 12 → 85
    base_headcount = 12

    # ── Generate data per month ──────────────────────────────────────────────
    revenue_rows = []
    expense_rows = []
    customer_rows = []
    pipeline_rows = []
    invoice_rows = []
    employee_rows = []
    marketing_rows = []
    balance_sheet_rows = []
    time_tracking_rows = []
    survey_rows = []
    support_rows = []
    product_usage_rows = []

    mrr = base_mrr
    total_customers = base_customers
    headcount = base_headcount
    cash_balance = 2_500_000  # $2.5M post-seed
    customer_id_counter = 100

    for idx, (yr, mo) in enumerate(periods):
        period = _date_str(yr, mo)
        month_frac = idx / max(months - 1, 1)  # 0→1 over time

        # ── MRR evolution (decelerating growth) ──────────────────────────────
        growth_rate = mrr_growth_monthly * (1 - month_frac * 0.5)  # slows from 5.5% to 2.75%
        mrr = mrr * (1 + growth_rate + random.uniform(-0.02, 0.02))
        arr = mrr * 12
        total_revenue = mrr * _jitter(1.08, 0.03)  # ~8% non-recurring on top

        # Recurring vs one-time split
        recurring_rev = mrr
        onetime_rev = total_revenue - recurring_rev

        # ── Customers ────────────────────────────────────────────────────────
        new_cust = max(2, int(monthly_new_customers * (1 + month_frac * 1.5) + random.randint(-3, 5)))
        churn_rate_base = max(0.5, 3.5 - month_frac * 2)  # 3.5% → 1.5% over time
        churned = max(0, int(total_customers * churn_rate_base / 100 + random.randint(-1, 2)))
        total_customers = max(10, total_customers + new_cust - churned)

        # ── Revenue transactions (5-15 per month for realism) ────────────────
        n_transactions = random.randint(5, 15)
        for ti in range(n_transactions):
            cid = f"CUST-{random.randint(100, 100 + total_customers)}"
            is_recurring = random.random() < 0.82
            amount = _jitter(recurring_rev / n_transactions if is_recurring else onetime_rev / max(1, n_transactions // 3))
            revenue_rows.append([
                f"rev-{yr}{mo:02d}-{ti}",         # source_id
                round(amount, 2),                   # amount
                "USD",                              # currency
                period,                             # period
                cid,                                # customer_id
                "recurring" if is_recurring else "one-time",  # subscription_type
                f"PROD-{random.choice(['A','B','C'])}",       # product_id
                period,                             # recognized_at
            ])

        # ── Expenses (categorized: COGS, S&M, G&A, R&D) ─────────────────────
        cogs = total_revenue * _jitter(0.32, 0.03)   # ~32% COGS → ~68% gross margin
        sm_spend = total_revenue * _jitter(0.38, 0.05)  # S&M ~38% of revenue
        ga_spend = total_revenue * _jitter(0.12, 0.03)   # G&A ~12%
        rd_spend = total_revenue * _jitter(0.22, 0.04)   # R&D ~22%

        expense_entries = [
            ("cogs", cogs, "Hosting & Infrastructure"),
            ("sales", sm_spend * 0.6, "Sales Team Costs"),
            ("marketing", sm_spend * 0.4, "Marketing Programs"),
            ("general", ga_spend, "General & Administrative"),
            ("research", rd_spend, "Product Development"),
        ]
        for ei, (cat, amt, desc) in enumerate(expense_entries):
            expense_rows.append([
                f"exp-{yr}{mo:02d}-{ei}",
                round(amt, 2),
                "USD",
                cat,
                f"Vendor-{cat[:3].upper()}",
                period,
                desc,
            ])

        # ── Customers (new acquisitions this month) ──────────────────────────
        for ci in range(new_cust):
            customer_id_counter += 1
            customer_rows.append([
                f"cust-{customer_id_counter}",
                f"Customer {customer_id_counter}",
                f"c{customer_id_counter}@example.com",
                f"Company {customer_id_counter}",
                period,
                random.choice(["prospect", "active", "enterprise"]),
            ])

        # ── Pipeline (8-20 deals per month) ──────────────────────────────────
        n_deals = random.randint(8, 20)
        for di in range(n_deals):
            stage = random.choice(["discovery", "proposal", "negotiation", "closed-won", "closed-lost"])
            deal_amt = _jitter(arr / total_customers * random.uniform(0.5, 3), 0.15)
            won = stage == "closed-won"
            close_dt = _date_str(yr, mo, random.randint(1, 28))
            pipeline_rows.append([
                f"deal-{yr}{mo:02d}-{di}",
                f"Deal {yr}{mo:02d}-{di}",
                round(deal_amt, 2),
                stage,
                close_dt if won or stage == "closed-lost" else "",
                round(random.uniform(0.1, 0.95), 2),
                f"Rep-{random.choice(['A','B','C','D'])}",
                _date_str(yr, mo, max(1, random.randint(1, 15))),
            ])

        # ── Invoices (one per customer subset) ───────────────────────────────
        n_invoices = min(total_customers, random.randint(15, 40))
        for ii in range(n_invoices):
            issue_day = random.randint(1, 28)
            due_day = min(28, issue_day + random.choice([15, 30, 45, 60]))
            inv_status = random.choices(["paid", "outstanding", "overdue"], weights=[70, 20, 10])[0]
            invoice_rows.append([
                f"inv-{yr}{mo:02d}-{ii}",
                round(_jitter(mrr / n_invoices * random.uniform(0.5, 2)), 2),
                "USD",
                f"CUST-{random.randint(100, 100 + total_customers)}",
                _date_str(yr, mo, issue_day),
                _date_str(yr, mo, due_day),
                inv_status,
                period,
            ])

        # ── Employees ────────────────────────────────────────────────────────
        headcount_growth = max(0, int(1 + month_frac * 2 + random.randint(-1, 2)))
        headcount += headcount_growth
        for ei in range(headcount_growth):
            dept = random.choice(["Engineering", "Sales", "Marketing", "CS", "G&A", "Product"])
            salary = _jitter({"Engineering": 140000, "Sales": 120000, "Marketing": 110000, "CS": 95000, "G&A": 100000, "Product": 130000}[dept])
            employee_rows.append([
                f"emp-{yr}{mo:02d}-{ei}",
                f"Employee {yr}{mo:02d}-{ei}",
                dept,
                dept,
                round(salary, 0),
                _date_str(yr, mo, random.randint(1, 28)),
                "active",
            ])

        # ── Marketing (channel-level spend) ──────────────────────────────────
        channels = [
            ("Google Ads", 0.35), ("LinkedIn", 0.25), ("Content", 0.15),
            ("Events", 0.10), ("SEO", 0.08), ("Referral", 0.07),
        ]
        total_mkt_spend = sm_spend * 0.4  # marketing portion of S&M
        total_leads = int(_jitter(new_cust * random.uniform(8, 15)))
        for ch_name, ch_share in channels:
            ch_leads = max(1, int(total_leads * ch_share * _jitter(1, 0.2)))
            ch_conv = max(0, int(ch_leads * random.uniform(0.03, 0.12)))
            marketing_rows.append([
                f"mkt-{yr}{mo:02d}-{ch_name[:3].lower()}",
                ch_name,
                round(total_mkt_spend * ch_share * _jitter(1, 0.1), 2),
                period,
                ch_leads,
                ch_conv,
            ])

        # ── Balance Sheet (monthly snapshot) ─────────────────────────────────
        monthly_burn = (cogs + sm_spend + ga_spend + rd_spend) - total_revenue
        cash_balance = max(0, cash_balance - monthly_burn)
        if idx == 18:  # Series A at month 18
            cash_balance += 8_000_000
        current_assets = cash_balance + total_revenue * 1.3
        current_liabilities = total_revenue * 0.4
        balance_sheet_rows.append([
            f"bs-{yr}{mo:02d}",
            period,
            round(cash_balance, 2),
            round(current_assets, 2),
            round(current_liabilities, 2),
        ])

        # ── Time Tracking (team utilization) ─────────────────────────────────
        n_workers = min(headcount, random.randint(5, 15))
        for wi in range(n_workers):
            total_hrs = _jitter(160, 0.05)  # ~160 hrs/month
            billable_pct = random.uniform(0.55, 0.85)
            time_tracking_rows.append([
                f"time-{yr}{mo:02d}-{wi}",
                f"worker-{wi}",
                period,
                round(total_hrs * billable_pct, 1),
                round(total_hrs, 1),
            ])

        # ── Surveys (quarterly NPS + monthly CSAT) ───────────────────────────
        n_responses = random.randint(8, 25)
        for si in range(n_responses):
            nps = random.choices(range(0, 11), weights=[1,1,1,1,1,2,3,5,8,6,4])[0]
            csat = round(random.uniform(3.0, 5.0), 1)
            survey_rows.append([
                f"survey-{yr}{mo:02d}-{si}",
                f"resp-{random.randint(100, 100 + total_customers)}",
                period,
                nps,
                csat,
            ])

        # ── Support (tickets) ────────────────────────────────────────────────
        n_tickets = max(3, int(total_customers * random.uniform(0.05, 0.15)))
        for ti in range(n_tickets):
            resolution_hrs = _jitter(random.choice([2, 4, 8, 16, 24, 48]), 0.3)
            support_rows.append([
                f"tkt-{yr}{mo:02d}-{ti}",
                f"TKT-{yr}{mo:02d}-{ti}",
                period,
                round(resolution_hrs, 1),
                f"CUST-{random.randint(100, 100 + total_customers)}",
            ])

        # ── Product Usage (feature adoption) ─────────────────────────────────
        n_active = max(5, int(total_customers * random.uniform(0.6, 0.9)))
        features = ["dashboard", "reports", "alerts", "api", "export", "import",
                     "workflows", "integrations", "analytics", "forecasting", "scenarios", "audit"]
        for ui in range(min(n_active, 30)):
            features_used = random.randint(3, len(features))
            days_to_activate = max(1, int(_jitter(random.choice([3, 7, 14, 21, 30]), 0.4)))
            product_usage_rows.append([
                f"usage-{yr}{mo:02d}-{ui}",
                f"user-{random.randint(100, 100 + total_customers)}",
                period,
                f"feature-{random.choice(features)}",
                random.randint(5, 200),
            ])

    # ── Write sheets ─────────────────────────────────────────────────────────

    # Sheet 1: Revenue
    ws = wb.active
    ws.title = "Revenue"
    _write_sheet(ws,
        ["source_id", "amount", "currency", "period", "customer_id", "subscription_type", "product_id", "recognized_at"],
        revenue_rows)

    # Sheet 2: Expenses
    ws2 = wb.create_sheet("Expenses")
    _write_sheet(ws2,
        ["source_id", "amount", "currency", "category", "vendor", "period", "description"],
        expense_rows)

    # Sheet 3: Customers
    ws3 = wb.create_sheet("Customers")
    _write_sheet(ws3,
        ["source_id", "name", "email", "company", "created_at", "lifecycle_stage"],
        customer_rows)

    # Sheet 4: Pipeline
    ws4 = wb.create_sheet("Pipeline")
    _write_sheet(ws4,
        ["source_id", "name", "amount", "stage", "close_date", "probability", "owner", "created_at"],
        pipeline_rows)

    # Sheet 5: Invoices
    ws5 = wb.create_sheet("Invoices")
    _write_sheet(ws5,
        ["source_id", "amount", "currency", "customer_id", "issue_date", "due_date", "status", "period"],
        invoice_rows)

    # Sheet 6: Employees
    ws6 = wb.create_sheet("Employees")
    _write_sheet(ws6,
        ["source_id", "name", "title", "department", "salary", "hire_date", "status"],
        employee_rows)

    # Sheet 7: Marketing
    ws7 = wb.create_sheet("Marketing")
    _write_sheet(ws7,
        ["source_id", "channel", "spend", "period", "leads", "conversions"],
        marketing_rows)

    # Sheet 8: Balance Sheet
    ws8 = wb.create_sheet("Balance Sheet")
    _write_sheet(ws8,
        ["source_id", "period", "cash_balance", "current_assets", "current_liabilities"],
        balance_sheet_rows)

    # Sheet 9: Time Tracking
    ws9 = wb.create_sheet("Time Tracking")
    _write_sheet(ws9,
        ["source_id", "worker_id", "period", "billable_hours", "total_hours"],
        time_tracking_rows)

    # Sheet 10: Surveys
    ws10 = wb.create_sheet("Surveys")
    _write_sheet(ws10,
        ["source_id", "respondent_id", "period", "nps_score", "csat_score"],
        survey_rows)

    # Sheet 11: Support
    ws11 = wb.create_sheet("Support")
    _write_sheet(ws11,
        ["source_id", "ticket_id", "period", "resolution_hours", "customer_id"],
        support_rows)

    # Sheet 12: Product Usage
    ws12 = wb.create_sheet("Product Usage")
    _write_sheet(ws12,
        ["source_id", "user_id", "period", "feature_id", "usage_count"],
        product_usage_rows)

    # Sheet 13: README
    ws_readme = wb.create_sheet("README")
    readme_text = [
        ("AXIOM INTELLIGENCE - Canonical Data Workbook", ""),
        ("", ""),
        ("This workbook contains 12 sheets of raw business data.", ""),
        ("Edit any values, then upload via Manual Upload > Canonical Workbook.", ""),
        ("The platform will compute all 61 KPIs, health scores, and narratives.", ""),
        ("", ""),
        ("SHEETS:", ""),
        ("Revenue", "Transaction-level revenue data (amount, customer, subscription type)"),
        ("Expenses", "Categorized expenses (COGS, S&M, G&A, R&D)"),
        ("Customers", "Customer records (name, email, lifecycle stage)"),
        ("Pipeline", "Sales deals (amount, stage, probability, close date)"),
        ("Invoices", "Accounts receivable (amount, issue/due date, status)"),
        ("Employees", "Team data (department, salary, hire date)"),
        ("Marketing", "Channel-level spend, leads, and conversions"),
        ("Balance Sheet", "Monthly snapshots (cash, assets, liabilities)"),
        ("Time Tracking", "Billable vs total hours per worker"),
        ("Surveys", "NPS and CSAT scores per respondent"),
        ("Support", "Ticket resolution times per customer"),
        ("Product Usage", "Feature-level usage counts per user"),
        ("", ""),
        ("NOTES:", ""),
        ("- 'period' columns use YYYY-MM-DD format (day is ignored, only year-month matters)", ""),
        ("- 'source_id' must be unique per sheet (used for deduplication)", ""),
        ("- Not all sheets are required - upload whatever data you have", ""),
        ("- Categories in Expenses: 'cogs', 'sales', 'marketing', 'general', 'research'", ""),
        ("- Pipeline stages: 'discovery', 'proposal', 'negotiation', 'closed-won', 'closed-lost'", ""),
        ("- Invoice status: 'paid', 'outstanding', 'overdue'", ""),
    ]
    for ri, (col1, col2) in enumerate(readme_text, 1):
        c = ws_readme.cell(row=ri, column=1, value=col1)
        if ri == 1:
            c.font = Font(bold=True, size=14)
        ws_readme.cell(row=ri, column=2, value=col2)
    ws_readme.column_dimensions["A"].width = 50
    ws_readme.column_dimensions["B"].width = 65

    return wb


if __name__ == "__main__":
    output = sys.argv[1] if len(sys.argv) > 1 else "axiom_canonical_workbook.xlsx"
    wb = generate_canonical_workbook()
    wb.save(output)
    # Count rows per sheet
    for ws in wb.worksheets:
        if ws.title != "README":
            print(f"  {ws.title}: {ws.max_row - 1} rows")
    print(f"\nSaved to: {output}")
