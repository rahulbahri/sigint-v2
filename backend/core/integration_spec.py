"""
core/integration_spec.py — Integration Specification data and Excel workbook
generator for customer engineering teams.

Contains:
  - SYSTEM_REGISTRY: all 15 source systems with field-level specs
  - CANONICAL_SCHEMAS: 6 canonical table definitions
  - generate_integration_spec_workbook(): builds the multi-tab .xlsx
"""

from datetime import datetime

# ─── Canonical Table Schemas ───────────────────────────────────────────────

CANONICAL_SCHEMAS = {
    "canonical_revenue": {
        "description": "Normalised revenue transactions from all payment and accounting systems",
        "fields": [
            ("source",            "string",   "Source system identifier (e.g. 'stripe', 'quickbooks')"),
            ("source_id",         "string",   "Unique transaction ID from the source system"),
            ("amount",            "currency", "Transaction amount normalised to decimal USD"),
            ("currency",          "string",   "ISO 4217 currency code (default: USD)"),
            ("period",            "date",     "Transaction period in YYYY-MM format"),
            ("customer_id",       "string",   "Customer reference linking to canonical_customers"),
            ("subscription_type", "string",   "Revenue type: recurring, one-time, usage-based"),
            ("product_id",        "string",   "Product or SKU identifier if available"),
            ("recognized_at",     "date",     "Revenue recognition date (ISO 8601)"),
        ],
    },
    "canonical_customers": {
        "description": "Normalised customer records from CRM and billing systems",
        "fields": [
            ("source",          "string", "Source system identifier"),
            ("source_id",       "string", "Unique customer ID from the source system"),
            ("name",            "string", "Customer display name"),
            ("email",           "string", "Primary email address"),
            ("company",         "string", "Company or organisation name"),
            ("phone",           "string", "Primary phone number"),
            ("country",         "string", "ISO 3166-1 country code"),
            ("created_at",      "date",   "Customer creation/acquisition date (ISO 8601)"),
            ("lifecycle_stage", "string", "Lifecycle stage: lead, prospect, customer, churned"),
        ],
    },
    "canonical_pipeline": {
        "description": "Normalised sales pipeline / deal records from CRM systems",
        "fields": [
            ("source",      "string",   "Source system identifier"),
            ("source_id",   "string",   "Unique deal/opportunity ID from the source system"),
            ("name",        "string",   "Deal or opportunity name"),
            ("amount",      "currency", "Deal value in decimal USD"),
            ("stage",       "string",   "Pipeline stage (e.g. prospecting, negotiation, closed won)"),
            ("close_date",  "date",     "Expected or actual close date (ISO 8601)"),
            ("probability", "number",   "Win probability as decimal 0-1"),
            ("owner",       "string",   "Sales rep or deal owner identifier"),
            ("created_at",  "date",     "Deal creation date (ISO 8601)"),
        ],
    },
    "canonical_employees": {
        "description": "Normalised employee / headcount records from HRIS and accounting systems",
        "fields": [
            ("source",     "string",   "Source system identifier"),
            ("source_id",  "string",   "Unique employee ID from the source system"),
            ("name",       "string",   "Employee full name"),
            ("email",      "string",   "Work email address"),
            ("title",      "string",   "Job title"),
            ("department", "string",   "Department or team name"),
            ("salary",     "currency", "Monthly or annual salary in decimal USD"),
            ("hire_date",  "date",     "Hire or start date (ISO 8601)"),
            ("status",     "string",   "Employment status: active, terminated, contractor, part-time"),
        ],
    },
    "canonical_expenses": {
        "description": "Normalised expense transactions from accounting and spend management systems",
        "fields": [
            ("source",      "string",   "Source system identifier"),
            ("source_id",   "string",   "Unique transaction ID from the source system"),
            ("amount",      "currency", "Expense amount normalised to decimal USD (positive)"),
            ("currency",    "string",   "ISO 4217 currency code (default: USD)"),
            ("category",    "string",   "Expense category: cogs, s&m, g&a, r&d, hosting, etc."),
            ("vendor",      "string",   "Vendor or merchant name"),
            ("period",      "date",     "Expense period in YYYY-MM format"),
            ("description", "string",   "Line-item description or memo"),
        ],
    },
    "canonical_invoices": {
        "description": "Normalised accounts receivable invoice records",
        "fields": [
            ("source",      "string",   "Source system identifier"),
            ("source_id",   "string",   "Unique invoice ID from the source system"),
            ("amount",      "currency", "Invoice amount normalised to decimal USD"),
            ("currency",    "string",   "ISO 4217 currency code (default: USD)"),
            ("customer_id", "string",   "Customer reference linking to canonical_customers"),
            ("issue_date",  "date",     "Invoice issue date (ISO 8601)"),
            ("due_date",    "date",     "Invoice due date (ISO 8601)"),
            ("status",      "string",   "Invoice status: draft, sent, paid, overdue, void"),
            ("period",      "date",     "Invoice period in YYYY-MM format"),
        ],
    },
}

# ─── KPI-to-Canonical Field Dependencies ───────────────────────────────────
# Maps each KPI to the canonical fields it depends on (table.field format)

KPI_FIELD_DEPS = {
    "revenue_growth":        ["canonical_revenue.amount", "canonical_revenue.period"],
    "arr_growth":            ["canonical_revenue.amount", "canonical_revenue.subscription_type", "canonical_revenue.period"],
    "recurring_revenue":     ["canonical_revenue.amount", "canonical_revenue.subscription_type"],
    "revenue_quality":       ["canonical_revenue.amount", "canonical_revenue.subscription_type"],
    "customer_concentration":["canonical_revenue.amount", "canonical_revenue.customer_id"],
    "avg_deal_size":         ["canonical_pipeline.amount", "canonical_pipeline.stage"],
    "expansion_rate":        ["canonical_revenue.amount", "canonical_revenue.customer_id", "canonical_revenue.period"],
    "gross_dollar_ret":      ["canonical_revenue.amount", "canonical_revenue.customer_id", "canonical_revenue.period"],
    "customer_ltv":          ["canonical_revenue.amount", "canonical_revenue.customer_id", "canonical_expenses.amount", "canonical_expenses.category"],
    "gross_margin":          ["canonical_revenue.amount", "canonical_expenses.amount", "canonical_expenses.category"],
    "operating_margin":      ["canonical_revenue.amount", "canonical_expenses.amount", "canonical_expenses.category"],
    "ebitda_margin":         ["canonical_revenue.amount", "canonical_expenses.amount", "canonical_expenses.category"],
    "contribution_margin":   ["canonical_revenue.amount", "canonical_expenses.amount", "canonical_expenses.category"],
    "opex_ratio":            ["canonical_revenue.amount", "canonical_expenses.amount"],
    "operating_leverage":    ["canonical_revenue.amount", "canonical_expenses.amount", "canonical_expenses.category"],
    "margin_volatility":     ["canonical_revenue.amount", "canonical_expenses.amount", "canonical_expenses.category"],
    "burn_multiple":         ["canonical_revenue.amount", "canonical_revenue.subscription_type", "canonical_expenses.amount"],
    "pricing_power_index":   ["canonical_revenue.amount", "canonical_revenue.customer_id", "canonical_revenue.period"],
    "churn_rate":            ["canonical_customers.source_id", "canonical_customers.created_at", "canonical_revenue.customer_id"],
    "nrr":                   ["canonical_revenue.amount", "canonical_revenue.customer_id", "canonical_revenue.period"],
    "logo_retention":        ["canonical_customers.source_id", "canonical_customers.created_at"],
    "customer_decay_slope":  ["canonical_customers.source_id", "canonical_customers.created_at"],
    "ltv_cac":               ["canonical_revenue.amount", "canonical_customers.source_id", "canonical_expenses.amount", "canonical_expenses.category"],
    "cac_payback":           ["canonical_revenue.amount", "canonical_customers.source_id", "canonical_expenses.amount", "canonical_expenses.category"],
    "sales_efficiency":      ["canonical_revenue.amount", "canonical_revenue.subscription_type", "canonical_expenses.amount", "canonical_expenses.category"],
    "headcount_eff":         ["canonical_revenue.amount", "canonical_employees.source_id", "canonical_employees.status"],
    "rev_per_employee":      ["canonical_revenue.amount", "canonical_employees.source_id", "canonical_employees.status"],
    "billable_utilization":  ["canonical_employees.source_id", "canonical_employees.status"],
    "dso":                   ["canonical_invoices.amount", "canonical_invoices.issue_date", "canonical_invoices.due_date", "canonical_invoices.status"],
    "ar_turnover":           ["canonical_invoices.amount", "canonical_invoices.issue_date", "canonical_revenue.amount"],
    "avg_collection_period": ["canonical_invoices.amount", "canonical_invoices.issue_date", "canonical_invoices.due_date"],
    "cash_conv_cycle":       ["canonical_invoices.amount", "canonical_invoices.issue_date", "canonical_invoices.due_date", "canonical_expenses.amount"],
    "cei":                   ["canonical_invoices.amount", "canonical_invoices.status", "canonical_invoices.issue_date"],
    "ar_aging_current":      ["canonical_invoices.amount", "canonical_invoices.due_date", "canonical_invoices.status"],
    "ar_aging_overdue":      ["canonical_invoices.amount", "canonical_invoices.due_date", "canonical_invoices.status"],
    "pipeline_conversion":   ["canonical_pipeline.amount", "canonical_pipeline.stage"],
    "win_rate":              ["canonical_pipeline.stage", "canonical_pipeline.source_id"],
    "growth_efficiency":     ["canonical_revenue.amount", "canonical_revenue.subscription_type", "canonical_expenses.amount"],
    "revenue_momentum":      ["canonical_revenue.amount", "canonical_revenue.period"],
    "revenue_fragility":     ["canonical_revenue.amount", "canonical_revenue.customer_id", "canonical_customers.source_id"],
    "burn_convexity":        ["canonical_revenue.amount", "canonical_expenses.amount"],
    "cpl":                   ["canonical_expenses.amount", "canonical_expenses.category", "canonical_customers.source_id"],
    "mql_sql_rate":          ["canonical_customers.lifecycle_stage", "canonical_customers.source_id"],
    "pipeline_velocity":     ["canonical_pipeline.amount", "canonical_pipeline.stage", "canonical_pipeline.close_date", "canonical_pipeline.created_at"],
    "quota_attainment":      ["canonical_pipeline.amount", "canonical_pipeline.stage", "canonical_pipeline.owner"],
    "marketing_roi":         ["canonical_revenue.amount", "canonical_expenses.amount", "canonical_expenses.category"],
    "product_nps":           [],
    "feature_adoption":      [],
    "activation_rate":       ["canonical_customers.created_at", "canonical_customers.lifecycle_stage"],
    "time_to_value":         ["canonical_customers.created_at"],
    "health_score":          ["canonical_customers.source_id", "canonical_revenue.customer_id"],
    "csat":                  [],
    "ramp_time":             ["canonical_employees.hire_date", "canonical_pipeline.owner", "canonical_pipeline.amount"],
    "support_volume":        [],
    "automation_rate":       [],
    "cash_runway":           ["canonical_revenue.amount", "canonical_expenses.amount"],
    "current_ratio":         ["canonical_invoices.amount", "canonical_expenses.amount"],
    "working_capital":       ["canonical_invoices.amount", "canonical_expenses.amount"],
    "contraction_rate":      ["canonical_revenue.amount", "canonical_revenue.customer_id", "canonical_revenue.period"],
    "payback_period":        ["canonical_revenue.amount", "canonical_expenses.amount"],
    "organic_traffic":       [],
    "brand_awareness":       [],
}


def _kpis_for_field(canonical_table, canonical_field):
    """Return comma-separated KPIs driven by a given canonical table.field."""
    key = f"{canonical_table}.{canonical_field}"
    kpis = sorted({k for k, deps in KPI_FIELD_DEPS.items() if key in deps})
    return ", ".join(kpis) if kpis else ""


# ─── System Registry with Field Specs ──────────────────────────────────────

def _f(entity, field, desc, dtype, required, freq, unit, canon_table, canon_field):
    """Shorthand helper to build a field dict."""
    return {
        "entity": entity, "source_field": field, "description": desc,
        "data_type": dtype, "required": required, "frequency": freq,
        "unit_of_ingest": unit, "canonical_table": canon_table,
        "canonical_field": canon_field,
        "kpis_driven": _kpis_for_field(canon_table, canon_field) if canon_table else "",
    }


SYSTEM_REGISTRY = [
    # ────────────────────── 1. STRIPE ──────────────────────────────────────
    {
        "key": "stripe", "name": "Stripe", "category": "Payments",
        "header_color": "635BFF", "has_connector": True,
        "fields": [
            _f("Charges", "id",          "Unique charge identifier",                      "string",   True,  "Transaction-level", "String (ch_xxx)",    "canonical_revenue",   "source_id"),
            _f("Charges", "amount",      "Charge amount in smallest currency unit",       "number",   True,  "Transaction-level", "Integer (cents)",    "canonical_revenue",   "amount"),
            _f("Charges", "currency",    "Three-letter ISO currency code",                "string",   True,  "Transaction-level", "ISO 4217 (e.g. usd)","canonical_revenue",  "currency"),
            _f("Charges", "customer",    "Stripe customer ID associated with charge",     "string",   True,  "Transaction-level", "String (cus_xxx)",   "canonical_revenue",   "customer_id"),
            _f("Charges", "created",     "Charge creation timestamp",                     "date",     True,  "Transaction-level", "Unix timestamp (s)", "canonical_revenue",   "period"),
            _f("Charges", "description", "Charge description or memo",                    "string",   False, "Transaction-level", "Free text",          "",                    ""),
            _f("Charges", "status",      "Charge status (succeeded, pending, failed)",     "string",   True,  "Transaction-level", "Enum string",        "",                    ""),
            _f("Customers", "id",        "Unique Stripe customer identifier",             "string",   True,  "Snapshot",          "String (cus_xxx)",   "canonical_customers", "source_id"),
            _f("Customers", "name",      "Customer full name",                            "string",   False, "Snapshot",          "Free text",          "canonical_customers", "name"),
            _f("Customers", "email",     "Customer email address",                        "string",   True,  "Snapshot",          "Email address",      "canonical_customers", "email"),
            _f("Customers", "description","Customer description or notes",                "string",   False, "Snapshot",          "Free text",          "",                    ""),
            _f("Customers", "address",   "Customer billing address object",               "object",   False, "Snapshot",          "JSON object",        "canonical_customers", "country"),
            _f("Customers", "phone",     "Customer phone number",                         "string",   False, "Snapshot",          "Phone string",       "canonical_customers", "phone"),
            _f("Customers", "created",   "Customer creation timestamp",                   "date",     True,  "Snapshot",          "Unix timestamp (s)", "canonical_customers", "created_at"),
            _f("Invoices", "id",         "Unique invoice identifier",                     "string",   True,  "Transaction-level", "String (in_xxx)",    "canonical_invoices",  "source_id"),
            _f("Invoices", "amount_due", "Total amount due on invoice",                   "number",   True,  "Transaction-level", "Integer (cents)",    "canonical_invoices",  "amount"),
            _f("Invoices", "amount_paid","Total amount already paid",                     "number",   True,  "Transaction-level", "Integer (cents)",    "",                    ""),
            _f("Invoices", "currency",   "Invoice currency",                              "string",   True,  "Transaction-level", "ISO 4217",           "canonical_invoices",  "currency"),
            _f("Invoices", "customer",   "Customer ID for this invoice",                  "string",   True,  "Transaction-level", "String (cus_xxx)",   "canonical_invoices",  "customer_id"),
            _f("Invoices", "date",       "Invoice issue date",                            "date",     True,  "Transaction-level", "Unix timestamp (s)", "canonical_invoices",  "issue_date"),
            _f("Invoices", "due_date",   "Invoice payment due date",                      "date",     True,  "Transaction-level", "Unix timestamp (s)", "canonical_invoices",  "due_date"),
            _f("Invoices", "status",     "Invoice status (draft, open, paid, void, etc.)", "string",  True,  "Transaction-level", "Enum string",        "canonical_invoices",  "status"),
            _f("Subscriptions", "id",          "Unique subscription identifier",           "string",  True,  "Transaction-level", "String (sub_xxx)",   "canonical_revenue",   "source_id"),
            _f("Subscriptions", "customer",    "Customer ID for this subscription",        "string",  True,  "Transaction-level", "String (cus_xxx)",   "canonical_revenue",   "customer_id"),
            _f("Subscriptions", "amount",      "Subscription amount per billing period",   "number",  True,  "Transaction-level", "Integer (cents)",    "canonical_revenue",   "amount"),
            _f("Subscriptions", "status",      "Subscription status (active, canceled, past_due)", "string", True, "Transaction-level", "Enum string", "",                    ""),
            _f("Subscriptions", "current_period_start", "Current billing period start",    "date",    True,  "Transaction-level", "Unix timestamp (s)", "canonical_revenue",   "period"),
            _f("Subscriptions", "current_period_end",   "Current billing period end",      "date",    True,  "Transaction-level", "Unix timestamp (s)", "",                    ""),
            _f("Subscriptions", "created",     "Subscription creation date",               "date",    True,  "Transaction-level", "Unix timestamp (s)", "canonical_revenue",   "recognized_at"),
        ],
    },

    # ────────────────────── 2. QUICKBOOKS ──────────────────────────────────
    {
        "key": "quickbooks", "name": "QuickBooks", "category": "Accounting",
        "header_color": "2CA01C", "has_connector": True,
        "fields": [
            _f("Payment",  "Id",             "Unique payment record ID",                    "string",   True,  "Transaction-level", "Integer ID",         "canonical_revenue",   "source_id"),
            _f("Payment",  "TxnDate",        "Payment transaction date",                    "date",     True,  "Transaction-level", "YYYY-MM-DD",         "canonical_revenue",   "period"),
            _f("Payment",  "Line.Amount",    "Payment line item amount",                    "currency", True,  "Transaction-level", "Decimal USD",        "canonical_revenue",   "amount"),
            _f("Payment",  "CustomerRef",    "Customer reference ID",                       "string",   True,  "Transaction-level", "Integer ID",         "canonical_revenue",   "customer_id"),
            _f("Payment",  "Balance",        "Remaining unapplied balance",                 "currency", False, "Transaction-level", "Decimal USD",        "",                    ""),
            _f("Invoice",  "Id",             "Unique invoice record ID",                    "string",   True,  "Transaction-level", "Integer ID",         "canonical_invoices",  "source_id"),
            _f("Invoice",  "DocNumber",      "User-visible invoice number",                 "string",   False, "Transaction-level", "String",             "",                    ""),
            _f("Invoice",  "TxnDate",        "Invoice transaction date",                    "date",     True,  "Transaction-level", "YYYY-MM-DD",         "canonical_invoices",  "issue_date"),
            _f("Invoice",  "DueDate",        "Invoice payment due date",                    "date",     True,  "Transaction-level", "YYYY-MM-DD",         "canonical_invoices",  "due_date"),
            _f("Invoice",  "CustomerRef",    "Customer reference ID",                       "string",   True,  "Transaction-level", "Integer ID",         "canonical_invoices",  "customer_id"),
            _f("Invoice",  "TotalAmt",       "Total invoice amount",                        "currency", True,  "Transaction-level", "Decimal USD",        "canonical_invoices",  "amount"),
            _f("Invoice",  "Balance",        "Unpaid balance on invoice",                   "currency", True,  "Transaction-level", "Decimal USD",        "",                    ""),
            _f("Customer", "Id",             "Unique customer record ID",                   "string",   True,  "Snapshot",          "Integer ID",         "canonical_customers", "source_id"),
            _f("Customer", "DisplayName",    "Customer display name",                       "string",   True,  "Snapshot",          "Free text",          "canonical_customers", "name"),
            _f("Customer", "PrimaryEmailAddr","Primary email address",                      "string",   False, "Snapshot",          "Email address",      "canonical_customers", "email"),
            _f("Customer", "BillingAddr",    "Customer billing address object",             "object",   False, "Snapshot",          "JSON object",        "canonical_customers", "country"),
            _f("Customer", "MetaData.CreateTime","Customer creation timestamp",             "date",     True,  "Snapshot",          "ISO 8601",           "canonical_customers", "created_at"),
            _f("Purchase", "Id",             "Unique purchase/expense record ID",           "string",   True,  "Transaction-level", "Integer ID",         "canonical_expenses",  "source_id"),
            _f("Purchase", "TxnDate",        "Purchase transaction date",                   "date",     True,  "Transaction-level", "YYYY-MM-DD",         "canonical_expenses",  "period"),
            _f("Purchase", "VendorRef",      "Vendor reference ID",                         "string",   True,  "Transaction-level", "Integer ID",         "canonical_expenses",  "vendor"),
            _f("Purchase", "Line.Amount",    "Purchase line item amount",                   "currency", True,  "Transaction-level", "Decimal USD",        "canonical_expenses",  "amount"),
            _f("Purchase", "Line.Description","Line item description",                      "string",   False, "Transaction-level", "Free text",          "canonical_expenses",  "description"),
            _f("Purchase", "AccountRef.name","Expense account name (used for categorisation)", "string", True, "Transaction-level", "Account name",       "canonical_expenses",  "category"),
            _f("Employee", "Id",             "Unique employee record ID",                   "string",   True,  "Snapshot",          "Integer ID",         "canonical_employees", "source_id"),
            _f("Employee", "DisplayName",    "Employee full name",                          "string",   True,  "Snapshot",          "Free text",          "canonical_employees", "name"),
            _f("Employee", "PrimaryEmailAddr","Employee email address",                     "string",   False, "Snapshot",          "Email address",      "canonical_employees", "email"),
            _f("Employee", "PrimaryPhone",   "Employee phone number",                       "string",   False, "Snapshot",          "Phone string",       "",                    ""),
            _f("Employee", "HiredDate",      "Employee hire date",                          "date",     True,  "Snapshot",          "YYYY-MM-DD",         "canonical_employees", "hire_date"),
            _f("Account",  "Id",             "Unique chart of accounts record ID",          "string",   True,  "Snapshot",          "Integer ID",         "",                    ""),
            _f("Account",  "Name",           "Account name",                                "string",   True,  "Snapshot",          "Free text",          "",                    ""),
            _f("Account",  "AccountType",    "Account type (Income, Expense, Asset, etc.)", "string",   True,  "Snapshot",          "Enum string",        "",                    ""),
            _f("Account",  "CurrentBalance", "Current account balance",                     "currency", True,  "Snapshot",          "Decimal USD",        "",                    ""),
        ],
    },

    # ────────────────────── 3. XERO ────────────────────────────────────────
    {
        "key": "xero", "name": "Xero", "category": "Accounting",
        "header_color": "13B5EA", "has_connector": True,
        "fields": [
            _f("Invoices", "InvoiceID",     "Unique Xero invoice ID",                      "string",   True,  "Transaction-level", "UUID string",        "canonical_invoices",  "source_id"),
            _f("Invoices", "InvoiceNumber", "User-visible invoice number",                  "string",   False, "Transaction-level", "String",             "",                    ""),
            _f("Invoices", "InvoiceDate",   "Invoice issue date",                           "date",     True,  "Transaction-level", "YYYY-MM-DD",         "canonical_invoices",  "issue_date"),
            _f("Invoices", "DueDate",       "Invoice payment due date",                     "date",     True,  "Transaction-level", "YYYY-MM-DD",         "canonical_invoices",  "due_date"),
            _f("Invoices", "Contact.Name",  "Customer name from contact reference",         "string",   True,  "Transaction-level", "Free text",          "canonical_invoices",  "customer_id"),
            _f("Invoices", "Total",         "Invoice total amount",                         "currency", True,  "Transaction-level", "Decimal (2dp)",      "canonical_invoices",  "amount"),
            _f("Invoices", "Status",        "Invoice status (DRAFT, SUBMITTED, AUTHORISED, PAID, VOIDED)", "string", True, "Transaction-level", "Enum string", "canonical_invoices", "status"),
            _f("Invoices", "CurrencyCode",  "Invoice currency",                             "string",   True,  "Transaction-level", "ISO 4217",           "canonical_invoices",  "currency"),
            _f("Invoices", "AmountDue",     "Outstanding amount due",                       "currency", True,  "Transaction-level", "Decimal (2dp)",      "",                    ""),
            _f("Contacts", "ContactID",     "Unique Xero contact ID",                       "string",   True,  "Snapshot",          "UUID string",        "canonical_customers", "source_id"),
            _f("Contacts", "Name",          "Contact display name",                          "string",   True,  "Snapshot",          "Free text",          "canonical_customers", "name"),
            _f("Contacts", "EmailAddress",  "Contact email address",                         "string",   False, "Snapshot",          "Email address",      "canonical_customers", "email"),
            _f("Contacts", "ContactStatus", "Contact status (ACTIVE, ARCHIVED)",             "string",   True,  "Snapshot",          "Enum string",        "canonical_customers", "lifecycle_stage"),
            _f("Contacts", "IsCustomer",    "Whether contact is a customer",                 "boolean",  True,  "Snapshot",          "Boolean",            "",                    ""),
            _f("Payments", "PaymentID",     "Unique payment ID",                             "string",   True,  "Transaction-level", "UUID string",        "canonical_revenue",   "source_id"),
            _f("Payments", "Date",          "Payment date",                                  "date",     True,  "Transaction-level", "YYYY-MM-DD",         "canonical_revenue",   "period"),
            _f("Payments", "Amount",        "Payment amount",                                "currency", True,  "Transaction-level", "Decimal (2dp)",      "canonical_revenue",   "amount"),
            _f("Payments", "InvoiceNumber", "Related invoice number",                        "string",   True,  "Transaction-level", "String",             "",                    ""),
            _f("Payments", "Status",        "Payment status (AUTHORISED, DELETED)",           "string",   True,  "Transaction-level", "Enum string",        "",                    ""),
            _f("Accounts", "AccountID",     "Unique account ID",                             "string",   True,  "Snapshot",          "UUID string",        "",                    ""),
            _f("Accounts", "Code",          "Account code",                                  "string",   True,  "Snapshot",          "String",             "",                    ""),
            _f("Accounts", "Name",          "Account name",                                  "string",   True,  "Snapshot",          "Free text",          "",                    ""),
            _f("Accounts", "Type",          "Account type (REVENUE, EXPENSE, ASSET, etc.)",  "string",   True,  "Snapshot",          "Enum string",        "",                    ""),
        ],
    },

    # ────────────────────── 4. NETSUITE ────────────────────────────────────
    {
        "key": "netsuite", "name": "NetSuite", "category": "ERP",
        "header_color": "1A3E72", "has_connector": True,
        "fields": [
            _f("Invoice",    "id",          "Internal NetSuite invoice ID",                 "string",   True,  "Transaction-level", "Integer ID",         "canonical_invoices",  "source_id"),
            _f("Invoice",    "tranDate",    "Transaction date",                             "date",     True,  "Transaction-level", "MM/DD/YYYY",         "canonical_invoices",  "issue_date"),
            _f("Invoice",    "entity",      "Customer entity reference",                    "string",   True,  "Transaction-level", "Integer ID",         "canonical_invoices",  "customer_id"),
            _f("Invoice",    "amount",      "Invoice total amount",                         "currency", True,  "Transaction-level", "Decimal USD",        "canonical_invoices",  "amount"),
            _f("Invoice",    "status",      "Invoice status",                               "string",   True,  "Transaction-level", "Enum string",        "canonical_invoices",  "status"),
            _f("Invoice",    "currency",    "Transaction currency",                         "string",   True,  "Transaction-level", "ISO 4217",           "canonical_invoices",  "currency"),
            _f("Invoice",    "dueDate",     "Invoice due date",                             "date",     True,  "Transaction-level", "MM/DD/YYYY",         "canonical_invoices",  "due_date"),
            _f("Invoice",    "terms",       "Payment terms (Net 30, Net 60, etc.)",         "string",   False, "Transaction-level", "String",             "",                    ""),
            _f("Invoice",    "subsidiary",  "Subsidiary entity",                            "string",   False, "Transaction-level", "Integer ID",         "",                    ""),
            _f("VendorBill", "id",          "Internal NetSuite vendor bill ID",             "string",   True,  "Transaction-level", "Integer ID",         "canonical_expenses",  "source_id"),
            _f("VendorBill", "tranDate",    "Transaction date",                             "date",     True,  "Transaction-level", "MM/DD/YYYY",         "canonical_expenses",  "period"),
            _f("VendorBill", "entity",      "Vendor entity reference",                      "string",   True,  "Transaction-level", "Integer ID",         "canonical_expenses",  "vendor"),
            _f("VendorBill", "amount",      "Bill total amount",                            "currency", True,  "Transaction-level", "Decimal USD",        "canonical_expenses",  "amount"),
            _f("VendorBill", "status",      "Bill status",                                  "string",   True,  "Transaction-level", "Enum string",        "",                    ""),
            _f("VendorBill", "account",     "Expense account reference",                    "string",   True,  "Transaction-level", "Integer ID",         "canonical_expenses",  "category"),
            _f("VendorBill", "memo",        "Bill memo or description",                     "string",   False, "Transaction-level", "Free text",          "canonical_expenses",  "description"),
            _f("Customer",   "id",          "Internal NetSuite customer ID",                "string",   True,  "Snapshot",          "Integer ID",         "canonical_customers", "source_id"),
            _f("Customer",   "entityId",    "Customer entity display ID",                   "string",   True,  "Snapshot",          "String",             "",                    ""),
            _f("Customer",   "companyName", "Company name",                                 "string",   True,  "Snapshot",          "Free text",          "canonical_customers", "company"),
            _f("Customer",   "email",       "Customer email address",                       "string",   False, "Snapshot",          "Email address",      "canonical_customers", "email"),
            _f("Customer",   "dateCreated", "Customer creation date",                       "date",     True,  "Snapshot",          "MM/DD/YYYY",         "canonical_customers", "created_at"),
            _f("Customer",   "balance",     "Outstanding AR balance",                       "currency", False, "Snapshot",          "Decimal USD",        "",                    ""),
            _f("Customer",   "salesRep",    "Assigned sales rep reference",                 "string",   False, "Snapshot",          "Integer ID",         "",                    ""),
            _f("Employee",   "id",          "Internal NetSuite employee ID",                "string",   True,  "Snapshot",          "Integer ID",         "canonical_employees", "source_id"),
            _f("Employee",   "entityId",    "Employee entity display ID",                   "string",   True,  "Snapshot",          "String",             "canonical_employees", "name"),
            _f("Employee",   "email",       "Employee email address",                       "string",   False, "Snapshot",          "Email address",      "canonical_employees", "email"),
            _f("Employee",   "title",       "Job title",                                    "string",   False, "Snapshot",          "Free text",          "canonical_employees", "title"),
            _f("Employee",   "department",  "Department reference",                         "string",   False, "Snapshot",          "Integer ID",         "canonical_employees", "department"),
            _f("Employee",   "hireDate",    "Employee hire date",                           "date",     True,  "Snapshot",          "MM/DD/YYYY",         "canonical_employees", "hire_date"),
            _f("Employee",   "isInactive",  "Whether employee is inactive",                 "boolean",  True,  "Snapshot",          "Boolean",            "canonical_employees", "status"),
        ],
    },

    # ────────────────────── 5. SAGE INTACCT ────────────────────────────────
    {
        "key": "sage_intacct", "name": "Sage Intacct", "category": "Accounting",
        "header_color": "00A651", "has_connector": True,
        "fields": [
            _f("ARINVOICE", "RECORDNO",     "Unique AR invoice record number",              "string",   True,  "Transaction-level", "Integer ID",         "canonical_invoices",  "source_id"),
            _f("ARINVOICE", "CUSTOMERID",   "Customer ID reference",                        "string",   True,  "Transaction-level", "String ID",          "canonical_invoices",  "customer_id"),
            _f("ARINVOICE", "CUSTOMERNAME", "Customer display name",                        "string",   False, "Transaction-level", "Free text",          "",                    ""),
            _f("ARINVOICE", "TOTALDUE",     "Total amount due",                             "currency", True,  "Transaction-level", "Decimal USD",        "canonical_invoices",  "amount"),
            _f("ARINVOICE", "TOTALENTERED", "Total amount originally entered",              "currency", True,  "Transaction-level", "Decimal USD",        "",                    ""),
            _f("ARINVOICE", "WHENDUE",      "Invoice due date",                             "date",     True,  "Transaction-level", "MM/DD/YYYY",         "canonical_invoices",  "due_date"),
            _f("ARINVOICE", "WHENCREATED",  "Invoice creation date",                        "date",     True,  "Transaction-level", "MM/DD/YYYY",         "canonical_invoices",  "issue_date"),
            _f("ARINVOICE", "STATE",        "Invoice state (Posted, Draft, etc.)",           "string",   True,  "Transaction-level", "Enum string",        "canonical_invoices",  "status"),
            _f("ARINVOICE", "CURRENCY",     "Invoice currency",                              "string",   True,  "Transaction-level", "ISO 4217",           "canonical_invoices",  "currency"),
            _f("APBILL",    "RECORDNO",     "Unique AP bill record number",                  "string",   True,  "Transaction-level", "Integer ID",         "canonical_expenses",  "source_id"),
            _f("APBILL",    "VENDORID",     "Vendor ID reference",                           "string",   True,  "Transaction-level", "String ID",          "canonical_expenses",  "vendor"),
            _f("APBILL",    "VENDORNAME",   "Vendor display name",                           "string",   False, "Transaction-level", "Free text",          "",                    ""),
            _f("APBILL",    "TOTALDUE",     "Total bill amount due",                         "currency", True,  "Transaction-level", "Decimal USD",        "canonical_expenses",  "amount"),
            _f("APBILL",    "TOTALENTERED", "Total amount originally entered",               "currency", True,  "Transaction-level", "Decimal USD",        "",                    ""),
            _f("APBILL",    "WHENDUE",      "Bill due date",                                 "date",     True,  "Transaction-level", "MM/DD/YYYY",         "",                    ""),
            _f("APBILL",    "WHENCREATED",  "Bill creation date",                            "date",     True,  "Transaction-level", "MM/DD/YYYY",         "canonical_expenses",  "period"),
            _f("APBILL",    "STATE",        "Bill state (Posted, Draft, etc.)",               "string",   True,  "Transaction-level", "Enum string",        "",                    ""),
            _f("APBILL",    "GLPOSTINGDATE","GL posting date",                                "date",     False, "Transaction-level", "MM/DD/YYYY",         "",                    ""),
            _f("CUSTOMER",  "CUSTOMERID",   "Unique customer ID",                            "string",   True,  "Snapshot",          "String ID",          "canonical_customers", "source_id"),
            _f("CUSTOMER",  "NAME",         "Customer display name",                          "string",   True,  "Snapshot",          "Free text",          "canonical_customers", "name"),
            _f("CUSTOMER",  "WHENCREATED",  "Customer creation date",                         "date",     True,  "Snapshot",          "MM/DD/YYYY",         "canonical_customers", "created_at"),
            _f("CUSTOMER",  "STATUS",       "Customer status (Active, Inactive)",             "string",   True,  "Snapshot",          "Enum string",        "canonical_customers", "lifecycle_stage"),
            _f("CUSTOMER",  "TOTALDUE",     "Outstanding balance",                            "currency", False, "Snapshot",          "Decimal USD",        "",                    ""),
            _f("CUSTOMER",  "EMAIL",        "Customer email",                                 "string",   False, "Snapshot",          "Email address",      "canonical_customers", "email"),
        ],
    },

    # ────────────────────── 6. PEACHTREE (SAGE 50) ─────────────────────────
    {
        "key": "peachtree", "name": "Peachtree (Sage 50)", "category": "Accounting",
        "header_color": "F7941D", "has_connector": False,
        "fields": [
            _f("Sales Journal",    "TransactionID",     "Unique sales transaction ID",               "string",   True,  "Transaction-level", "Integer ID",         "canonical_revenue",   "source_id"),
            _f("Sales Journal",    "Date",              "Transaction date",                          "date",     True,  "Transaction-level", "MM/DD/YYYY",         "canonical_revenue",   "period"),
            _f("Sales Journal",    "CustomerID",        "Customer ID reference",                     "string",   True,  "Transaction-level", "String ID",          "canonical_revenue",   "customer_id"),
            _f("Sales Journal",    "Amount",            "Transaction total amount",                  "currency", True,  "Transaction-level", "Decimal USD",        "canonical_revenue",   "amount"),
            _f("Sales Journal",    "InvoiceNumber",     "Invoice number reference",                  "string",   False, "Transaction-level", "String",             "",                    ""),
            _f("Sales Journal",    "GLAccount",         "General ledger account code",               "string",   True,  "Transaction-level", "Account code",       "",                    ""),
            _f("Sales Journal",    "Description",       "Transaction description",                   "string",   False, "Transaction-level", "Free text",          "",                    ""),
            _f("Purchase Journal", "TransactionID",     "Unique purchase transaction ID",            "string",   True,  "Transaction-level", "Integer ID",         "canonical_expenses",  "source_id"),
            _f("Purchase Journal", "Date",              "Transaction date",                          "date",     True,  "Transaction-level", "MM/DD/YYYY",         "canonical_expenses",  "period"),
            _f("Purchase Journal", "VendorID",          "Vendor ID reference",                       "string",   True,  "Transaction-level", "String ID",          "canonical_expenses",  "vendor"),
            _f("Purchase Journal", "Amount",            "Purchase amount",                           "currency", True,  "Transaction-level", "Decimal USD",        "canonical_expenses",  "amount"),
            _f("Purchase Journal", "GLAccount",         "Expense GL account code",                   "string",   True,  "Transaction-level", "Account code",       "canonical_expenses",  "category"),
            _f("Purchase Journal", "Description",       "Purchase description",                      "string",   False, "Transaction-level", "Free text",          "canonical_expenses",  "description"),
            _f("Cash Receipts",    "ReceiptID",         "Unique cash receipt ID",                    "string",   True,  "Transaction-level", "Integer ID",         "canonical_invoices",  "source_id"),
            _f("Cash Receipts",    "Date",              "Receipt date",                              "date",     True,  "Transaction-level", "MM/DD/YYYY",         "canonical_invoices",  "issue_date"),
            _f("Cash Receipts",    "CustomerID",        "Customer ID reference",                     "string",   True,  "Transaction-level", "String ID",          "canonical_invoices",  "customer_id"),
            _f("Cash Receipts",    "Amount",            "Receipt amount",                            "currency", True,  "Transaction-level", "Decimal USD",        "canonical_invoices",  "amount"),
            _f("Cash Receipts",    "AppliedToInvoice",  "Invoice number payment is applied to",      "string",   False, "Transaction-level", "String",             "",                    ""),
            _f("Payroll",          "EmployeeID",        "Unique employee ID",                        "string",   True,  "Monthly",           "String ID",          "canonical_employees", "source_id"),
            _f("Payroll",          "EmployeeName",      "Employee full name",                        "string",   True,  "Monthly",           "Free text",          "canonical_employees", "name"),
            _f("Payroll",          "PayDate",           "Pay period end date",                       "date",     True,  "Monthly",           "MM/DD/YYYY",         "",                    ""),
            _f("Payroll",          "GrossPay",          "Gross pay amount",                          "currency", True,  "Monthly",           "Decimal USD",        "canonical_employees", "salary"),
            _f("Payroll",          "Department",        "Department assignment",                      "string",   False, "Monthly",           "String",             "canonical_employees", "department"),
            _f("Payroll",          "HireDate",          "Employee hire date",                        "date",     True,  "Snapshot",          "MM/DD/YYYY",         "canonical_employees", "hire_date"),
            _f("Payroll",          "Status",            "Employment status",                         "string",   True,  "Snapshot",          "Enum string",        "canonical_employees", "status"),
            _f("Chart of Accounts","AccountID",         "Unique account code",                       "string",   True,  "Snapshot",          "String",             "",                    ""),
            _f("Chart of Accounts","AccountName",       "Account display name",                      "string",   True,  "Snapshot",          "Free text",          "",                    ""),
            _f("Chart of Accounts","AccountType",       "Account type (Income, Expense, Asset, etc.)", "string", True,  "Snapshot",          "Enum string",        "",                    ""),
            _f("Chart of Accounts","Balance",           "Current account balance",                   "currency", True,  "Snapshot",          "Decimal USD",        "",                    ""),
        ],
    },

    # ────────────────────── 7. SALESFORCE ──────────────────────────────────
    {
        "key": "salesforce", "name": "Salesforce", "category": "CRM",
        "header_color": "00A1E0", "has_connector": True,
        "fields": [
            _f("Opportunity", "Id",           "Unique Salesforce opportunity ID",             "string",   True,  "Transaction-level", "18-char SFID",       "canonical_pipeline",  "source_id"),
            _f("Opportunity", "Name",         "Opportunity name",                             "string",   True,  "Transaction-level", "Free text",          "canonical_pipeline",  "name"),
            _f("Opportunity", "Amount",       "Opportunity value",                            "currency", True,  "Transaction-level", "Decimal USD",        "canonical_pipeline",  "amount"),
            _f("Opportunity", "StageName",    "Pipeline stage name",                          "string",   True,  "Transaction-level", "Picklist value",     "canonical_pipeline",  "stage"),
            _f("Opportunity", "CloseDate",    "Expected or actual close date",                "date",     True,  "Transaction-level", "YYYY-MM-DD",         "canonical_pipeline",  "close_date"),
            _f("Opportunity", "CreatedDate",  "Opportunity creation date",                    "date",     True,  "Transaction-level", "ISO 8601",           "canonical_pipeline",  "created_at"),
            _f("Opportunity", "OwnerId",      "Opportunity owner (sales rep) reference",      "string",   True,  "Transaction-level", "18-char SFID",       "canonical_pipeline",  "owner"),
            _f("Opportunity", "AccountId",    "Related account ID",                           "string",   True,  "Transaction-level", "18-char SFID",       "",                    ""),
            _f("Opportunity", "Probability",  "Win probability percentage",                   "number",   True,  "Transaction-level", "Percentage (0-100)", "canonical_pipeline",  "probability"),
            _f("Opportunity", "Type",         "Opportunity type (New, Expansion, Renewal)",   "string",   False, "Transaction-level", "Picklist value",     "",                    ""),
            _f("Opportunity", "LeadSource",   "Lead source channel",                          "string",   False, "Transaction-level", "Picklist value",     "",                    ""),
            _f("Opportunity", "ForecastCategory","Forecast category",                         "string",   False, "Transaction-level", "Picklist value",     "",                    ""),
            _f("Contact",     "Id",           "Unique Salesforce contact ID",                 "string",   True,  "Snapshot",          "18-char SFID",       "canonical_customers", "source_id"),
            _f("Contact",     "Name",         "Contact full name",                            "string",   True,  "Snapshot",          "Free text",          "canonical_customers", "name"),
            _f("Contact",     "Email",        "Contact email address",                        "string",   True,  "Snapshot",          "Email address",      "canonical_customers", "email"),
            _f("Contact",     "Phone",        "Contact phone number",                         "string",   False, "Snapshot",          "Phone string",       "canonical_customers", "phone"),
            _f("Contact",     "Title",        "Job title",                                    "string",   False, "Snapshot",          "Free text",          "",                    ""),
            _f("Contact",     "AccountId",    "Related account ID",                           "string",   True,  "Snapshot",          "18-char SFID",       "canonical_customers", "company"),
            _f("Contact",     "CreatedDate",  "Contact creation date",                        "date",     True,  "Snapshot",          "ISO 8601",           "canonical_customers", "created_at"),
            _f("Contact",     "LeadSource",   "Original lead source",                         "string",   False, "Snapshot",          "Picklist value",     "",                    ""),
            _f("Account",     "Id",           "Unique Salesforce account ID",                 "string",   True,  "Snapshot",          "18-char SFID",       "canonical_customers", "source_id"),
            _f("Account",     "Name",         "Account (company) name",                       "string",   True,  "Snapshot",          "Free text",          "canonical_customers", "company"),
            _f("Account",     "Industry",     "Industry classification",                      "string",   False, "Snapshot",          "Picklist value",     "",                    ""),
            _f("Account",     "NumberOfEmployees","Company headcount",                        "number",   False, "Snapshot",          "Integer count",      "",                    ""),
            _f("Account",     "AnnualRevenue","Company annual revenue",                       "currency", False, "Snapshot",          "Decimal USD",        "",                    ""),
            _f("Account",     "BillingCountry","Billing country",                             "string",   False, "Snapshot",          "Country name/code",  "canonical_customers", "country"),
            _f("Account",     "CreatedDate",  "Account creation date",                        "date",     True,  "Snapshot",          "ISO 8601",           "",                    ""),
        ],
    },

    # ────────────────────── 8. HUBSPOT ─────────────────────────────────────
    {
        "key": "hubspot", "name": "HubSpot", "category": "CRM",
        "header_color": "FF7A59", "has_connector": True,
        "fields": [
            _f("Contacts",  "id",              "Unique HubSpot contact ID",                   "string",   True,  "Snapshot",          "Integer ID",         "canonical_customers", "source_id"),
            _f("Contacts",  "firstname",       "Contact first name",                           "string",   True,  "Snapshot",          "Free text",          "canonical_customers", "name"),
            _f("Contacts",  "lastname",        "Contact last name",                            "string",   True,  "Snapshot",          "Free text",          "canonical_customers", "name"),
            _f("Contacts",  "email",           "Contact email address",                        "string",   True,  "Snapshot",          "Email address",      "canonical_customers", "email"),
            _f("Contacts",  "company",         "Associated company name",                      "string",   False, "Snapshot",          "Free text",          "canonical_customers", "company"),
            _f("Contacts",  "phone",           "Contact phone number",                         "string",   False, "Snapshot",          "Phone string",       "canonical_customers", "phone"),
            _f("Contacts",  "createdate",      "Contact creation date",                        "date",     True,  "Snapshot",          "Unix timestamp (ms)","canonical_customers", "created_at"),
            _f("Contacts",  "hs_lead_status",  "Lead status (New, Open, In Progress, etc.)",   "string",   True,  "Snapshot",          "Enum string",        "canonical_customers", "lifecycle_stage"),
            _f("Contacts",  "lifecyclestage",  "Lifecycle stage (subscriber, lead, MQL, SQL, customer)", "string", True, "Snapshot", "Enum string",        "canonical_customers", "lifecycle_stage"),
            _f("Contacts",  "hs_analytics_source","Original traffic source",                   "string",   False, "Snapshot",          "Enum string",        "",                    ""),
            _f("Deals",     "id",              "Unique HubSpot deal ID",                       "string",   True,  "Transaction-level", "Integer ID",         "canonical_pipeline",  "source_id"),
            _f("Deals",     "dealname",        "Deal name",                                    "string",   True,  "Transaction-level", "Free text",          "canonical_pipeline",  "name"),
            _f("Deals",     "amount",          "Deal value",                                   "currency", True,  "Transaction-level", "Decimal USD",        "canonical_pipeline",  "amount"),
            _f("Deals",     "dealstage",       "Deal stage ID",                                "string",   True,  "Transaction-level", "Stage ID string",    "canonical_pipeline",  "stage"),
            _f("Deals",     "closedate",       "Expected or actual close date",                "date",     True,  "Transaction-level", "Unix timestamp (ms)","canonical_pipeline",  "close_date"),
            _f("Deals",     "createdate",      "Deal creation date",                           "date",     True,  "Transaction-level", "Unix timestamp (ms)","canonical_pipeline",  "created_at"),
            _f("Deals",     "pipeline",        "Pipeline ID (for multi-pipeline setups)",      "string",   True,  "Transaction-level", "Pipeline ID string", "",                    ""),
            _f("Deals",     "hubspot_owner_id","Deal owner (sales rep) ID",                    "string",   True,  "Transaction-level", "Integer ID",         "canonical_pipeline",  "owner"),
            _f("Deals",     "hs_deal_stage_probability","Stage-based win probability",         "number",   False, "Transaction-level", "Decimal (0-1)",      "canonical_pipeline",  "probability"),
            _f("Companies", "id",              "Unique HubSpot company ID",                    "string",   True,  "Snapshot",          "Integer ID",         "canonical_customers", "source_id"),
            _f("Companies", "name",            "Company name",                                 "string",   True,  "Snapshot",          "Free text",          "canonical_customers", "company"),
            _f("Companies", "domain",          "Company website domain",                       "string",   False, "Snapshot",          "Domain string",      "",                    ""),
            _f("Companies", "industry",        "Industry classification",                      "string",   False, "Snapshot",          "Free text",          "",                    ""),
            _f("Companies", "numberofemployees","Company headcount",                           "number",   False, "Snapshot",          "Integer count",      "",                    ""),
            _f("Companies", "annualrevenue",   "Company annual revenue",                       "currency", False, "Snapshot",          "Decimal USD",        "",                    ""),
            _f("Companies", "createdate",      "Company creation date",                        "date",     True,  "Snapshot",          "Unix timestamp (ms)","canonical_customers", "created_at"),
        ],
    },

    # ────────────────────── 9. SHOPIFY ─────────────────────────────────────
    {
        "key": "shopify", "name": "Shopify", "category": "E-commerce",
        "header_color": "96BF48", "has_connector": True,
        "fields": [
            _f("Orders",    "id",              "Unique Shopify order ID",                      "string",   True,  "Transaction-level", "Integer ID",         "canonical_revenue",   "source_id"),
            _f("Orders",    "created_at",      "Order creation timestamp",                     "date",     True,  "Transaction-level", "ISO 8601",           "canonical_revenue",   "period"),
            _f("Orders",    "total_price",     "Order total amount",                           "currency", True,  "Transaction-level", "Decimal string",     "canonical_revenue",   "amount"),
            _f("Orders",    "currency",        "Order currency",                               "string",   True,  "Transaction-level", "ISO 4217",           "canonical_revenue",   "currency"),
            _f("Orders",    "customer.id",     "Customer ID reference",                        "string",   True,  "Transaction-level", "Integer ID",         "canonical_revenue",   "customer_id"),
            _f("Orders",    "customer.email",  "Customer email from order",                    "string",   False, "Transaction-level", "Email address",      "",                    ""),
            _f("Orders",    "financial_status", "Payment status (paid, pending, refunded)",     "string",   True,  "Transaction-level", "Enum string",        "",                    ""),
            _f("Orders",    "line_items[].price","Line item unit price",                       "currency", True,  "Transaction-level", "Decimal string",     "",                    ""),
            _f("Orders",    "line_items[].quantity","Line item quantity",                       "number",   True,  "Transaction-level", "Integer count",      "",                    ""),
            _f("Orders",    "line_items[].sku", "Line item SKU",                               "string",   False, "Transaction-level", "String",             "canonical_revenue",   "product_id"),
            _f("Orders",    "line_items[].title","Product title",                              "string",   False, "Transaction-level", "Free text",          "",                    ""),
            _f("Orders",    "discount_codes",  "Applied discount codes",                       "string",   False, "Transaction-level", "JSON array",         "",                    ""),
            _f("Customers", "id",              "Unique Shopify customer ID",                   "string",   True,  "Snapshot",          "Integer ID",         "canonical_customers", "source_id"),
            _f("Customers", "email",           "Customer email address",                       "string",   True,  "Snapshot",          "Email address",      "canonical_customers", "email"),
            _f("Customers", "first_name",      "Customer first name",                          "string",   True,  "Snapshot",          "Free text",          "canonical_customers", "name"),
            _f("Customers", "last_name",       "Customer last name",                           "string",   True,  "Snapshot",          "Free text",          "canonical_customers", "name"),
            _f("Customers", "created_at",      "Customer creation date",                       "date",     True,  "Snapshot",          "ISO 8601",           "canonical_customers", "created_at"),
            _f("Customers", "total_spent",     "Lifetime total spend",                         "currency", False, "Snapshot",          "Decimal string",     "",                    ""),
            _f("Customers", "orders_count",    "Total number of orders",                       "number",   False, "Snapshot",          "Integer count",      "",                    ""),
            _f("Products",  "id",              "Unique Shopify product ID",                    "string",   True,  "Snapshot",          "Integer ID",         "",                    ""),
            _f("Products",  "title",           "Product title",                                "string",   True,  "Snapshot",          "Free text",          "",                    ""),
            _f("Products",  "created_at",      "Product creation date",                        "date",     True,  "Snapshot",          "ISO 8601",           "",                    ""),
            _f("Products",  "variants[].price","Variant price",                                "currency", True,  "Snapshot",          "Decimal string",     "",                    ""),
            _f("Products",  "variants[].sku",  "Variant SKU",                                  "string",   False, "Snapshot",          "String",             "",                    ""),
        ],
    },

    # ────────────────────── 10. BREX ───────────────────────────────────────
    {
        "key": "brex", "name": "Brex", "category": "Spend Management",
        "header_color": "000000", "has_connector": True,
        "fields": [
            _f("Transactions", "id",            "Unique Brex transaction ID",                  "string",   True,  "Transaction-level", "UUID string",        "canonical_expenses",  "source_id"),
            _f("Transactions", "amount",        "Transaction amount",                          "currency", True,  "Transaction-level", "Integer (cents)",    "canonical_expenses",  "amount"),
            _f("Transactions", "merchant_name", "Merchant or vendor name",                     "string",   True,  "Transaction-level", "Free text",          "canonical_expenses",  "vendor"),
            _f("Transactions", "posted_date",   "Transaction posted date",                     "date",     True,  "Transaction-level", "YYYY-MM-DD",         "canonical_expenses",  "period"),
            _f("Transactions", "user_id",       "Card holder user ID",                         "string",   True,  "Transaction-level", "UUID string",        "",                    ""),
            _f("Transactions", "status",        "Transaction status (APPROVED, PENDING, etc.)", "string",  True,  "Transaction-level", "Enum string",        "",                    ""),
            _f("Transactions", "category",      "Expense category",                            "string",   False, "Transaction-level", "String",             "canonical_expenses",  "category"),
            _f("Transactions", "memo",          "Transaction memo or note",                    "string",   False, "Transaction-level", "Free text",          "canonical_expenses",  "description"),
            _f("Users",        "id",            "Unique Brex user ID",                         "string",   True,  "Snapshot",          "UUID string",        "canonical_employees", "source_id"),
            _f("Users",        "firstName",     "User first name",                             "string",   True,  "Snapshot",          "Free text",          "canonical_employees", "name"),
            _f("Users",        "lastName",      "User last name",                              "string",   True,  "Snapshot",          "Free text",          "canonical_employees", "name"),
            _f("Users",        "email",         "User email address",                          "string",   True,  "Snapshot",          "Email address",      "canonical_employees", "email"),
        ],
    },

    # ────────────────────── 11. RAMP ───────────────────────────────────────
    {
        "key": "ramp", "name": "Ramp", "category": "Spend Management",
        "header_color": "04AA6D", "has_connector": True,
        "fields": [
            _f("Transactions", "id",               "Unique Ramp transaction ID",                "string",   True,  "Transaction-level", "UUID string",        "canonical_expenses",  "source_id"),
            _f("Transactions", "merchant_name",    "Merchant or vendor name",                    "string",   True,  "Transaction-level", "Free text",          "canonical_expenses",  "vendor"),
            _f("Transactions", "amount",           "Transaction amount",                         "currency", True,  "Transaction-level", "Decimal USD",        "canonical_expenses",  "amount"),
            _f("Transactions", "transaction_date", "Transaction date",                           "date",     True,  "Transaction-level", "YYYY-MM-DD",         "canonical_expenses",  "period"),
            _f("Transactions", "user_id",          "Card holder user ID",                        "string",   True,  "Transaction-level", "UUID string",        "",                    ""),
            _f("Transactions", "status",           "Transaction status",                         "string",   True,  "Transaction-level", "Enum string",        "",                    ""),
            _f("Transactions", "category_name",    "Expense category name",                      "string",   False, "Transaction-level", "String",             "canonical_expenses",  "category"),
            _f("Transactions", "memo",             "Transaction memo",                           "string",   False, "Transaction-level", "Free text",          "canonical_expenses",  "description"),
            _f("Transactions", "department_name",  "Department name",                            "string",   False, "Transaction-level", "String",             "",                    ""),
            _f("Users",        "id",               "Unique Ramp user ID",                        "string",   True,  "Snapshot",          "UUID string",        "canonical_employees", "source_id"),
            _f("Users",        "name",             "User full name",                             "string",   True,  "Snapshot",          "Free text",          "canonical_employees", "name"),
            _f("Users",        "email",            "User email address",                         "string",   True,  "Snapshot",          "Email address",      "canonical_employees", "email"),
        ],
    },

    # ────────────────────── 12. WORKDAY ────────────────────────────────────
    {
        "key": "workday", "name": "Workday", "category": "HRIS",
        "header_color": "005CB9", "has_connector": False,
        "fields": [
            _f("Workers",       "Worker_ID",              "Unique Workday worker ID",                  "string",   True,  "Snapshot",          "WID string",         "canonical_employees", "source_id"),
            _f("Workers",       "Legal_Name_Full",        "Worker full legal name",                    "string",   True,  "Snapshot",          "Free text",          "canonical_employees", "name"),
            _f("Workers",       "Email_Address",          "Primary work email",                        "string",   True,  "Snapshot",          "Email address",      "canonical_employees", "email"),
            _f("Workers",       "Business_Title",         "Current business title",                    "string",   True,  "Snapshot",          "Free text",          "canonical_employees", "title"),
            _f("Workers",       "Supervisory_Organization","Department or team name",                  "string",   True,  "Snapshot",          "Org name string",    "canonical_employees", "department"),
            _f("Workers",       "Hire_Date",              "Original hire date",                        "date",     True,  "Snapshot",          "YYYY-MM-DD",         "canonical_employees", "hire_date"),
            _f("Workers",       "Worker_Status",          "Employment status (Active, Terminated, LOA)", "string", True,  "Snapshot",          "Enum string",        "canonical_employees", "status"),
            _f("Workers",       "Worker_Type",            "Worker type (Employee, Contingent)",         "string",   True,  "Snapshot",          "Enum string",        "",                    ""),
            _f("Workers",       "Location",               "Work location name",                        "string",   False, "Snapshot",          "Free text",          "",                    ""),
            _f("Workers",       "Manager_ID",             "Direct manager worker ID",                  "string",   False, "Snapshot",          "WID string",         "",                    ""),
            _f("Workers",       "Cost_Center",            "Cost center assignment",                    "string",   False, "Snapshot",          "String",             "",                    ""),
            _f("Compensation",  "Worker_ID",              "Worker ID reference",                       "string",   True,  "Monthly",           "WID string",         "canonical_employees", "source_id"),
            _f("Compensation",  "Base_Pay_Amount",        "Base salary amount",                        "currency", True,  "Monthly",           "Decimal USD",        "canonical_employees", "salary"),
            _f("Compensation",  "Pay_Frequency",          "Pay frequency (Monthly, Bi-weekly, Annual)","string",   True,  "Monthly",           "Enum string",        "",                    ""),
            _f("Compensation",  "Currency",               "Compensation currency",                     "string",   True,  "Monthly",           "ISO 4217",           "",                    ""),
            _f("Compensation",  "Total_Compensation",     "Total compensation including benefits",     "currency", False, "Monthly",           "Decimal USD",        "",                    ""),
            _f("Compensation",  "Effective_Date",         "Compensation effective date",               "date",     True,  "Monthly",           "YYYY-MM-DD",         "",                    ""),
            _f("Time Tracking", "Worker_ID",              "Worker ID reference",                       "string",   True,  "Transaction-level", "WID string",         "canonical_employees", "source_id"),
            _f("Time Tracking", "Date",                   "Time entry date",                           "date",     True,  "Transaction-level", "YYYY-MM-DD",         "",                    ""),
            _f("Time Tracking", "Hours_Worked",           "Total hours worked",                        "number",   True,  "Transaction-level", "Decimal hours",      "",                    ""),
            _f("Time Tracking", "Time_Type",              "Time type (Regular, Overtime, PTO)",         "string",   True,  "Transaction-level", "Enum string",        "",                    ""),
            _f("Time Tracking", "Billable",               "Whether hours are billable",                "boolean",  False, "Transaction-level", "Boolean",            "",                    ""),
            _f("Headcount",     "Period",                 "Reporting period",                          "date",     True,  "Monthly",           "YYYY-MM",            "",                    ""),
            _f("Headcount",     "Department",             "Department name",                           "string",   True,  "Monthly",           "Free text",          "canonical_employees", "department"),
            _f("Headcount",     "Active_Count",           "Active employee headcount",                 "number",   True,  "Monthly",           "Integer count",      "",                    ""),
            _f("Headcount",     "New_Hires",              "New hires in period",                       "number",   True,  "Monthly",           "Integer count",      "",                    ""),
            _f("Headcount",     "Terminations",           "Terminations in period",                    "number",   True,  "Monthly",           "Integer count",      "",                    ""),
        ],
    },

    # ────────────────────── 13. ADP ────────────────────────────────────────
    {
        "key": "adp", "name": "ADP", "category": "HRIS",
        "header_color": "D0271D", "has_connector": False,
        "fields": [
            _f("Employee Demographics", "associateOID",         "Unique ADP associate ID",                  "string",   True,  "Snapshot",          "OID string",         "canonical_employees", "source_id"),
            _f("Employee Demographics", "legalName.formattedName","Employee full legal name",                "string",   True,  "Snapshot",          "Free text",          "canonical_employees", "name"),
            _f("Employee Demographics", "businessCommunication.emailUri","Work email address",               "string",   True,  "Snapshot",          "Email address",      "canonical_employees", "email"),
            _f("Employee Demographics", "jobTitle",             "Current job title",                        "string",   True,  "Snapshot",          "Free text",          "canonical_employees", "title"),
            _f("Employee Demographics", "homeOrganizationalUnit.nameCode","Department code",                "string",   True,  "Snapshot",          "Code string",        "canonical_employees", "department"),
            _f("Employee Demographics", "originalHireDate",     "Original hire date",                       "date",     True,  "Snapshot",          "YYYY-MM-DD",         "canonical_employees", "hire_date"),
            _f("Employee Demographics", "workerStatus.statusCode","Employment status (Active, Terminated)", "string",   True,  "Snapshot",          "Enum string",        "canonical_employees", "status"),
            _f("Employee Demographics", "workerType",           "Worker type (Employee, Contractor)",       "string",   True,  "Snapshot",          "Enum string",        "",                    ""),
            _f("Employee Demographics", "workLocation",         "Work location name or code",               "string",   False, "Snapshot",          "String",             "",                    ""),
            _f("Employee Demographics", "reportsTo.associateOID","Direct manager ADP ID",                   "string",   False, "Snapshot",          "OID string",         "",                    ""),
            _f("Payroll Summary",       "associateOID",         "Employee ADP ID reference",                "string",   True,  "Monthly",           "OID string",         "canonical_employees", "source_id"),
            _f("Payroll Summary",       "payPeriodEndDate",     "Pay period end date",                      "date",     True,  "Monthly",           "YYYY-MM-DD",         "",                    ""),
            _f("Payroll Summary",       "grossPayAmount",       "Gross pay amount",                         "currency", True,  "Monthly",           "Decimal USD",        "canonical_employees", "salary"),
            _f("Payroll Summary",       "netPayAmount",         "Net pay amount after deductions",          "currency", True,  "Monthly",           "Decimal USD",        "",                    ""),
            _f("Payroll Summary",       "totalDeductions",      "Total deductions amount",                  "currency", False, "Monthly",           "Decimal USD",        "",                    ""),
            _f("Payroll Summary",       "overtimeHours",        "Overtime hours worked",                    "number",   False, "Monthly",           "Decimal hours",      "",                    ""),
            _f("Payroll Summary",       "regularHours",         "Regular hours worked",                     "number",   False, "Monthly",           "Decimal hours",      "",                    ""),
            _f("Benefits",             "associateOID",          "Employee ADP ID reference",                "string",   True,  "Snapshot",          "OID string",         "canonical_employees", "source_id"),
            _f("Benefits",             "planName",              "Benefit plan name",                        "string",   True,  "Snapshot",          "Free text",          "",                    ""),
            _f("Benefits",             "coverageLevel",         "Coverage level (Employee, Family, etc.)",  "string",   True,  "Snapshot",          "Enum string",        "",                    ""),
            _f("Benefits",             "employerContribution",  "Employer contribution amount",             "currency", False, "Snapshot",          "Decimal USD",        "",                    ""),
            _f("Benefits",             "effectiveDate",         "Benefit effective date",                   "date",     True,  "Snapshot",          "YYYY-MM-DD",         "",                    ""),
            _f("Time & Attendance",    "associateOID",          "Employee ADP ID reference",                "string",   True,  "Transaction-level", "OID string",         "canonical_employees", "source_id"),
            _f("Time & Attendance",    "date",                  "Time entry date",                          "date",     True,  "Transaction-level", "YYYY-MM-DD",         "",                    ""),
            _f("Time & Attendance",    "hoursWorked",           "Total hours worked",                       "number",   True,  "Transaction-level", "Decimal hours",      "",                    ""),
            _f("Time & Attendance",    "timeCode",              "Time code (Regular, OT, PTO, Sick)",       "string",   True,  "Transaction-level", "Enum string",        "",                    ""),
        ],
    },

    # ────────────────────── 14. SNOWFLAKE ──────────────────────────────────
    {
        "key": "snowflake", "name": "Snowflake", "category": "Data Warehouse",
        "header_color": "29B5E8", "has_connector": True,
        "fields": [
            _f("User-Defined Table", "PERIOD",         "Reporting period column",                     "date",     True,  "Monthly",           "YYYY-MM string",     "canonical_revenue",   "period"),
            _f("User-Defined Table", "<KPI_COLUMN>",   "Any column matching a known KPI key (e.g. revenue_growth, gross_margin). Multiple KPI columns can be present in a single table.", "number", True, "Monthly", "Decimal value", "", ""),
            _f("User-Defined Table", "REVENUE",        "Total revenue for the period",                "currency", False, "Monthly",           "Decimal USD",        "canonical_revenue",   "amount"),
            _f("User-Defined Table", "COGS",           "Cost of goods sold for the period",           "currency", False, "Monthly",           "Decimal USD",        "canonical_expenses",  "amount"),
            _f("User-Defined Table", "EXPENSES",       "Total operating expenses for the period",     "currency", False, "Monthly",           "Decimal USD",        "canonical_expenses",  "amount"),
            _f("User-Defined Table", "HEADCOUNT",      "Active employee count for the period",        "number",   False, "Monthly",           "Integer count",      "canonical_employees", "source_id"),
            _f("User-Defined Table", "CUSTOMERS",      "Active customer count for the period",        "number",   False, "Monthly",           "Integer count",      "canonical_customers", "source_id"),
            _f("User-Defined Table", "ARR",            "Annual recurring revenue for the period",     "currency", False, "Monthly",           "Decimal USD",        "canonical_revenue",   "amount"),
        ],
    },

    # ────────────────────── 15. GOOGLE SHEETS ──────────────────────────────
    {
        "key": "google_sheets", "name": "Google Sheets", "category": "Data Warehouse",
        "header_color": "34A853", "has_connector": True,
        "fields": [
            _f("Sheet (Row-per-Month)", "month",           "Reporting period column header",              "date",     True,  "Monthly",           "YYYY-MM string",     "canonical_revenue",   "period"),
            _f("Sheet (Row-per-Month)", "<KPI_COLUMN>",    "Any column header matching a known KPI key (e.g. revenue_growth, gross_margin). Multiple KPI columns supported.", "number", True, "Monthly", "Decimal value", "", ""),
            _f("Sheet (Row-per-Month)", "revenue",         "Total revenue for the period",                "currency", False, "Monthly",           "Decimal USD",        "canonical_revenue",   "amount"),
            _f("Sheet (Row-per-Month)", "cogs",            "Cost of goods sold for the period",           "currency", False, "Monthly",           "Decimal USD",        "canonical_expenses",  "amount"),
            _f("Sheet (Row-per-Month)", "expenses",        "Total operating expenses for the period",     "currency", False, "Monthly",           "Decimal USD",        "canonical_expenses",  "amount"),
            _f("Sheet (Row-per-Month)", "headcount",       "Active employee count",                       "number",   False, "Monthly",           "Integer count",      "canonical_employees", "source_id"),
            _f("Sheet (Row-per-Month)", "customers",       "Active customer count",                       "number",   False, "Monthly",           "Integer count",      "canonical_customers", "source_id"),
            _f("Sheet (Row-per-Month)", "arr",             "Annual recurring revenue",                    "currency", False, "Monthly",           "Decimal USD",        "canonical_revenue",   "amount"),
        ],
    },
]


# ─── ELT Pipeline Rules ───────────────────────────────────────────────────

ELT_PIPELINE_RULES = [
    # canonical_revenue consumption
    {"canonical_table": "canonical_revenue", "canonical_field": "amount",            "aggregation": "SUM by month",                    "output_kpis": "revenue_growth, gross_margin, operating_margin, ebitda_margin, contribution_margin, opex_ratio, burn_multiple, cash_runway, arr_growth"},
    {"canonical_table": "canonical_revenue", "canonical_field": "subscription_type", "aggregation": "Filter recurring vs one-time",    "output_kpis": "recurring_revenue, revenue_quality, arr_growth, mrr"},
    {"canonical_table": "canonical_revenue", "canonical_field": "customer_id",       "aggregation": "COUNT DISTINCT by month",         "output_kpis": "customer_concentration, churn_rate, nrr, expansion_rate, contraction_rate"},
    {"canonical_table": "canonical_revenue", "canonical_field": "period",            "aggregation": "GROUP BY year-month",             "output_kpis": "revenue_growth, revenue_momentum, pricing_power_index"},
    # canonical_expenses consumption
    {"canonical_table": "canonical_expenses", "canonical_field": "amount",           "aggregation": "SUM(ABS) by month",               "output_kpis": "gross_margin, operating_margin, ebitda_margin, opex_ratio, burn_multiple, cash_burn"},
    {"canonical_table": "canonical_expenses", "canonical_field": "category",         "aggregation": "SUM WHERE category matches S&M",  "output_kpis": "sales_efficiency, cac_payback, cpl, marketing_roi"},
    {"canonical_table": "canonical_expenses", "canonical_field": "category",         "aggregation": "SUM WHERE category matches COGS", "output_kpis": "gross_margin, contribution_margin"},
    # canonical_customers consumption
    {"canonical_table": "canonical_customers", "canonical_field": "source_id",       "aggregation": "COUNT by acquisition month",      "output_kpis": "churn_rate, logo_retention, customer_decay_slope, cac_payback, ltv_cac"},
    {"canonical_table": "canonical_customers", "canonical_field": "created_at",      "aggregation": "Parse to year-month cohort",      "output_kpis": "churn_rate, logo_retention, activation_rate"},
    {"canonical_table": "canonical_customers", "canonical_field": "lifecycle_stage",  "aggregation": "Filter by stage transitions",     "output_kpis": "mql_sql_rate, activation_rate"},
    # canonical_pipeline consumption
    {"canonical_table": "canonical_pipeline", "canonical_field": "amount",           "aggregation": "SUM by close month",              "output_kpis": "pipeline_conversion, avg_deal_size, pipeline_velocity"},
    {"canonical_table": "canonical_pipeline", "canonical_field": "stage",            "aggregation": "Filter won vs total",             "output_kpis": "win_rate, pipeline_conversion, quota_attainment"},
    {"canonical_table": "canonical_pipeline", "canonical_field": "close_date",       "aggregation": "GROUP BY year-month",             "output_kpis": "pipeline_conversion, pipeline_velocity"},
    # canonical_invoices consumption
    {"canonical_table": "canonical_invoices", "canonical_field": "amount",           "aggregation": "SUM by issue month",              "output_kpis": "dso, ar_turnover, avg_collection_period, cei"},
    {"canonical_table": "canonical_invoices", "canonical_field": "due_date",         "aggregation": "Compare to current date",         "output_kpis": "ar_aging_current, ar_aging_overdue, dso"},
    {"canonical_table": "canonical_invoices", "canonical_field": "status",           "aggregation": "Filter paid vs overdue",          "output_kpis": "ar_aging_current, ar_aging_overdue, cei"},
    # canonical_employees consumption
    {"canonical_table": "canonical_employees", "canonical_field": "source_id",       "aggregation": "COUNT active by month",           "output_kpis": "headcount_eff, rev_per_employee"},
    {"canonical_table": "canonical_employees", "canonical_field": "salary",          "aggregation": "SUM by month",                    "output_kpis": "headcount_eff"},
    {"canonical_table": "canonical_employees", "canonical_field": "hire_date",       "aggregation": "Parse to year-month",             "output_kpis": "ramp_time"},
    {"canonical_table": "canonical_employees", "canonical_field": "status",          "aggregation": "Filter active/employed",          "output_kpis": "headcount_eff, rev_per_employee, billable_utilization"},
]


# ─── Excel Workbook Generator ─────────────────────────────────────────────

def generate_integration_spec_workbook():
    """Build a multi-tab openpyxl Workbook with the full integration spec."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Import KPI definitions
    from core.kpi_defs import KPI_DEFS, EXTENDED_ONTOLOGY_METRICS

    wb = Workbook()

    # ── Styling helpers ────────────────────────────────────────────────
    BRAND = "0055A4"
    alt_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    thin_border = Border(bottom=Side(style="thin", color="E2E8F0"))
    wrap = Alignment(wrap_text=True, vertical="top")

    def header_fill(color):
        return PatternFill(start_color=color, end_color=color, fill_type="solid")

    header_font = Font(bold=True, color="FFFFFF", size=10)

    def auto_width(ws, max_w=45):
        for col in ws.columns:
            mx = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    mx = max(mx, min(len(str(cell.value)), max_w))
            ws.column_dimensions[col_letter].width = max(mx + 3, 12)

    def style_header(ws, headers, color=BRAND):
        ws.append(headers)
        fill = header_fill(color)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = fill
            cell.alignment = wrap
        ws.freeze_panes = "A2"

    def style_rows(ws, start=2):
        for idx, row in enumerate(ws.iter_rows(min_row=start, max_row=ws.max_row), start=0):
            for cell in row:
                cell.border = thin_border
                cell.alignment = wrap
                if idx % 2 == 1:
                    cell.fill = alt_fill

    # ── 1. Overview Tab ────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Overview"
    style_header(ws, ["Axiom Intelligence - Integration Specification"], BRAND)
    ws.merge_cells("A1:D1")
    ws[1][0].font = Font(bold=True, color="FFFFFF", size=14)
    ws.append([])
    ws.append(["Generated", datetime.utcnow().strftime("%Y-%m-%d")])
    ws.append(["Total Source Systems", len(SYSTEM_REGISTRY)])
    total_fields = sum(len(s["fields"]) for s in SYSTEM_REGISTRY)
    ws.append(["Total Field Definitions", total_fields])
    all_kpis = {d["key"] for d in KPI_DEFS} | {d["key"] for d in EXTENDED_ONTOLOGY_METRICS}
    ws.append(["Total KPIs Tracked", len(all_kpis)])
    ws.append(["Canonical Tables", len(CANONICAL_SCHEMAS)])
    ws.append([])

    # System summary table
    ws.append(["System", "Category", "Entities", "Fields", "Connector Status"])
    row_num = ws.max_row
    for cell in ws[row_num]:
        cell.font = Font(bold=True, size=10)
        cell.fill = header_fill(BRAND)
        cell.font = header_font
    for sys in SYSTEM_REGISTRY:
        entities = sorted(set(f["entity"] for f in sys["fields"]))
        ws.append([
            sys["name"], sys["category"],
            ", ".join(entities), len(sys["fields"]),
            "Active" if sys["has_connector"] else "Planned",
        ])

    ws.append([])
    ws.append(["Canonical Table", "Description", "Field Count"])
    row_num = ws.max_row
    for cell in ws[row_num]:
        cell.font = header_font
        cell.fill = header_fill(BRAND)
    for tname, tdef in CANONICAL_SCHEMAS.items():
        ws.append([tname, tdef["description"], len(tdef["fields"])])

    auto_width(ws)
    style_rows(ws, start=9)

    # ── 2. Canonical Schema Tab ────────────────────────────────────────
    ws2 = wb.create_sheet("Canonical Schema")
    style_header(ws2, ["Table", "Field Name", "Data Type", "Description", "Source Systems"])
    for tname, tdef in CANONICAL_SCHEMAS.items():
        for fname, ftype, fdesc in tdef["fields"]:
            # Find which systems map to this field
            sources = []
            for sys in SYSTEM_REGISTRY:
                for f in sys["fields"]:
                    if f["canonical_table"] == tname and f["canonical_field"] == fname:
                        if sys["name"] not in sources:
                            sources.append(sys["name"])
            ws2.append([tname, fname, ftype, fdesc, ", ".join(sources) if sources else ""])
    auto_width(ws2)
    style_rows(ws2)

    # ── 3. KPI Reference Tab ──────────────────────────────────────────
    ws3 = wb.create_sheet("KPI Reference")
    style_header(ws3, ["KPI Key", "KPI Name", "Domain", "Unit", "Direction", "Formula", "Required Canonical Fields"])

    # Domain map for core KPIs
    from core.kpi_defs import CAUSATION_RULES
    # Build domain from extended + explicit
    domain_map = {}
    for d in EXTENDED_ONTOLOGY_METRICS:
        domain_map[d["key"]] = d.get("domain", "")
    for d in KPI_DEFS:
        domain_map[d["key"]] = d.get("domain", "")

    # Infer domain for core KPIs without explicit domain
    for d in KPI_DEFS:
        key = d["key"]
        if not domain_map.get(key):
            if key in ("revenue_growth", "arr_growth", "recurring_revenue", "revenue_quality",
                       "customer_concentration", "customer_ltv", "pricing_power_index"):
                domain_map[key] = "revenue"
            elif key in ("gross_margin", "operating_margin", "ebitda_margin", "contribution_margin",
                         "opex_ratio", "operating_leverage", "margin_volatility", "burn_multiple"):
                domain_map[key] = "profitability"
            elif key in ("churn_rate", "nrr", "logo_retention", "customer_decay_slope"):
                domain_map[key] = "retention"
            elif key in ("cac_payback", "sales_efficiency"):
                domain_map[key] = "efficiency"

    all_kpi_list = []
    seen = set()
    for d in KPI_DEFS:
        if d["key"] not in seen:
            seen.add(d["key"])
            all_kpi_list.append(d)
    for d in EXTENDED_ONTOLOGY_METRICS:
        if d["key"] not in seen:
            seen.add(d["key"])
            all_kpi_list.append(d)

    for d in all_kpi_list:
        key = d["key"]
        deps = KPI_FIELD_DEPS.get(key, [])
        ws3.append([
            key, d["name"], domain_map.get(key, d.get("domain", "")),
            d.get("unit", ""), d.get("direction", ""),
            d.get("formula", ""), ", ".join(deps),
        ])
    auto_width(ws3)
    style_rows(ws3)

    # ── 4-18. Per-System Tabs ──────────────────────────────────────────
    FIELD_HEADERS = [
        "Source Object", "Source Field", "Description", "Data Type",
        "Required", "Frequency", "Unit of Ingest",
        "Canonical Table", "Canonical Field", "KPIs Driven",
    ]

    for sys in SYSTEM_REGISTRY:
        ws_sys = wb.create_sheet(sys["name"])
        style_header(ws_sys, FIELD_HEADERS, sys["header_color"])

        if not sys["has_connector"]:
            ws_sys.append(["NOTE: Connector planned - fields represent expected integration specification. Actual field names may vary upon implementation."] + [""] * 9)
            ws_sys.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)
            note_cell = ws_sys.cell(row=2, column=1)
            note_cell.font = Font(italic=True, color="B45309", size=10)
            note_cell.fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

        for f in sys["fields"]:
            ws_sys.append([
                f["entity"], f["source_field"], f["description"], f["data_type"],
                "Required" if f["required"] else "Optional",
                f["frequency"], f["unit_of_ingest"],
                f["canonical_table"], f["canonical_field"], f["kpis_driven"],
            ])

        auto_width(ws_sys)
        start = 3 if not sys["has_connector"] else 2
        style_rows(ws_sys, start=start)

    # ── 19. ELT Pipeline Tab ──────────────────────────────────────────
    ws_elt = wb.create_sheet("ELT Pipeline")
    style_header(ws_elt, ["Canonical Table", "Canonical Field", "Aggregation Method", "Output KPIs"])

    for rule in ELT_PIPELINE_RULES:
        ws_elt.append([
            rule["canonical_table"], rule["canonical_field"],
            rule["aggregation"], rule["output_kpis"],
        ])
    auto_width(ws_elt)
    style_rows(ws_elt)

    return wb
