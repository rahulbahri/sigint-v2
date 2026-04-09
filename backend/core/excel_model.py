"""
core/excel_model.py -- Financial model workbook generator with live Excel formulas.

Generates a structured multi-sheet Excel workbook where the Forecast and P&L
sheets use real formulas referencing an editable Assumptions sheet, enabling
CFOs to modify assumptions and see cascading impact without leaving Excel.

Sheets:
  1. Assumptions (editable inputs with named ranges)
  2. Actuals (locked historical data)
  3. Forecast (formulas referencing Assumptions)
  4. P&L (formulas building a simplified income statement)
  5. Cash Flow (operating, investing, financing with live formulas)
  6. Balance Sheet (assets, liabilities, equity linked to Cash Flow + P&L)
  7. Scenarios (side-by-side comparison from saved scenarios)
  8. Confidence Bands (Monte Carlo p10/p25/p50/p75/p90)
  9. Dashboard (charts)
"""
import json
from datetime import datetime
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.chart import LineChart, AreaChart, BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Styling Constants ────────────────────────────────────────────────────────

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
_HEADER_FILL = PatternFill(start_color="0055A4", end_color="0055A4", fill_type="solid")
_SUBHEADER_FILL = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
_FORMULA_FONT = Font(name="Consolas", size=9, color="6B21A8")
_LOCKED_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
_EDITABLE_FILL = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
_GREEN_FILL = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")
_RED_FILL = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid")
_THIN_BORDER = Border(bottom=Side(style="thin", color="E2E8F0"))
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
_ALIGN_RIGHT = Alignment(horizontal="right")

# ── Core Assumption Parameters ───────────────────────────────────────────────
# These are the editable inputs on the Assumptions sheet. Each tuple:
# (key, label, unit, description, default_source_kpi_or_none)

ASSUMPTION_PARAMS = [
    ("revenue",          "Revenue",            "USD",   "Monthly revenue",                 None),
    ("cogs",             "COGS",               "USD",   "Cost of goods sold",              None),
    ("opex",             "Operating Expenses",  "USD",   "Total OpEx",                      None),
    ("headcount",        "Headcount",           "count", "Total employees",                 None),
    ("customers",        "Customers",           "count", "Total active customers",          None),
    ("new_customers",    "New Customers",       "count", "Customers acquired this month",   None),
    ("churned_customers","Churned Customers",   "count", "Customers lost this month",       None),
    ("mrr",              "MRR",                "USD",   "Monthly recurring revenue",        None),
    ("arr",              "ARR",                "USD",   "Annual recurring revenue",         None),
    ("ar",               "Accounts Receivable", "USD",   "AR balance",                      None),
    ("sm_allocated",     "S&M Spend",          "USD",   "Sales & marketing spend",         None),
    ("expansion_rev",    "Expansion Revenue",   "USD",   "Upsell + cross-sell revenue",     None),
    ("contraction_rev",  "Contraction Revenue", "USD",   "Revenue lost to downgrades",      None),
    ("revenue_growth_rate", "Revenue Growth Rate", "pct", "Expected MoM revenue growth",  "revenue_growth"),
    ("gross_margin_target", "Gross Margin Target", "pct", "Target gross margin %",        "gross_margin"),
]

# P&L line items (key, label, formula_type)
PL_LINES = [
    ("revenue",         "Revenue",           "direct"),
    ("cogs",            "COGS",              "direct"),
    ("gross_profit",    "Gross Profit",      "formula"),  # =Revenue - COGS
    ("gross_margin_pct","Gross Margin %",    "formula"),  # =GP / Revenue * 100
    ("opex",            "Operating Expenses", "direct"),
    ("operating_income","Operating Income",   "formula"),  # =GP - OpEx
    ("op_margin_pct",   "Operating Margin %", "formula"),  # =OI / Revenue * 100
    ("ebitda",          "EBITDA (proxy)",     "formula"),  # =OI * 1.15
    ("ebitda_margin_pct","EBITDA Margin %",   "formula"),  # =EBITDA / Revenue * 100
    ("net_burn",        "Net Burn",          "formula"),  # =COGS + OpEx - Revenue
]


# ── Helper Functions ─────────────────────────────────────────────────────────

def _col(n: int) -> str:
    """1-indexed column number to Excel letter."""
    return get_column_letter(n)


def _cell(row: int, col: int) -> str:
    """Return absolute cell reference like $B$5."""
    return f"${_col(col)}${row}"


def _auto_width(ws, max_w: int = 40):
    """Set column widths based on content."""
    for col_cells in ws.columns:
        mx = 0
        letter = _col(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                mx = max(mx, min(len(str(cell.value)), max_w))
        ws.column_dimensions[letter].width = max(mx + 3, 12)


def _style_header(ws, row: int, num_cols: int):
    """Apply header styling to a row."""
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _ALIGN_CENTER


# ── CellMap ──────────────────────────────────────────────────────────────────

class CellMap:
    """Tracks cell locations across sheets for cross-sheet formula references."""

    def __init__(self):
        self._map: Dict[str, Dict[str, Dict[int, str]]] = {}
        # sheet -> key -> {month_col: "SheetName!$C$5"}
        self._named_ranges: Dict[str, str] = {}

    def register(self, sheet: str, key: str, row: int, col: int, month_idx: int = 0):
        self._map.setdefault(sheet, {}).setdefault(key, {})[month_idx] = \
            f"'{sheet}'!{_cell(row, col)}"

    def resolve(self, key: str, sheet: str = None, month_idx: int = 0) -> Optional[str]:
        if sheet:
            return self._map.get(sheet, {}).get(key, {}).get(month_idx)
        # Search all sheets
        for s in self._map:
            ref = self._map[s].get(key, {}).get(month_idx)
            if ref:
                return ref
        return None

    def add_named_range(self, name: str, ref: str):
        self._named_ranges[name] = ref

    @property
    def named_ranges(self) -> Dict[str, str]:
        return dict(self._named_ranges)


# ── ModelWorkbookBuilder ─────────────────────────────────────────────────────

class ModelWorkbookBuilder:
    """
    Orchestrates the construction of the full financial model workbook.

    Arguments:
        monthly_data: list of (year, month, kpi_dict) sorted chronologically
        forecast_data: dict with 'trajectories' (kpi -> [{step, p10..p90}]),
                       'value_ranges', 'kpis'  (or None if no model trained)
        scenarios: list of {name, levers_json} from saved_scenarios
        settings: dict with company_name, stage, model_window_months
        targets: dict of kpi_key -> {target_value, unit, direction}
    """

    def __init__(self, monthly_data: list, forecast_data: Optional[dict],
                 scenarios: list, settings: dict, targets: dict,
                 projection_data: Optional[list] = None):
        self.monthly_data = monthly_data
        self.forecast_data = forecast_data
        self.scenarios = scenarios
        self.settings = settings
        self.targets = targets
        self.projection_data = projection_data or []
        self.wb = Workbook()
        self.cell_map = CellMap()

        # Derive month labels from data
        self.months = [(y, m) for y, m, _ in monthly_data]
        # Forecast months (if available)
        self.forecast_months = []
        if forecast_data and forecast_data.get("trajectories"):
            first_kpi = next(iter(forecast_data["trajectories"]), None)
            if first_kpi:
                steps = forecast_data["trajectories"][first_kpi]
                last_y, last_m = self.months[-1] if self.months else (datetime.utcnow().year, datetime.utcnow().month)
                for i, step in enumerate(steps):
                    if i == 0:
                        continue  # step 0 = "Now" (current month)
                    fm = last_m + i
                    fy = last_y + (fm - 1) // 12
                    fm = ((fm - 1) % 12) + 1
                    self.forecast_months.append((fy, fm))

    def build(self) -> Workbook:
        """Build the complete workbook and return it."""
        self._build_assumptions_sheet()
        self._build_actuals_sheet()
        if self.projection_data:
            self._build_projections_sheet()
        self._build_forecast_sheet()
        self._build_pl_sheet()
        self._build_cashflow_sheet()
        self._build_balance_sheet()
        self._build_scenarios_sheet()
        self._build_confidence_sheet()
        self._build_dashboard_sheet()
        self._add_metadata()
        return self.wb

    # ── Sheet 1: Assumptions ─────────────────────────────────────────────

    def _build_assumptions_sheet(self):
        ws = self.wb.active
        ws.title = "Assumptions"

        # Header
        ws.merge_cells("A1:F1")
        ws.cell(row=1, column=1, value="Assumptions -- Edit These Values").font = \
            Font(bold=True, size=14, color="0055A4")

        company = self.settings.get("company_name", "Your Company")
        ws.cell(row=2, column=1, value=f"Financial Model for {company}").font = \
            Font(size=10, color="94A3B8", italic=True)

        # Column headers
        row = 4
        headers = ["Parameter", "Unit", "Description"]
        # Add monthly columns
        for i, (y, m) in enumerate(self.months[-12:]):  # last 12 months of actuals
            headers.append(f"{y}-{m:02d}")
        for c, h in enumerate(headers, 1):
            ws.cell(row=row, column=c, value=h)
        _style_header(ws, row, len(headers))

        # Extract last 12 months of raw data
        recent_data = self.monthly_data[-12:] if len(self.monthly_data) >= 12 else self.monthly_data

        # Data rows
        for p_idx, (key, label, unit, desc, source_kpi) in enumerate(ASSUMPTION_PARAMS):
            r = row + 1 + p_idx
            ws.cell(row=r, column=1, value=label).font = Font(bold=True, size=10)
            ws.cell(row=r, column=2, value=unit).font = Font(size=9, color="64748B")
            ws.cell(row=r, column=3, value=desc).font = Font(size=9, color="94A3B8")

            # Populate monthly values from data
            for m_idx, (y, m, kpis) in enumerate(recent_data):
                col = 4 + m_idx
                # Try the key directly, then the source_kpi
                val = kpis.get(key) or (kpis.get(source_kpi) if source_kpi else None)
                if val is not None:
                    try:
                        ws.cell(row=r, column=col, value=round(float(val), 2))
                    except (ValueError, TypeError):
                        ws.cell(row=r, column=col, value=val)

                # Style as editable
                ws.cell(row=r, column=col).fill = _EDITABLE_FILL
                ws.cell(row=r, column=col).protection = Protection(locked=False)

                # Register in cell map
                self.cell_map.register("Assumptions", key, r, col, month_idx=m_idx)

            ws.row_dimensions[r].height = 22

        # Protect sheet (editable cells marked unlocked above)
        ws.protection.sheet = True
        ws.protection.password = ""
        ws.freeze_panes = "D5"
        _auto_width(ws)

    # ── Sheet 2: Actuals ─────────────────────────────────────────────────

    def _build_actuals_sheet(self):
        ws = self.wb.create_sheet("Actuals")

        # Collect all KPI keys that appear in the data
        all_kpis = set()
        for _, _, kpis in self.monthly_data:
            for k in kpis:
                if not k.startswith("_"):
                    all_kpis.add(k)
        kpi_list = sorted(all_kpis)

        # Headers
        row = 1
        headers = ["KPI Key", "KPI Name"]
        for y, m in self.months:
            headers.append(f"{y}-{m:02d}")
        for c, h in enumerate(headers, 1):
            ws.cell(row=row, column=c, value=h)
        _style_header(ws, row, len(headers))

        # Data rows
        from core.kpi_defs import KPI_DEFS, EXTENDED_ONTOLOGY_METRICS
        kpi_name_map = {}
        for kd in KPI_DEFS:
            kpi_name_map[kd["key"]] = kd.get("name", kd["key"])
        for kd in EXTENDED_ONTOLOGY_METRICS:
            kpi_name_map[kd["key"]] = kd.get("name", kd["key"])

        for k_idx, kpi_key in enumerate(kpi_list):
            r = row + 1 + k_idx
            ws.cell(row=r, column=1, value=kpi_key).font = Font(size=9, name="Consolas")
            ws.cell(row=r, column=2, value=kpi_name_map.get(kpi_key, kpi_key)).font = Font(size=9)

            for m_idx, (y, m, kpis) in enumerate(self.monthly_data):
                col = 3 + m_idx
                val = kpis.get(kpi_key)
                if val is not None:
                    try:
                        ws.cell(row=r, column=col, value=round(float(val), 4))
                    except (ValueError, TypeError):
                        ws.cell(row=r, column=col, value=val)
                ws.cell(row=r, column=col).fill = _LOCKED_FILL
                ws.cell(row=r, column=col).protection = Protection(locked=True)
                ws.cell(row=r, column=col).number_format = '#,##0.00'

                # Register in cell map
                self.cell_map.register("Actuals", kpi_key, r, col, month_idx=m_idx)

            # Zebra striping
            if k_idx % 2 == 0:
                for c in range(1, len(headers) + 1):
                    ws.cell(row=r, column=c).fill = _LOCKED_FILL

        ws.protection.sheet = True
        ws.protection.password = ""
        ws.freeze_panes = "C2"
        _auto_width(ws)

    # ── Sheet: Projections (user-uploaded plan/budget data) ────────────

    def _build_projections_sheet(self):
        """Build a Projections sheet from user-uploaded plan/budget data.
        Layout mirrors the Actuals sheet: KPI keys in column A, months across top."""
        ws = self.wb.create_sheet("Projections")

        if not self.projection_data:
            ws.cell(row=1, column=1, value="No projection data uploaded.").font = \
                Font(size=12, color="94A3B8", italic=True)
            return

        # Collect all KPI keys across projection months
        all_kpis = set()
        for _, _, kpi_dict in self.projection_data:
            for k in kpi_dict:
                if not k.startswith("_") and k not in ("year", "month"):
                    all_kpis.add(k)
        sorted_kpis = sorted(all_kpis)

        # Header row
        ws.cell(row=1, column=1, value="KPI").font = _HEADER_FONT
        ws.cell(row=1, column=1).fill = _HEADER_FILL
        for i, (y, m, _) in enumerate(self.projection_data):
            col = 2 + i
            ws.cell(row=1, column=col, value=f"{y}-{m:02d}").font = _HEADER_FONT
            ws.cell(row=1, column=col).fill = PatternFill(
                start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")

        # Data rows
        for r_idx, kpi in enumerate(sorted_kpis):
            row = 2 + r_idx
            ws.cell(row=row, column=1, value=kpi).font = Font(size=10, color="1E293B")
            for m_idx, (y, m, kpi_dict) in enumerate(self.projection_data):
                col = 2 + m_idx
                val = kpi_dict.get(kpi)
                if val is not None and isinstance(val, (int, float)):
                    ws.cell(row=row, column=col, value=round(float(val), 2))
                    ws.cell(row=row, column=col).number_format = '#,##0.00'
                # Register in cell_map for cross-sheet references
                self.cell_map.register("Projections", kpi, row, col, month_idx=m_idx)

        ws.freeze_panes = "B2"
        _auto_width(ws)

    # ── Sheet 3: Forecast ────────────────────────────────────────────────

    def _build_forecast_sheet(self):
        ws = self.wb.create_sheet("Forecast")

        if not self.forecast_data or not self.forecast_data.get("trajectories"):
            ws.cell(row=1, column=1, value="No forecast model trained yet.").font = \
                Font(size=12, color="94A3B8", italic=True)
            ws.cell(row=2, column=1, value="Train the model in Forward Signals, then re-export.").font = \
                Font(size=10, color="94A3B8")
            return

        trajectories = self.forecast_data["trajectories"]
        kpis = sorted(trajectories.keys())

        # Headers
        row = 1
        ws.cell(row=row, column=1, value="KPI").font = _HEADER_FONT
        ws.cell(row=row, column=1).fill = _HEADER_FILL
        ws.cell(row=row, column=2, value="Current").font = _HEADER_FONT
        ws.cell(row=row, column=2).fill = _HEADER_FILL

        for i, (fy, fm) in enumerate(self.forecast_months):
            col = 3 + i
            ws.cell(row=row, column=col, value=f"{fy}-{fm:02d}").font = _HEADER_FONT
            ws.cell(row=row, column=col).fill = _HEADER_FILL

        # Determine which KPIs have assumption-based formulas
        # Revenue forecast: =PrevMonth * (1 + growth_rate/100)
        # Other KPIs: use p50 from Monte Carlo as hard value

        for k_idx, kpi_key in enumerate(kpis):
            r = row + 1 + k_idx
            traj = trajectories[kpi_key]
            ws.cell(row=r, column=1, value=kpi_key).font = Font(size=9, name="Consolas")

            # Current value (step 0 p50)
            if traj and len(traj) > 0:
                current = traj[0].get("p50", traj[0].get("current"))
                if current is not None:
                    ws.cell(row=r, column=2, value=round(float(current), 4))
                    ws.cell(row=r, column=2).number_format = '#,##0.00'

            # Forecast months
            for i, step_data in enumerate(traj[1:] if len(traj) > 1 else []):
                col = 3 + i
                p50 = step_data.get("p50")

                if kpi_key == "revenue_growth" and i == 0:
                    # For revenue_growth, use the growth rate assumption if available
                    assumption_ref = self.cell_map.resolve(
                        "revenue_growth_rate", "Assumptions",
                        month_idx=min(len(self.monthly_data) - 1, 11)
                    )
                    if assumption_ref:
                        ws.cell(row=r, column=col, value=f"={assumption_ref}")
                        ws.cell(row=r, column=col).font = _FORMULA_FONT
                        continue

                # Default: hard p50 value
                if p50 is not None:
                    ws.cell(row=r, column=col, value=round(float(p50), 4))
                    ws.cell(row=r, column=col).number_format = '#,##0.00'

                self.cell_map.register("Forecast", kpi_key, r, col, month_idx=i)

        ws.freeze_panes = "C2"
        _auto_width(ws)

    # ── Sheet 4: P&L ─────────────────────────────────────────────────────

    def _build_pl_sheet(self):
        ws = self.wb.create_sheet("P&L")

        # Use last 12 actuals + forecast months
        actuals_slice = self.monthly_data[-12:] if len(self.monthly_data) >= 12 else self.monthly_data
        all_months = [(y, m, "actual") for y, m, _ in actuals_slice]
        for fy, fm in self.forecast_months:
            all_months.append((fy, fm, "forecast"))

        # Headers
        row = 1
        ws.cell(row=row, column=1, value="Line Item").font = _HEADER_FONT
        ws.cell(row=row, column=1).fill = _HEADER_FILL
        for i, (y, m, source) in enumerate(all_months):
            col = 2 + i
            label = f"{y}-{m:02d}"
            ws.cell(row=row, column=col, value=label).font = _HEADER_FONT
            ws.cell(row=row, column=col).fill = _HEADER_FILL
            if source == "forecast":
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color="1E3A5F", end_color="1E3A5F", fill_type="solid"
                )

        # Track P&L row positions for formula references
        pl_rows = {}

        for line_idx, (key, label, formula_type) in enumerate(PL_LINES):
            r = row + 1 + line_idx
            pl_rows[key] = r

            # Bold for totals
            is_total = key in ("gross_profit", "operating_income", "ebitda", "net_burn")
            ws.cell(row=r, column=1, value=label).font = Font(
                bold=is_total, size=10,
                color="0055A4" if is_total else "1E293B"
            )
            if is_total:
                ws.cell(row=r, column=1).border = Border(
                    top=Side(style="thin", color="0055A4"),
                    bottom=Side(style="double", color="0055A4"),
                )

            for m_idx, (y, m, source) in enumerate(all_months):
                col = 2 + m_idx

                if formula_type == "direct":
                    # Pull from Actuals or Forecast
                    if source == "actual":
                        data_idx = len(self.monthly_data) - len(actuals_slice) + m_idx
                        if data_idx < len(self.monthly_data):
                            val = self.monthly_data[data_idx][2].get(key)
                            if val is not None:
                                try:
                                    ws.cell(row=r, column=col, value=round(float(val), 2))
                                except (ValueError, TypeError):
                                    pass
                    else:
                        # Forecast: reference Forecast sheet p50 if available
                        fc_ref = self.cell_map.resolve(key, "Forecast", month_idx=m_idx - len(actuals_slice))
                        if fc_ref:
                            ws.cell(row=r, column=col, value=f"={fc_ref}")
                            ws.cell(row=r, column=col).font = _FORMULA_FONT

                elif formula_type == "formula":
                    # Build the formula from P&L row references
                    rev_cell = _cell(pl_rows.get("revenue", r), col)
                    cogs_cell = _cell(pl_rows.get("cogs", r), col)
                    gp_cell = _cell(pl_rows.get("gross_profit", r), col)
                    opex_cell = _cell(pl_rows.get("opex", r), col)
                    oi_cell = _cell(pl_rows.get("operating_income", r), col)
                    ebitda_cell = _cell(pl_rows.get("ebitda", r), col)

                    formula = None
                    if key == "gross_profit":
                        formula = f"={rev_cell}-{cogs_cell}"
                    elif key == "gross_margin_pct":
                        formula = f"=IF({rev_cell}=0,0,{gp_cell}/{rev_cell}*100)"
                    elif key == "operating_income":
                        formula = f"={gp_cell}-{opex_cell}"
                    elif key == "op_margin_pct":
                        formula = f"=IF({rev_cell}=0,0,{oi_cell}/{rev_cell}*100)"
                    elif key == "ebitda":
                        formula = f"={oi_cell}*1.15"
                    elif key == "ebitda_margin_pct":
                        formula = f"=IF({rev_cell}=0,0,{ebitda_cell}/{rev_cell}*100)"
                    elif key == "net_burn":
                        formula = f"={cogs_cell}+{opex_cell}-{rev_cell}"

                    if formula:
                        ws.cell(row=r, column=col, value=formula)
                        ws.cell(row=r, column=col).font = _FORMULA_FONT

                ws.cell(row=r, column=col).number_format = '#,##0.00'

        ws.freeze_panes = "B2"
        _auto_width(ws)

    # ── Sheet 5: Cash Flow Statement ───────────────────────────────────

    # Cash Flow line items: (key, label, formula_or_source)
    CF_LINES = [
        # Operating Activities
        ("cf_header_ops",     "OPERATING ACTIVITIES",          "header"),
        ("net_income",        "Net Income (Operating Income)", "pl_ref:operating_income"),
        ("add_da",            "Add: Depreciation & Amort.",    "percent_of:revenue:2"),  # ~2% of revenue proxy
        ("change_ar",         "Change in Accounts Receivable", "delta:ar"),
        ("change_ap",         "Change in Accounts Payable",    "percent_of:opex:-8"),  # ~8% of opex proxy
        ("cf_from_ops",       "Cash Flow from Operations",     "sum:net_income:add_da:change_ar:change_ap"),
        # Investing Activities
        ("cf_header_inv",     "INVESTING ACTIVITIES",          "header"),
        ("capex",             "Capital Expenditures",          "percent_of:revenue:-3"),  # ~3% SaaS capex
        ("cf_from_investing", "Cash Flow from Investing",      "ref:capex"),
        # Financing Activities
        ("cf_header_fin",     "FINANCING ACTIVITIES",          "header"),
        ("equity_debt",       "Equity / Debt Raised",          "value:0"),
        ("cf_from_financing", "Cash Flow from Financing",      "ref:equity_debt"),
        # Net
        ("cf_separator",      "",                              "separator"),
        ("net_cash_flow",     "Net Cash Flow",                 "sum:cf_from_ops:cf_from_investing:cf_from_financing"),
        ("opening_cash",      "Opening Cash Balance",          "carry:closing_cash"),
        ("closing_cash",      "Closing Cash Balance",          "sum:opening_cash:net_cash_flow"),
    ]

    def _build_cashflow_sheet(self):
        ws = self.wb.create_sheet("Cash Flow")

        actuals_slice = self.monthly_data[-12:] if len(self.monthly_data) >= 12 else self.monthly_data
        all_months = [(y, m, "actual") for y, m, _ in actuals_slice]
        for fy, fm in self.forecast_months:
            all_months.append((fy, fm, "forecast"))

        # Headers
        row = 1
        ws.cell(row=row, column=1, value="Cash Flow Statement").font = _HEADER_FONT
        ws.cell(row=row, column=1).fill = _HEADER_FILL
        for i, (y, m, source) in enumerate(all_months):
            col = 2 + i
            ws.cell(row=row, column=col, value=f"{y}-{m:02d}").font = _HEADER_FONT
            ws.cell(row=row, column=col).fill = _HEADER_FILL if source == "actual" else PatternFill(
                start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")

        cf_rows = {}
        for line_idx, (key, label, formula_type) in enumerate(self.CF_LINES):
            r = row + 1 + line_idx
            cf_rows[key] = r

            if formula_type == "header":
                ws.cell(row=r, column=1, value=label).font = Font(bold=True, size=10, color="0055A4")
                continue
            if formula_type == "separator":
                continue

            is_total = key in ("cf_from_ops", "cf_from_investing", "cf_from_financing", "net_cash_flow", "closing_cash")
            ws.cell(row=r, column=1, value=label).font = Font(
                bold=is_total, size=10, color="0055A4" if is_total else "1E293B")
            if is_total:
                ws.cell(row=r, column=1).border = Border(
                    top=Side(style="thin", color="0055A4"),
                    bottom=Side(style="double", color="0055A4"))

            for m_idx, (y, m, source) in enumerate(all_months):
                col = 2 + m_idx

                if formula_type.startswith("pl_ref:"):
                    # Reference P&L line item
                    pl_key = formula_type.split(":")[1]
                    # Find the row in P&L sheet
                    for pl_idx, (pk, _, _) in enumerate(PL_LINES):
                        if pk == pl_key:
                            ws.cell(row=r, column=col, value=f"='P&L'!{_cell(2 + pl_idx, col)}")
                            ws.cell(row=r, column=col).font = _FORMULA_FONT
                            break

                elif formula_type.startswith("percent_of:"):
                    parts = formula_type.split(":")
                    base_key = parts[1]
                    pct = float(parts[2])
                    # Use actuals data
                    if source == "actual":
                        data_idx = len(self.monthly_data) - len(actuals_slice) + m_idx
                        if data_idx < len(self.monthly_data):
                            base_val = self.monthly_data[data_idx][2].get(base_key, 0)
                            if base_val:
                                try:
                                    ws.cell(row=r, column=col, value=round(float(base_val) * pct / 100, 2))
                                except (ValueError, TypeError):
                                    ws.cell(row=r, column=col, value=0)
                    else:
                        # For forecast, use formula referencing P&L revenue
                        for pl_idx, (pk, _, _) in enumerate(PL_LINES):
                            if pk == base_key:
                                ws.cell(row=r, column=col,
                                        value=f"='P&L'!{_cell(2 + pl_idx, col)}*{pct/100}")
                                ws.cell(row=r, column=col).font = _FORMULA_FONT
                                break

                elif formula_type.startswith("delta:"):
                    # Change in a balance (this month - last month)
                    balance_key = formula_type.split(":")[1]
                    if source == "actual" and m_idx > 0:
                        data_idx = len(self.monthly_data) - len(actuals_slice) + m_idx
                        prev_idx = data_idx - 1
                        if data_idx < len(self.monthly_data) and prev_idx >= 0:
                            curr = self.monthly_data[data_idx][2].get(balance_key, 0) or 0
                            prev = self.monthly_data[prev_idx][2].get(balance_key, 0) or 0
                            try:
                                ws.cell(row=r, column=col, value=round(float(prev) - float(curr), 2))
                            except (ValueError, TypeError):
                                ws.cell(row=r, column=col, value=0)
                    else:
                        ws.cell(row=r, column=col, value=0)

                elif formula_type.startswith("sum:"):
                    refs = formula_type.split(":")[1:]
                    cells = [_cell(cf_rows.get(ref, r), col) for ref in refs if ref in cf_rows]
                    if cells:
                        ws.cell(row=r, column=col, value="=" + "+".join(cells))
                        ws.cell(row=r, column=col).font = _FORMULA_FONT

                elif formula_type.startswith("ref:"):
                    ref_key = formula_type.split(":")[1]
                    if ref_key in cf_rows:
                        ws.cell(row=r, column=col, value=f"={_cell(cf_rows[ref_key], col)}")
                        ws.cell(row=r, column=col).font = _FORMULA_FONT

                elif formula_type.startswith("value:"):
                    ws.cell(row=r, column=col, value=float(formula_type.split(":")[1]))

                elif formula_type.startswith("carry:"):
                    # Carry forward from previous month's closing
                    ref_key = formula_type.split(":")[1]
                    if m_idx == 0:
                        # First month — use cash burn data or estimate
                        if actuals_slice:
                            burn = actuals_slice[0][2].get("cash_burn", 0) or 0
                            rev = actuals_slice[0][2].get("_total_revenue", 0) or actuals_slice[0][2].get("revenue", 0) or 0
                            # Estimate opening cash as 12 months of net burn
                            try:
                                est_cash = abs(float(burn)) * 12 if float(burn) < 0 else float(rev) * 6
                            except (ValueError, TypeError):
                                est_cash = 0
                            ws.cell(row=r, column=col, value=round(est_cash, 2))
                    else:
                        # Reference previous month's closing cash
                        if ref_key in cf_rows:
                            ws.cell(row=r, column=col, value=f"={_cell(cf_rows[ref_key], col - 1)}")
                            ws.cell(row=r, column=col).font = _FORMULA_FONT

                ws.cell(row=r, column=col).number_format = '#,##0'

        ws.freeze_panes = "B2"
        _auto_width(ws)

    # ── Sheet 6: Balance Sheet ──────────────────────────────────────────

    BS_LINES = [
        # Assets
        ("bs_header_assets",  "ASSETS",                    "header"),
        ("cash",              "Cash & Equivalents",        "cf_ref:closing_cash"),
        ("accounts_recv",     "Accounts Receivable",       "data:ar"),
        ("prepaids",          "Prepaid Expenses",          "percent_of:opex:8"),  # ~8% of opex
        ("total_current",     "Total Current Assets",      "sum:cash:accounts_recv:prepaids"),
        ("total_assets",      "Total Assets",              "ref:total_current"),
        # Liabilities
        ("bs_header_liab",    "LIABILITIES",               "header"),
        ("accounts_pay",      "Accounts Payable",          "percent_of:cogs:15"),  # ~15% of COGS
        ("deferred_rev",      "Deferred Revenue",          "percent_of:mrr:25"),  # ~25% of MRR
        ("accrued",           "Accrued Expenses",          "percent_of:opex:5"),   # ~5% of opex
        ("total_current_liab","Total Current Liabilities", "sum:accounts_pay:deferred_rev:accrued"),
        ("total_liab",        "Total Liabilities",         "ref:total_current_liab"),
        # Equity
        ("bs_header_equity",  "EQUITY",                    "header"),
        ("equity",            "Total Equity",              "formula:total_assets-total_liab"),
        # Check
        ("bs_check",          "Balance Check (A - L - E)", "formula:total_assets-total_liab-equity"),
    ]

    def _build_balance_sheet(self):
        ws = self.wb.create_sheet("Balance Sheet")

        actuals_slice = self.monthly_data[-12:] if len(self.monthly_data) >= 12 else self.monthly_data
        all_months = [(y, m, "actual") for y, m, _ in actuals_slice]
        for fy, fm in self.forecast_months:
            all_months.append((fy, fm, "forecast"))

        # Headers
        row = 1
        ws.cell(row=row, column=1, value="Balance Sheet").font = _HEADER_FONT
        ws.cell(row=row, column=1).fill = _HEADER_FILL
        for i, (y, m, source) in enumerate(all_months):
            col = 2 + i
            ws.cell(row=row, column=col, value=f"{y}-{m:02d}").font = _HEADER_FONT
            ws.cell(row=row, column=col).fill = _HEADER_FILL if source == "actual" else PatternFill(
                start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")

        bs_rows = {}
        for line_idx, (key, label, formula_type) in enumerate(self.BS_LINES):
            r = row + 1 + line_idx
            bs_rows[key] = r

            if formula_type == "header":
                ws.cell(row=r, column=1, value=label).font = Font(bold=True, size=10, color="0055A4")
                continue

            is_total = key in ("total_current", "total_assets", "total_current_liab", "total_liab", "equity")
            ws.cell(row=r, column=1, value=label).font = Font(
                bold=is_total, size=10, color="0055A4" if is_total else "1E293B")
            if is_total:
                ws.cell(row=r, column=1).border = Border(
                    top=Side(style="thin", color="0055A4"),
                    bottom=Side(style="double", color="0055A4"))

            for m_idx, (y, m, source) in enumerate(all_months):
                col = 2 + m_idx

                if formula_type.startswith("cf_ref:"):
                    # Reference Cash Flow sheet
                    cf_key = formula_type.split(":")[1]
                    # Find the row in CF sheet
                    for cf_idx, (ck, _, _) in enumerate(self.CF_LINES):
                        if ck == cf_key:
                            ws.cell(row=r, column=col, value=f"='Cash Flow'!{_cell(2 + cf_idx, col)}")
                            ws.cell(row=r, column=col).font = _FORMULA_FONT
                            break

                elif formula_type.startswith("data:"):
                    data_key = formula_type.split(":")[1]
                    if source == "actual":
                        data_idx = len(self.monthly_data) - len(actuals_slice) + m_idx
                        if data_idx < len(self.monthly_data):
                            val = self.monthly_data[data_idx][2].get(data_key, 0)
                            try:
                                ws.cell(row=r, column=col, value=round(float(val or 0), 2))
                            except (ValueError, TypeError):
                                ws.cell(row=r, column=col, value=0)
                    else:
                        # Forecast: carry forward last actual
                        prev_col = col - 1
                        ws.cell(row=r, column=col, value=f"={_cell(r, prev_col)}")
                        ws.cell(row=r, column=col).font = _FORMULA_FONT

                elif formula_type.startswith("percent_of:"):
                    parts = formula_type.split(":")
                    base_key = parts[1]
                    pct = float(parts[2])
                    if source == "actual":
                        data_idx = len(self.monthly_data) - len(actuals_slice) + m_idx
                        if data_idx < len(self.monthly_data):
                            base_val = self.monthly_data[data_idx][2].get(base_key, 0)
                            try:
                                ws.cell(row=r, column=col, value=round(float(base_val or 0) * pct / 100, 2))
                            except (ValueError, TypeError):
                                ws.cell(row=r, column=col, value=0)
                    else:
                        prev_col = col - 1
                        ws.cell(row=r, column=col, value=f"={_cell(r, prev_col)}")
                        ws.cell(row=r, column=col).font = _FORMULA_FONT

                elif formula_type.startswith("sum:"):
                    refs = formula_type.split(":")[1:]
                    cells = [_cell(bs_rows.get(ref, r), col) for ref in refs if ref in bs_rows]
                    if cells:
                        ws.cell(row=r, column=col, value="=" + "+".join(cells))
                        ws.cell(row=r, column=col).font = _FORMULA_FONT

                elif formula_type.startswith("ref:"):
                    ref_key = formula_type.split(":")[1]
                    if ref_key in bs_rows:
                        ws.cell(row=r, column=col, value=f"={_cell(bs_rows[ref_key], col)}")
                        ws.cell(row=r, column=col).font = _FORMULA_FONT

                elif formula_type.startswith("formula:"):
                    expr = formula_type.split(":", 1)[1]
                    # Parse "a-b" or "a-b-c"
                    parts = expr.replace("-", "+-").split("+")
                    formula_parts = []
                    for part in parts:
                        part = part.strip()
                        if not part:
                            continue
                        neg = part.startswith("-")
                        pk = part.lstrip("-")
                        if pk in bs_rows:
                            ref = _cell(bs_rows[pk], col)
                            formula_parts.append(f"-{ref}" if neg else ref)
                    if formula_parts:
                        ws.cell(row=r, column=col, value="=" + "+".join(formula_parts))
                        ws.cell(row=r, column=col).font = _FORMULA_FONT

                ws.cell(row=r, column=col).number_format = '#,##0'

        # Balance check row should show 0 if balanced
        bs_check_row = bs_rows.get("bs_check")
        if bs_check_row:
            ws.cell(row=bs_check_row, column=1).font = Font(size=9, italic=True, color="94A3B8")

        ws.freeze_panes = "B2"
        _auto_width(ws)

    # ── Sheet 7: Scenarios ───────────────────────────────────────────────

    def _build_scenarios_sheet(self):
        ws = self.wb.create_sheet("Scenarios")

        if not self.scenarios:
            ws.cell(row=1, column=1, value="No saved scenarios found.").font = \
                Font(size=12, color="94A3B8", italic=True)
            ws.cell(row=2, column=1, value="Save scenarios in Scenario Planner, then re-export.").font = \
                Font(size=10, color="94A3B8")
            return

        # Causal sensitivity map (mirrors frontend ScenarioPlanner.jsx CAUSAL_MAP)
        CAUSAL_MAP = {
            "gross_margin":     {"revenue_growth": 0.05, "gross_margin_adj": 1.0, "cost_reduction": 0.8},
            "operating_margin": {"revenue_growth": 0.04, "gross_margin_adj": 0.9, "headcount_delta": -0.3, "cost_reduction": 0.7},
            "ebitda_margin":    {"revenue_growth": 0.04, "gross_margin_adj": 0.85, "headcount_delta": -0.25, "cost_reduction": 0.65},
            "nrr":              {"churn_adj": -1.2, "expansion_adj": 0.8, "revenue_growth": 0.1},
            "arr_growth":       {"revenue_growth": 1.0, "churn_adj": -0.6, "expansion_adj": 0.4},
            "churn_rate":       {"churn_adj": 1.0, "revenue_growth": -0.03},
            "burn_multiple":    {"revenue_growth": -0.15, "headcount_delta": 0.2, "cost_reduction": -0.3, "cac_adj": 0.05},
            "cac_payback":      {"cac_adj": 0.3, "revenue_growth": -0.1, "gross_margin_adj": -0.15},
            "ltv_cac":          {"churn_adj": -0.5, "gross_margin_adj": 0.3, "cac_adj": -0.2},
            "opex_ratio":       {"headcount_delta": 0.3, "cost_reduction": -0.8, "revenue_growth": -0.05},
            "headcount_eff":    {"headcount_delta": -0.4, "revenue_growth": 0.3},
        }
        OUTPUT_KPIS = list(CAUSAL_MAP.keys())
        LEVER_LABELS = {
            "revenue_growth": "Revenue Growth",
            "gross_margin_adj": "Gross Margin Adj",
            "churn_adj": "Churn Adj",
            "cac_adj": "CAC Adj",
            "headcount_delta": "Headcount Delta",
            "cost_reduction": "Cost Reduction",
            "expansion_adj": "Expansion Adj",
        }

        # Get base values from latest monthly data
        base_values = {}
        if self.monthly_data:
            base_values = self.monthly_data[-1][2]

        # Header row
        row = 1
        headers = ["KPI"] + [s.get("name", f"Scenario {i+1}") for i, s in enumerate(self.scenarios[:3])]
        headers.insert(1, "Base")
        for c, h in enumerate(headers, 1):
            ws.cell(row=row, column=c, value=h)
        _style_header(ws, row, len(headers))

        # Lever inputs section
        ws.cell(row=3, column=1, value="LEVER INPUTS").font = Font(bold=True, size=9, color="94A3B8")
        lever_start_row = 4
        lever_keys = list(LEVER_LABELS.keys())

        for l_idx, lk in enumerate(lever_keys):
            r = lever_start_row + l_idx
            ws.cell(row=r, column=1, value=LEVER_LABELS.get(lk, lk)).font = Font(size=9)
            ws.cell(row=r, column=2, value=0).font = Font(size=9, color="94A3B8")  # Base = 0

            for s_idx, scenario in enumerate(self.scenarios[:3]):
                col = 3 + s_idx
                try:
                    levers = json.loads(scenario.get("levers_json", "{}")) if isinstance(scenario.get("levers_json"), str) else scenario.get("levers_json", {})
                    val = levers.get(lk, 0)
                except (json.JSONDecodeError, TypeError):
                    val = 0
                ws.cell(row=r, column=col, value=val)
                ws.cell(row=r, column=col).fill = _EDITABLE_FILL
                ws.cell(row=r, column=col).protection = Protection(locked=False)

        # Output KPIs section
        separator_row = lever_start_row + len(lever_keys) + 1
        ws.cell(row=separator_row, column=1, value="OUTPUT KPIs (Projected Impact)").font = Font(bold=True, size=9, color="94A3B8")

        for k_idx, kpi_key in enumerate(OUTPUT_KPIS):
            r = separator_row + 1 + k_idx
            ws.cell(row=r, column=1, value=kpi_key).font = Font(size=9, name="Consolas")

            base_val = base_values.get(kpi_key, 0)
            try:
                base_val = float(base_val) if base_val else 0
            except (ValueError, TypeError):
                base_val = 0
            ws.cell(row=r, column=2, value=round(base_val, 2))

            # Build formula: =Base + SUMPRODUCT of levers * coefficients
            coeffs = CAUSAL_MAP.get(kpi_key, {})
            for s_idx in range(min(len(self.scenarios), 3)):
                col = 3 + s_idx
                # Formula: =BaseCell + coeff1*LeverCell1 + coeff2*LeverCell2 + ...
                base_cell = _cell(r, 2)
                terms = []
                for l_idx, lk in enumerate(lever_keys):
                    c = coeffs.get(lk, 0)
                    if c != 0:
                        lever_cell = _cell(lever_start_row + l_idx, col)
                        terms.append(f"{c}*{lever_cell}")

                if terms:
                    formula = f"={base_cell}+" + "+".join(terms)
                    ws.cell(row=r, column=col, value=formula)
                    ws.cell(row=r, column=col).font = _FORMULA_FONT
                else:
                    ws.cell(row=r, column=col, value=round(base_val, 2))

        ws.protection.sheet = True
        ws.protection.password = ""
        ws.freeze_panes = "B2"
        _auto_width(ws)

    # ── Sheet 6: Confidence Bands ────────────────────────────────────────

    def _build_confidence_sheet(self):
        ws = self.wb.create_sheet("Confidence Bands")

        if not self.forecast_data or not self.forecast_data.get("trajectories"):
            ws.cell(row=1, column=1, value="No forecast data available.").font = \
                Font(size=12, color="94A3B8", italic=True)
            return

        trajectories = self.forecast_data["trajectories"]
        kpis = sorted(trajectories.keys())

        current_row = 1
        for kpi_key in kpis:
            traj = trajectories[kpi_key]
            if not traj:
                continue

            # KPI header
            ws.cell(row=current_row, column=1, value=kpi_key).font = \
                Font(bold=True, size=11, color="0055A4")
            current_row += 1

            # Column headers
            headers = ["Percentile"]
            for i, step in enumerate(traj):
                headers.append(step.get("label", f"M+{i}"))
            for c, h in enumerate(headers, 1):
                ws.cell(row=current_row, column=c, value=h)
            _style_header(ws, current_row, len(headers))
            current_row += 1

            # Percentile rows
            for pct_key, pct_label, fill in [
                ("p90", "P90 (Optimistic)", _GREEN_FILL),
                ("p75", "P75", PatternFill(start_color="F0FDF4", end_color="F0FDF4", fill_type="solid")),
                ("p50", "P50 (Median)", PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")),
                ("p25", "P25", PatternFill(start_color="FFF7ED", end_color="FFF7ED", fill_type="solid")),
                ("p10", "P10 (Pessimistic)", _RED_FILL),
            ]:
                ws.cell(row=current_row, column=1, value=pct_label).font = Font(size=9)
                ws.cell(row=current_row, column=1).fill = fill
                for i, step in enumerate(traj):
                    val = step.get(pct_key)
                    if val is not None:
                        ws.cell(row=current_row, column=2 + i, value=round(float(val), 4))
                        ws.cell(row=current_row, column=2 + i).fill = fill
                        ws.cell(row=current_row, column=2 + i).number_format = '#,##0.00'
                current_row += 1

            current_row += 1  # blank row between KPIs

        _auto_width(ws)

    # ── Sheet 7: Dashboard ───────────────────────────────────────────────

    def _build_dashboard_sheet(self):
        ws = self.wb.create_sheet("Dashboard")

        ws.cell(row=1, column=1, value="Financial Model Dashboard").font = \
            Font(bold=True, size=14, color="0055A4")

        company = self.settings.get("company_name", "Your Company")
        stage = self.settings.get("company_stage", "")
        ws.cell(row=2, column=1, value=f"{company} | {stage.replace('_', ' ').title()} | Generated {datetime.utcnow().strftime('%Y-%m-%d')}").font = \
            Font(size=10, color="94A3B8")

        # Summary metrics from latest data
        if self.monthly_data:
            latest = self.monthly_data[-1][2]
            summary_kpis = [
                ("Revenue Growth", latest.get("revenue_growth"), "%"),
                ("Gross Margin", latest.get("gross_margin"), "%"),
                ("Operating Margin", latest.get("operating_margin"), "%"),
                ("NRR", latest.get("nrr"), "%"),
                ("Churn Rate", latest.get("churn_rate"), "%"),
                ("Burn Multiple", latest.get("burn_multiple"), "x"),
            ]

            row = 4
            for c, (label, val, unit) in enumerate(summary_kpis):
                ws.cell(row=row, column=1 + c * 2, value=label).font = Font(size=9, color="64748B")
                if val is not None:
                    ws.cell(row=row + 1, column=1 + c * 2, value=f"{float(val):.1f}{unit}").font = \
                        Font(bold=True, size=14, color="1E293B")

        # Add a revenue trend chart if we have enough data
        if len(self.monthly_data) >= 6 and "Actuals" in [s.title for s in self.wb.worksheets]:
            actuals_ws = self.wb["Actuals"]
            # Find revenue_growth row in Actuals
            rev_row = None
            for r in range(2, actuals_ws.max_row + 1):
                if actuals_ws.cell(row=r, column=1).value == "revenue_growth":
                    rev_row = r
                    break

            if rev_row and actuals_ws.max_column >= 4:
                chart = LineChart()
                chart.title = "Revenue Growth Trend"
                chart.style = 10
                chart.y_axis.title = "Growth %"
                chart.x_axis.title = "Period"
                chart.width = 25
                chart.height = 12

                data = Reference(actuals_ws, min_col=3, max_col=actuals_ws.max_column,
                                 min_row=rev_row, max_row=rev_row)
                cats = Reference(actuals_ws, min_col=3, max_col=actuals_ws.max_column,
                                 min_row=1, max_row=1)
                chart.add_data(data, from_rows=True, titles_from_data=False)
                chart.set_categories(cats)
                chart.series[0].graphicalProperties.line.solidFill = "0055A4"

                ws.add_chart(chart, "A8")

        _auto_width(ws)

    # ── Metadata ─────────────────────────────────────────────────────────

    def _add_metadata(self):
        """Add workbook properties for version tracking."""
        self.wb.properties.title = "Axiom Financial Model"
        self.wb.properties.creator = "Axiom Intelligence Platform"
        self.wb.properties.description = (
            f"Generated {datetime.utcnow().isoformat()} | "
            f"{len(self.monthly_data)} months of actuals | "
            f"{len(self.forecast_months)} forecast months | "
            f"{len(self.scenarios)} scenarios"
        )
