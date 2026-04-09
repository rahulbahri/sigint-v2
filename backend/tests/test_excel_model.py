"""
tests/test_excel_model.py -- Tests for the financial model workbook generator.
"""
import json
import pytest
from openpyxl import load_workbook
from io import BytesIO

from core.excel_model import CellMap, ModelWorkbookBuilder, ASSUMPTION_PARAMS, PL_LINES


# ── Test Data Fixtures ───────────────────────────────────────────────────────

def _sample_monthly(n_months=12, start_year=2024, start_month=1):
    """Generate n months of realistic sample data."""
    data = []
    y, m = start_year, start_month
    for i in range(n_months):
        rev = 100000 + i * 5000
        cogs = rev * 0.38
        opex = rev * 0.30
        kpis = {
            "revenue": rev,
            "cogs": cogs,
            "opex": opex,
            "revenue_growth": 5.0 + i * 0.1,
            "gross_margin": 62.0 + i * 0.2,
            "operating_margin": 32.0 + i * 0.15,
            "ebitda_margin": 36.8 + i * 0.15,
            "nrr": 105.0,
            "churn_rate": 2.5 - i * 0.05,
            "burn_multiple": 1.2 - i * 0.02,
            "arr_growth": 6.0 + i * 0.1,
            "headcount": 50 + i,
            "customers": 100 + i * 3,
            "dso": 35.0,
        }
        data.append((y, m, kpis))
        m += 1
        if m > 12:
            m = 1
            y += 1
    return data


def _sample_forecast(kpis=None):
    """Generate sample forecast trajectories."""
    if kpis is None:
        kpis = ["revenue_growth", "gross_margin", "operating_margin", "nrr", "churn_rate"]
    trajectories = {}
    for kpi in kpis:
        steps = []
        base = 60.0 if "margin" in kpi else 5.0
        for s in range(7):  # 6 forecast months + current
            steps.append({
                "step": s,
                "label": "Now" if s == 0 else f"M+{s}",
                "p10": round(base - 5 + s * 0.3, 2),
                "p25": round(base - 2 + s * 0.3, 2),
                "p50": round(base + s * 0.3, 2),
                "p75": round(base + 2 + s * 0.3, 2),
                "p90": round(base + 5 + s * 0.3, 2),
            })
        trajectories[kpi] = steps
    return {"trajectories": trajectories, "kpis": kpis}


def _sample_scenarios():
    return [
        {"name": "Base Case", "levers_json": json.dumps({"revenue_growth": 0, "churn_adj": 0})},
        {"name": "Optimistic", "levers_json": json.dumps({"revenue_growth": 5, "churn_adj": -2, "expansion_adj": 3})},
        {"name": "Conservative", "levers_json": json.dumps({"revenue_growth": -3, "churn_adj": 1, "cost_reduction": 5})},
    ]


def _sample_settings():
    return {"company_name": "TestCorp", "company_stage": "series_b", "model_window_months": 48}


def _sample_targets():
    return {
        "revenue_growth": {"target_value": 6.0, "unit": "pct", "direction": "higher"},
        "gross_margin": {"target_value": 65.0, "unit": "pct", "direction": "higher"},
    }


def _build_and_load(**overrides):
    """Build the workbook and load it back for inspection."""
    monthly = overrides.get("monthly_data", _sample_monthly())
    forecast = overrides.get("forecast_data", _sample_forecast())
    scenarios = overrides.get("scenarios", _sample_scenarios())
    settings = overrides.get("settings", _sample_settings())
    targets = overrides.get("targets", _sample_targets())

    builder = ModelWorkbookBuilder(monthly, forecast, scenarios, settings, targets)
    wb = builder.build()

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return load_workbook(buf, data_only=False)


# ── CellMap Tests ────────────────────────────────────────────────────────────

def test_cell_map_register_and_resolve():
    cm = CellMap()
    cm.register("Actuals", "revenue_growth", 5, 3, month_idx=0)
    ref = cm.resolve("revenue_growth", "Actuals", month_idx=0)
    assert ref is not None
    assert "Actuals" in ref
    assert "$C$5" in ref


def test_cell_map_resolve_missing():
    cm = CellMap()
    assert cm.resolve("nonexistent") is None


def test_cell_map_cross_sheet():
    cm = CellMap()
    cm.register("Assumptions", "revenue", 5, 4, month_idx=0)
    cm.register("Actuals", "revenue", 3, 5, month_idx=0)
    # Resolve from specific sheet
    ref_a = cm.resolve("revenue", "Assumptions", month_idx=0)
    ref_b = cm.resolve("revenue", "Actuals", month_idx=0)
    assert ref_a != ref_b
    assert "Assumptions" in ref_a
    assert "Actuals" in ref_b


# ── Workbook Structure Tests ─────────────────────────────────────────────────

def test_workbook_has_nine_sheets():
    wb = _build_and_load()
    assert len(wb.sheetnames) == 9
    expected = ["Assumptions", "Actuals", "Forecast", "P&L", "Cash Flow", "Balance Sheet", "Scenarios", "Confidence Bands", "Dashboard"]
    assert wb.sheetnames == expected


def test_workbook_no_forecast():
    """Without forecast data, Forecast and Confidence sheets show placeholder."""
    wb = _build_and_load(forecast_data=None)
    ws = wb["Forecast"]
    assert "No forecast model" in str(ws.cell(row=1, column=1).value)


def test_workbook_no_scenarios():
    """Without scenarios, Scenarios sheet shows placeholder."""
    wb = _build_and_load(scenarios=[])
    ws = wb["Scenarios"]
    assert "No saved scenarios" in str(ws.cell(row=1, column=1).value)


# ── Assumptions Sheet Tests ──────────────────────────────────────────────────

def test_assumptions_sheet_has_parameters():
    wb = _build_and_load()
    ws = wb["Assumptions"]
    # Row 4 is header, Row 5+ are parameters
    param_labels = []
    for r in range(5, 5 + len(ASSUMPTION_PARAMS)):
        val = ws.cell(row=r, column=1).value
        if val:
            param_labels.append(val)
    assert "Revenue" in param_labels
    assert "COGS" in param_labels
    assert "Headcount" in param_labels


def test_assumptions_cells_have_values():
    wb = _build_and_load()
    ws = wb["Assumptions"]
    # Revenue row (first param) should have monthly values
    # Row 5, columns 4+ should have data from the sample
    rev_row = 5
    val = ws.cell(row=rev_row, column=4).value
    assert val is not None  # Should have the first month's revenue


# ── Actuals Sheet Tests ──────────────────────────────────────────────────────

def test_actuals_sheet_has_kpis():
    wb = _build_and_load()
    ws = wb["Actuals"]
    # Row 1 is header, Row 2+ are KPIs
    kpi_keys = []
    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val:
            kpi_keys.append(val)
    assert "revenue_growth" in kpi_keys
    assert "gross_margin" in kpi_keys


def test_actuals_has_monthly_data():
    wb = _build_and_load()
    ws = wb["Actuals"]
    # Should have 12 month columns (col 3 through 14)
    header_val = ws.cell(row=1, column=3).value
    assert header_val is not None
    assert "2024" in str(header_val)


# ── P&L Sheet Tests ─────────────────────────────────────────────────────────

def test_pl_has_line_items():
    wb = _build_and_load()
    ws = wb["P&L"]
    items = []
    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val:
            items.append(val)
    assert "Revenue" in items
    assert "Gross Profit" in items
    assert "Operating Income" in items


def test_pl_gross_profit_is_formula():
    wb = _build_and_load()
    ws = wb["P&L"]
    # Find Gross Profit row
    gp_row = None
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == "Gross Profit":
            gp_row = r
            break
    assert gp_row is not None
    # First data column should be a formula
    val = ws.cell(row=gp_row, column=2).value
    assert val is not None
    assert str(val).startswith("=")


# ── Scenarios Sheet Tests ────────────────────────────────────────────────────

def test_scenarios_has_three_columns():
    wb = _build_and_load()
    ws = wb["Scenarios"]
    # Row 1 headers should include scenario names
    h2 = ws.cell(row=1, column=2).value
    h3 = ws.cell(row=1, column=3).value
    assert h2 == "Base"
    assert h3 is not None  # First scenario name


def test_scenarios_output_has_formulas():
    wb = _build_and_load()
    ws = wb["Scenarios"]
    # Find "OUTPUT KPIs" row
    output_row = None
    for r in range(1, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val and "OUTPUT" in str(val):
            output_row = r
            break
    assert output_row is not None
    # First output KPI, first scenario column should be a formula
    formula_cell = ws.cell(row=output_row + 1, column=3).value
    assert formula_cell is not None
    assert str(formula_cell).startswith("=")


# ── Confidence Bands Tests ───────────────────────────────────────────────────

def test_confidence_has_percentiles():
    wb = _build_and_load()
    ws = wb["Confidence Bands"]
    # Should have "P50 (Median)" somewhere
    found = False
    for r in range(1, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val and "P50" in str(val):
            found = True
            break
    assert found


def test_confidence_percentile_ordering():
    wb = _build_and_load()
    ws = wb["Confidence Bands"]
    # Find first KPI's percentile block
    # Row 2 is header, rows 3-7 are percentiles
    p90 = ws.cell(row=3, column=2).value
    p50 = ws.cell(row=5, column=2).value
    p10 = ws.cell(row=7, column=2).value
    if p90 is not None and p50 is not None and p10 is not None:
        assert float(p10) <= float(p50) <= float(p90)


# ── Dashboard Tests ──────────────────────────────────────────────────────────

def test_dashboard_has_title():
    wb = _build_and_load()
    ws = wb["Dashboard"]
    assert "Financial Model Dashboard" in str(ws.cell(row=1, column=1).value)


def test_dashboard_has_company_name():
    wb = _build_and_load()
    ws = wb["Dashboard"]
    assert "TestCorp" in str(ws.cell(row=2, column=1).value)


# ── Round-trip Test ──────────────────────────────────────────────────────────

def test_full_workbook_round_trip():
    """Build a workbook, save to bytes, reload, verify it is valid."""
    monthly = _sample_monthly(36, start_year=2022)
    forecast = _sample_forecast()
    scenarios = _sample_scenarios()
    settings = _sample_settings()
    targets = _sample_targets()

    builder = ModelWorkbookBuilder(monthly, forecast, scenarios, settings, targets)
    wb = builder.build()

    buf = BytesIO()
    wb.save(buf)
    assert buf.tell() > 0  # Non-empty file

    buf.seek(0)
    wb2 = load_workbook(buf)
    assert len(wb2.sheetnames) == 9

    # Verify metadata
    assert wb2.properties.title == "Axiom Financial Model"
    assert "36 months" in wb2.properties.description
