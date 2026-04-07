"""
tests/test_export_model.py -- Tests for the financial model export endpoint.
"""
import io
import json
import pytest
from httpx import AsyncClient
from openpyxl import load_workbook

from core.database import get_db


# ── Helpers ──────────────────────────────────────────────────────────────────

WS = "testcorp.com"


def _seed_monthly(workspace_id: str, n_months: int = 12):
    conn = get_db()
    y, m = 2024, 1
    for i in range(n_months):
        kpis = {
            "revenue": 100000 + i * 5000,
            "cogs": 38000 + i * 1000,
            "opex": 30000 + i * 500,
            "revenue_growth": 5.0 + i * 0.1,
            "gross_margin": 62.0 + i * 0.2,
            "operating_margin": 32.0,
            "nrr": 105.0,
            "churn_rate": 2.5,
            "customers": 100 + i * 3,
            "headcount": 50 + i,
        }
        conn.execute(
            "INSERT INTO monthly_data (upload_id, year, month, data_json, workspace_id) VALUES (?,?,?,?,?)",
            (1, y, m, json.dumps(kpis), workspace_id),
        )
        m += 1
        if m > 12:
            m = 1
            y += 1
    conn.commit()
    conn.close()


def _clear_data(workspace_id: str):
    conn = get_db()
    conn.execute("DELETE FROM monthly_data WHERE workspace_id=?", [workspace_id])
    try:
        conn.execute("DELETE FROM model_exports WHERE workspace_id=?", [workspace_id])
    except Exception:
        pass
    conn.commit()
    conn.close()


# ── Tests ────────────────────────────────────────────────────────────────────

@pytest.mark.anyio
async def test_export_requires_auth(client: AsyncClient):
    """GET /api/export/financial-model.xlsx without auth returns 401."""
    resp = await client.get("/api/export/financial-model.xlsx")
    assert resp.status_code == 401


@pytest.mark.anyio
async def test_export_empty_workspace(client: AsyncClient, auth_headers):
    """Export with no data returns a valid but minimal XLSX."""
    _clear_data(WS)
    resp = await client.get("/api/export/financial-model.xlsx", headers=auth_headers)
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers.get("content-type", "")

    # Should be a valid XLSX
    wb = load_workbook(io.BytesIO(resp.content))
    assert "Assumptions" in wb.sheetnames
    assert len(wb.sheetnames) == 7


@pytest.mark.anyio
async def test_export_with_data(client: AsyncClient, auth_headers):
    """Export with seeded data produces a valid XLSX with populated sheets."""
    _clear_data(WS)
    _seed_monthly(WS, 12)

    resp = await client.get("/api/export/financial-model.xlsx", headers=auth_headers)
    assert resp.status_code == 200

    wb = load_workbook(io.BytesIO(resp.content), data_only=False)
    assert len(wb.sheetnames) == 7

    # Actuals sheet should have data rows
    ws = wb["Actuals"]
    assert ws.max_row > 2  # Header + at least some KPI rows

    # P&L should have formulas
    ws_pl = wb["P&L"]
    found_formula = False
    for r in range(2, ws_pl.max_row + 1):
        for c in range(2, ws_pl.max_column + 1):
            val = ws_pl.cell(row=r, column=c).value
            if val and str(val).startswith("="):
                found_formula = True
                break
    assert found_formula, "P&L sheet should contain at least one formula"

    _clear_data(WS)


@pytest.mark.anyio
async def test_export_creates_version_record(client: AsyncClient, auth_headers):
    """Each export creates a row in model_exports."""
    _clear_data(WS)
    _seed_monthly(WS, 6)

    await client.get("/api/export/financial-model.xlsx", headers=auth_headers)

    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM model_exports WHERE workspace_id=?", [WS]
    ).fetchall()
    conn.close()

    assert len(rows) >= 1
    assert rows[-1]["months_of_data"] == 6
    assert "axiom-financial-model" in rows[-1]["filename"]

    _clear_data(WS)


@pytest.mark.anyio
async def test_export_populates_assumption_snapshot(client: AsyncClient, auth_headers):
    """Export stores assumption values in assumption_snapshot for Phase 3 diff."""
    _clear_data(WS)
    _seed_monthly(WS, 6)

    await client.get("/api/export/financial-model.xlsx", headers=auth_headers)

    conn = get_db()
    rows = conn.execute(
        "SELECT assumption_snapshot FROM model_exports WHERE workspace_id=? ORDER BY id DESC LIMIT 1",
        [WS]
    ).fetchall()
    conn.close()

    assert len(rows) >= 1
    snapshot_raw = rows[0]["assumption_snapshot"]
    assert snapshot_raw is not None, "assumption_snapshot should be populated"

    snapshot = json.loads(snapshot_raw)
    assert isinstance(snapshot, dict)
    assert len(snapshot) > 0
    # Should contain key KPIs from the seeded data
    assert "revenue_growth" in snapshot or "gross_margin" in snapshot

    _clear_data(WS)
