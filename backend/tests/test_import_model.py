"""
tests/test_import_model.py -- Tests for the financial model import + diff engine.
"""
import io
import json
import pytest
from httpx import AsyncClient
from openpyxl import Workbook

from core.database import get_db


WS = "testcorp.com"


def _seed_monthly(workspace_id: str, n_months: int = 6):
    conn = get_db()
    y, m = 2024, 1
    for i in range(n_months):
        kpis = {
            "revenue": 100000 + i * 5000,
            "cogs": 38000 + i * 1000,
            "opex": 30000 + i * 500,
            "revenue_growth": 5.0 + i * 0.1,
            "gross_margin": 62.0 + i * 0.2,
            "churn_rate": 2.5,
            "nrr": 105.0,
            "headcount": 50 + i,
            "customers": 100 + i * 3,
        }
        conn.execute(
            "INSERT INTO monthly_data (upload_id, year, month, data_json, workspace_id) VALUES (?,?,?,?,?)",
            (1, y, m, json.dumps(kpis), workspace_id),
        )
        m += 1
        if m > 12:
            m = 1
            y += 1
    conn.commit()
    conn.close()


def _seed_export_snapshot(workspace_id: str, snapshot: dict):
    """Manually insert a model_exports row with a known assumption snapshot."""
    conn = get_db()
    conn.execute(
        "INSERT INTO model_exports (version, filename, assumption_snapshot, months_of_data, workspace_id) "
        "VALUES (?,?,?,?,?)",
        ("20260407120000", "test-export.xlsx", json.dumps(snapshot), 6, workspace_id),
    )
    conn.commit()
    conn.close()


def _clear(workspace_id: str):
    conn = get_db()
    conn.execute("DELETE FROM monthly_data WHERE workspace_id=?", [workspace_id])
    try:
        conn.execute("DELETE FROM model_exports WHERE workspace_id=?", [workspace_id])
    except Exception:
        pass
    try:
        conn.execute("DELETE FROM saved_scenarios WHERE workspace_id=?", [workspace_id])
    except Exception:
        pass
    conn.commit()
    conn.close()


def _make_assumptions_xlsx(values: dict) -> bytes:
    """Create an .xlsx with an Assumptions sheet matching our export format."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Assumptions"
    # Header
    ws.cell(row=1, column=1, value="Assumptions -- Edit These Values")
    ws.cell(row=4, column=1, value="Parameter")
    ws.cell(row=4, column=2, value="Unit")
    ws.cell(row=4, column=3, value="Description")
    ws.cell(row=4, column=4, value="2024-06")

    from core.excel_model import ASSUMPTION_PARAMS
    for i, (key, label, unit, desc, _) in enumerate(ASSUMPTION_PARAMS):
        r = 5 + i
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=unit)
        ws.cell(row=r, column=3, value=desc)
        if key in values:
            ws.cell(row=r, column=4, value=values[key])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ── Import Tests ─────────────────────────────────────────────────────────────

@pytest.mark.anyio
async def test_import_requires_auth(client: AsyncClient):
    """POST without auth returns 401."""
    xlsx = _make_assumptions_xlsx({"revenue_growth": 5.0})
    files = {"file": ("model.xlsx", io.BytesIO(xlsx),
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    resp = await client.post("/api/import/financial-model", files=files)
    assert resp.status_code == 401


@pytest.mark.anyio
async def test_import_bad_file(client: AsyncClient, auth_headers):
    """Non-XLSX file returns 400."""
    files = {"file": ("model.csv", io.BytesIO(b"a,b,c"), "text/csv")}
    resp = await client.post("/api/import/financial-model", files=files, headers=auth_headers)
    assert resp.status_code == 400


@pytest.mark.anyio
async def test_import_no_assumptions_sheet(client: AsyncClient, auth_headers):
    """XLSX without Assumptions sheet returns 400."""
    wb = Workbook()
    wb.active.title = "WrongSheet"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    files = {"file": ("model.xlsx", buf,
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    resp = await client.post("/api/import/financial-model", files=files, headers=auth_headers)
    assert resp.status_code == 400
    assert "Assumptions" in resp.json()["detail"]


@pytest.mark.anyio
async def test_import_no_baseline(client: AsyncClient, auth_headers):
    """Import without a prior export returns no_baseline status."""
    _clear(WS)
    # Use ASSUMPTION_PARAMS keys: "revenue" (raw), "revenue_growth_rate" (mapped)
    xlsx = _make_assumptions_xlsx({"revenue": 120000, "cogs": 45000, "opex": 35000})
    files = {"file": ("model.xlsx", io.BytesIO(xlsx),
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    resp = await client.post("/api/import/financial-model", files=files, headers=auth_headers)
    assert resp.status_code == 200
    data = resp.json()
    assert data["status"] == "no_baseline"
    assert len(data["imported_values"]) > 0
    _clear(WS)


@pytest.mark.anyio
async def test_import_no_changes(client: AsyncClient, auth_headers):
    """Import matching the baseline shows zero changes."""
    _clear(WS)
    snapshot = {"revenue_growth": 5.0, "gross_margin": 62.0}
    _seed_export_snapshot(WS, snapshot)

    xlsx = _make_assumptions_xlsx({"revenue_growth": 5.0, "gross_margin": 62.0})
    files = {"file": ("model.xlsx", io.BytesIO(xlsx),
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    resp = await client.post("/api/import/financial-model", files=files, headers=auth_headers)
    assert resp.status_code == 200
    data = resp.json()
    assert data["status"] == "diff_computed"
    assert len(data["changes"]) == 0
    _clear(WS)


@pytest.mark.anyio
async def test_import_detects_changes(client: AsyncClient, auth_headers):
    """Import with modified assumptions detects the delta."""
    _clear(WS)
    # Snapshot uses KPI keys (as stored from monthly_data)
    snapshot = {"revenue": 100000, "cogs": 38000, "opex": 30000, "revenue_growth": 5.0, "gross_margin": 62.0}
    _seed_export_snapshot(WS, snapshot)

    # User changed revenue from 100000 to 120000 and opex from 30000 to 25000
    # (Using ASSUMPTION_PARAMS keys which match directly for raw values)
    xlsx = _make_assumptions_xlsx({
        "revenue": 120000,       # changed +20000
        "cogs": 38000,           # unchanged
        "opex": 25000,           # changed -5000
        "revenue_growth_rate": 5.0,  # maps to "revenue_growth" in snapshot — unchanged
    })
    files = {"file": ("model.xlsx", io.BytesIO(xlsx),
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    resp = await client.post("/api/import/financial-model", files=files, headers=auth_headers)
    assert resp.status_code == 200
    data = resp.json()
    assert data["status"] == "diff_computed"
    assert len(data["changes"]) == 2

    # Verify specific changes
    change_map = {c["param"]: c for c in data["changes"]}
    assert "revenue" in change_map
    assert change_map["revenue"]["delta"] == pytest.approx(20000.0, abs=1)
    assert "opex" in change_map
    assert change_map["opex"]["delta"] == pytest.approx(-5000.0, abs=1)

    # Verify scenario mapping
    assert "cost_reduction" in data["scenario_mapping"]  # opex maps to cost_reduction lever

    _clear(WS)


# ── Apply Tests ──────────────────────────────────────────────────────────────

@pytest.mark.anyio
async def test_apply_creates_scenario(client: AsyncClient, auth_headers):
    """Apply creates a saved scenario from the diff."""
    _clear(WS)
    resp = await client.post(
        "/api/import/financial-model/apply",
        headers=auth_headers,
        json={
            "scenario_mapping": {"revenue_growth": 3.0, "churn_adj": -1.0},
            "scenario_name": "Excel Import Test",
        },
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data["status"] == "applied"
    assert data["scenario_id"] is not None

    # Verify scenario was created
    conn = get_db()
    rows = conn.execute(
        "SELECT name, levers_json FROM saved_scenarios WHERE workspace_id=? ORDER BY id DESC LIMIT 1",
        [WS],
    ).fetchall()
    conn.close()
    assert len(rows) >= 1
    assert rows[0]["name"] == "Excel Import Test"
    levers = json.loads(rows[0]["levers_json"])
    assert levers["revenue_growth"] == 3.0

    _clear(WS)


@pytest.mark.anyio
async def test_apply_empty_mapping(client: AsyncClient, auth_headers):
    """Apply with empty mapping returns 400."""
    resp = await client.post(
        "/api/import/financial-model/apply",
        headers=auth_headers,
        json={"scenario_mapping": {}, "scenario_name": "Empty"},
    )
    assert resp.status_code == 400
