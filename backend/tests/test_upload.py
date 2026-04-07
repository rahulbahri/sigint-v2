"""
tests/test_upload.py — CSV/XLSX upload and workspace isolation tests.
"""
import io
import pytest
from httpx import AsyncClient
from openpyxl import Workbook as _Workbook

CSV_CONTENT = b"""date,revenue,cogs,opex,customers
2025-01-01,100000,40000,30000,50
2025-02-01,110000,42000,31000,55
2025-03-01,120000,44000,32000,60
"""


@pytest.mark.anyio
async def test_upload_no_auth_behavior(client: AsyncClient):
    """
    POST /api/upload without auth — _get_workspace returns '' so workspace_id is ''.
    The upload itself may succeed (200) with an empty workspace_id (M7 bug),
    OR the endpoint may return 200 regardless.  We just assert it does NOT return 401.
    """
    files = {"file": ("test.csv", io.BytesIO(CSV_CONTENT), "text/csv")}
    resp = await client.post("/api/upload", files=files)
    # The endpoint uses _get_workspace (not _require_workspace) so it will succeed
    # with workspace_id="" — test documents the current (unguarded) behaviour.
    assert resp.status_code == 200


@pytest.mark.anyio
async def test_upload_with_auth(client: AsyncClient, auth_headers):
    """POST /api/upload with auth + valid CSV → 200, returns upload_id."""
    files = {"file": ("test_alice.csv", io.BytesIO(CSV_CONTENT), "text/csv")}
    resp = await client.post("/api/upload", files=files, headers=auth_headers)
    assert resp.status_code == 200
    data = resp.json()
    assert "upload_id" in data
    assert data["upload_id"] is not None
    assert data["rows_processed"] == 3
    assert data["months_detected"] == 3


@pytest.mark.anyio
async def test_list_uploads_with_auth(client: AsyncClient, auth_headers):
    """GET /api/uploads with auth → 200, returns alice's uploads (at least one)."""
    # Seed an upload first
    files = {"file": ("list_test.csv", io.BytesIO(CSV_CONTENT), "text/csv")}
    await client.post("/api/upload", files=files, headers=auth_headers)

    resp = await client.get("/api/uploads", headers=auth_headers)
    assert resp.status_code == 200
    data = resp.json()
    assert isinstance(data, list)
    assert len(data) >= 1
    # Every returned entry must have the expected shape
    for entry in data:
        assert "id" in entry
        assert "filename" in entry


@pytest.mark.anyio
async def test_workspace_isolation(client: AsyncClient, auth_headers, other_auth_headers):
    """GET /api/uploads for bob should not include alice's uploads."""
    # Ensure alice has at least one upload
    files = {"file": ("isolation_test.csv", io.BytesIO(CSV_CONTENT), "text/csv")}
    alice_upload = await client.post("/api/upload", files=files, headers=auth_headers)
    assert alice_upload.status_code == 200

    # Bob checks his uploads — his workspace is completely separate
    bob_resp = await client.get("/api/uploads", headers=other_auth_headers)
    assert bob_resp.status_code == 200
    bob_data = bob_resp.json()

    alice_id = alice_upload.json()["upload_id"]
    bob_ids = [entry["id"] for entry in bob_data]
    assert alice_id not in bob_ids, "Bob must NOT see Alice's upload"


@pytest.mark.anyio
async def test_delete_upload_no_auth(client: AsyncClient, auth_headers):
    """DELETE /api/uploads/{id} without auth → 401."""
    # Create an upload as alice first
    files = {"file": ("to_delete.csv", io.BytesIO(CSV_CONTENT), "text/csv")}
    create_resp = await client.post("/api/upload", files=files, headers=auth_headers)
    upload_id = create_resp.json()["upload_id"]

    # Attempt deletion with no auth — _get_workspace returns '' which triggers 401
    resp = await client.delete(f"/api/uploads/{upload_id}")
    assert resp.status_code == 401


@pytest.mark.anyio
async def test_delete_upload_wrong_workspace(client: AsyncClient, auth_headers, other_auth_headers):
    """DELETE /api/uploads/{id} with wrong workspace → 404."""
    # Alice creates an upload
    files = {"file": ("wrong_ws.csv", io.BytesIO(CSV_CONTENT), "text/csv")}
    create_resp = await client.post("/api/upload", files=files, headers=auth_headers)
    upload_id = create_resp.json()["upload_id"]

    # Bob tries to delete Alice's upload — should get 404 (not found in his workspace)
    resp = await client.delete(f"/api/uploads/{upload_id}", headers=other_auth_headers)
    assert resp.status_code == 404


# ── XLSX Upload Tests ────────────────────────────────────────────────────────

def _make_xlsx_bytes(rows: list, headers: list) -> bytes:
    """Create an in-memory .xlsx file and return its bytes."""
    wb = _Workbook()
    ws = wb.active
    ws.append(headers)
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@pytest.mark.anyio
async def test_xlsx_upload_single_sheet(client: AsyncClient, auth_headers):
    """POST /api/upload with a valid .xlsx file parses correctly."""
    xlsx_bytes = _make_xlsx_bytes(
        rows=[
            ["2025-01-01", 100000, 40000, 30000, 50],
            ["2025-02-01", 110000, 42000, 31000, 55],
            ["2025-03-01", 120000, 44000, 32000, 60],
        ],
        headers=["date", "revenue", "cogs", "opex", "customers"],
    )
    files = {"file": ("test_data.xlsx", io.BytesIO(xlsx_bytes),
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    resp = await client.post("/api/upload", files=files, headers=auth_headers)
    assert resp.status_code == 200
    data = resp.json()
    assert data["rows_processed"] == 3
    assert data["months_detected"] == 3


@pytest.mark.anyio
async def test_xlsx_upload_bad_file(client: AsyncClient, auth_headers):
    """Uploading garbage bytes as .xlsx returns 400."""
    files = {"file": ("bad.xlsx", io.BytesIO(b"not a real xlsx file"),
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    resp = await client.post("/api/upload", files=files, headers=auth_headers)
    assert resp.status_code == 400


def _make_multi_sheet_xlsx(sheets: dict) -> bytes:
    """Create multi-sheet .xlsx. sheets = {name: (headers, rows)}."""
    wb = _Workbook()
    first = True
    for name, (headers, rows) in sheets.items():
        ws = wb.active if first else wb.create_sheet(title=name)
        if first:
            ws.title = name
            first = False
        ws.append(headers)
        for row in rows:
            ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@pytest.mark.anyio
async def test_xlsx_multi_sheet_yearly(client: AsyncClient, auth_headers):
    """Multi-sheet workbook with year-named sheets concatenates correctly."""
    xlsx_bytes = _make_multi_sheet_xlsx({
        "2024": (
            ["date", "revenue", "cogs", "opex", "customers"],
            [
                ["2024-01-01", 100000, 40000, 30000, 50],
                ["2024-02-01", 110000, 42000, 31000, 55],
            ],
        ),
        "2025": (
            ["date", "revenue", "cogs", "opex", "customers"],
            [
                ["2025-01-01", 130000, 46000, 33000, 65],
                ["2025-02-01", 140000, 48000, 34000, 70],
            ],
        ),
    })
    files = {"file": ("multi_year.xlsx", io.BytesIO(xlsx_bytes),
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    resp = await client.post("/api/upload", files=files, headers=auth_headers)
    assert resp.status_code == 200
    data = resp.json()
    # 4 rows across 2 sheets = 4 months (Jan 2024, Feb 2024, Jan 2025, Feb 2025)
    assert data["months_detected"] == 4


@pytest.mark.anyio
async def test_xlsx_multi_sheet_year_no_date_column(client: AsyncClient, auth_headers):
    """Year-named sheets without a date column get dates synthesised from sheet name."""
    xlsx_bytes = _make_multi_sheet_xlsx({
        "2024": (
            ["revenue", "cogs", "opex", "customers"],
            [
                [100000, 40000, 30000, 50],
                [110000, 42000, 31000, 55],
                [120000, 44000, 32000, 60],
            ],
        ),
        "2025": (
            ["revenue", "cogs", "opex", "customers"],
            [
                [130000, 46000, 33000, 65],
                [140000, 48000, 34000, 70],
            ],
        ),
    })
    files = {"file": ("no_dates.xlsx", io.BytesIO(xlsx_bytes),
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    resp = await client.post("/api/upload", files=files, headers=auth_headers)
    assert resp.status_code == 200
    data = resp.json()
    # 3 rows in 2024 (Jan-Mar) + 2 rows in 2025 (Jan-Feb) = 5 months
    assert data["months_detected"] == 5


@pytest.mark.anyio
async def test_xlsx_multi_sheet_unknown_names(client: AsyncClient, auth_headers):
    """Unrecognised sheet names fall back to first sheet only."""
    xlsx_bytes = _make_multi_sheet_xlsx({
        "Summary": (
            ["date", "revenue", "cogs", "opex", "customers"],
            [
                ["2025-01-01", 100000, 40000, 30000, 50],
                ["2025-02-01", 110000, 42000, 31000, 55],
            ],
        ),
        "Notes": (
            ["comment"],
            [["This is a note"]],
        ),
    })
    files = {"file": ("unknown_sheets.xlsx", io.BytesIO(xlsx_bytes),
                       "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
    resp = await client.post("/api/upload", files=files, headers=auth_headers)
    assert resp.status_code == 200
    data = resp.json()
    assert data["months_detected"] == 2


@pytest.mark.anyio
async def test_csv_upload_regression(client: AsyncClient, auth_headers):
    """CSV upload still works after XLSX changes (regression test)."""
    files = {"file": ("regression.csv", io.BytesIO(CSV_CONTENT), "text/csv")}
    resp = await client.post("/api/upload", files=files, headers=auth_headers)
    assert resp.status_code == 200
    assert resp.json()["months_detected"] == 3
